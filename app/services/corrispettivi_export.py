# app/services/corrispettivi_export.py
# @version: v1.0
# Esportazione corrispettivi da DB → Excel e generazione template
# I campi canonici sono definiti qui e sono la fonte di verità

import io
import sqlite3
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional, List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

from app.utils.locale_data import locale_data_path

# R6.5 — path tenant-aware. Modulo: cassa (corrispettivi export).
DB_PATH = str(locale_data_path("admin_finance.sqlite3"))

# ══════════════════════════════════════════════
# CANONICAL FIELDS — fonte di verità
# ══════════════════════════════════════════════

# Ordine colonne nell'Excel
CANONICAL_COLUMNS = [
    {"db": "date",             "excel": "Data",              "type": "date",   "required": True,  "description": "Data chiusura (YYYY-MM-DD)"},
    {"db": "weekday",          "excel": "Giorno",            "type": "text",   "required": False, "description": "Giorno della settimana"},
    {"db": "corrispettivi",    "excel": "Corrispettivi",     "type": "euro",   "required": False, "description": "Chiusura RT (IVA 10% + IVA 22%)"},
    {"db": "iva_10",           "excel": "IVA 10%",           "type": "euro",   "required": False, "description": "Corrispettivi aliquota 10%"},
    {"db": "iva_22",           "excel": "IVA 22%",           "type": "euro",   "required": False, "description": "Corrispettivi aliquota 22%"},
    {"db": "fatture",          "excel": "Fatture",           "type": "euro",   "required": False, "description": "Fatture emesse"},
    {"db": "corrispettivi_tot","excel": "Corrispettivi Tot",  "type": "euro",   "required": False, "description": "Corrispettivi + Fatture (calcolato se omesso)"},
    {"db": "contanti_finali",  "excel": "Contanti",          "type": "euro",   "required": False, "description": "Incasso contanti"},
    {"db": "pos_bpm",          "excel": "POS BPM",           "type": "euro",   "required": False, "description": "POS BPM (Risto)"},
    {"db": "pos_sella",        "excel": "POS Sella",         "type": "euro",   "required": False, "description": "POS Sella"},
    {"db": "theforkpay",       "excel": "TheForkPay",        "type": "euro",   "required": False, "description": "TheFork Pay"},
    {"db": "other_e_payments", "excel": "Stripe/PayPal",     "type": "euro",   "required": False, "description": "Stripe, PayPal, altri digitali"},
    {"db": "bonifici",         "excel": "Bonifici",          "type": "euro",   "required": False, "description": "Bonifici bancari"},
    {"db": "mance",            "excel": "Mance",             "type": "euro",   "required": False, "description": "Mance digitali (non incasso)"},
    {"db": "note",             "excel": "Note",              "type": "text",   "required": False, "description": "Note libere"},
    {"db": "is_closed",        "excel": "Chiuso",            "type": "int",    "required": False, "description": "1 = giorno chiuso, 0 = aperto"},
]

EXCEL_HEADERS = [c["excel"] for c in CANONICAL_COLUMNS]
DB_FIELDS = [c["db"] for c in CANONICAL_COLUMNS]

# Mapping Excel header → DB field (per import)
HEADER_TO_DB = {c["excel"].upper(): c["db"] for c in CANONICAL_COLUMNS}
# Aliases for backward compatibility with old Excel formats
HEADER_ALIASES = {
    "CORRISPETTIVI-TOT": "corrispettivi_tot",
    "CORRISPETTIVI TOT": "corrispettivi_tot",
    "POS": "pos_bpm",
    "POSBPM": "pos_bpm",
    "POS RISTO": "pos_bpm",
    "SELLA": "pos_sella",
    "POSSELLA": "pos_sella",
    "THEFORK": "theforkpay",
    "THEFORKPAY": "theforkpay",
    "PAYPAL": "other_e_payments",
    "STRIPE": "other_e_payments",
    "PAYPAL/STRIPE": "other_e_payments",
    "STRIPE/PAYPAL": "other_e_payments",
    "MANCE DIG": "mance",
    "IVA10%": "iva_10",
    "IVA22%": "iva_22",
    "IVA 10": "iva_10",
    "IVA 22": "iva_22",
    "GIORNO": "weekday",
    "DATA": "date",
}


# ══════════════════════════════════════════════
# STILI EXCEL
# ══════════════════════════════════════════════

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
EURO_FORMAT = '#,##0.00'
DATE_FORMAT = 'YYYY-MM-DD'


def _style_header(ws, ncols: int):
    """Applica stile all'header."""
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws, ncols: int, nrows: int):
    """Auto-width basato su contenuto."""
    for col_idx in range(1, ncols + 1):
        max_len = 0
        for row_idx in range(1, min(nrows + 2, 100)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(max_len + 3, 12)


# ══════════════════════════════════════════════
# EXPORT DB → EXCEL
# ══════════════════════════════════════════════

def _merge_shift_and_daily(conn, where_sql: str, params: list, ym_prefix: str = None) -> list:
    """
    Merge shift_closures (primario) e daily_closures (fallback) in righe
    compatibili con il formato export canonico.
    """
    from datetime import datetime as dt

    WEEKDAY_IT = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]

    # 1. Leggi daily_closures
    query = f"""
        SELECT date, weekday,
               corrispettivi, iva_10, iva_22, fatture, corrispettivi_tot,
               contanti_finali, pos_bpm, pos_sella, theforkpay, other_e_payments,
               bonifici, mance, note, COALESCE(is_closed, 0) as is_closed
        FROM daily_closures
        {where_sql}
        ORDER BY date ASC
    """
    daily_rows = conn.execute(query, params).fetchall()
    daily_map = {r["date"]: dict(r) for r in daily_rows}

    # 2. Leggi shift_closures per lo stesso periodo
    shift_where = where_sql.replace("daily_closures", "shift_closures") if "daily_closures" in where_sql else where_sql
    shift_query = f"""
        SELECT date, turno, preconto, fatture, contanti,
               pos_bpm, pos_sella, theforkpay, other_e_payments,
               bonifici, mance, note
        FROM shift_closures
        {where_sql}
        ORDER BY date ASC
    """
    shift_rows = conn.execute(shift_query, params).fetchall()

    # 3. Aggrega shift_closures per data
    shift_by_date = {}
    for r in shift_rows:
        d = r["date"]
        shift_by_date.setdefault(d, []).append(r)

    shift_map = {}
    for date_str, turni in shift_by_date.items():
        pranzo = None
        cena = None
        for t in turni:
            if t["turno"] == "pranzo":
                pranzo = t
            else:
                cena = t

        # Base giornaliera = cena se esiste, altrimenti pranzo
        base = cena or pranzo
        chiusura = base["preconto"] or 0
        fatture_tot = (pranzo["fatture"] if pranzo else 0) + (cena["fatture"] if cena else 0)
        contanti = base["contanti"] or 0
        pos_bpm = base["pos_bpm"] or 0
        pos_sella = base["pos_sella"] or 0
        theforkpay = base["theforkpay"] or 0
        other_e = base["other_e_payments"] or 0
        bonifici = base["bonifici"] or 0
        mance = base["mance"] or 0

        note_parts = []
        if pranzo and pranzo["note"]:
            note_parts.append(f"P: {pranzo['note']}")
        if cena and cena["note"]:
            note_parts.append(f"C: {cena['note']}")

        d = dt.strptime(date_str, "%Y-%m-%d").date()
        weekday = WEEKDAY_IT[d.weekday()]

        shift_map[date_str] = {
            "date": date_str,
            "weekday": weekday,
            "corrispettivi": chiusura,
            "iva_10": 0.0,
            "iva_22": 0.0,
            "fatture": fatture_tot,
            "corrispettivi_tot": chiusura + fatture_tot,
            "contanti_finali": contanti,
            "pos_bpm": pos_bpm,
            "pos_sella": pos_sella,
            "theforkpay": theforkpay,
            "other_e_payments": other_e,
            "bonifici": bonifici,
            "mance": mance,
            "note": " | ".join(note_parts) if note_parts else "",
            "is_closed": 0,
        }

    # 4. Merge: shift_closures è primario, daily_closures è fallback
    all_dates = sorted(set(list(shift_map.keys()) + list(daily_map.keys())))
    merged = []
    for d in all_dates:
        if d in shift_map:
            merged.append(shift_map[d])
        else:
            merged.append(daily_map[d])

    return merged


def export_corrispettivi_to_excel(
    year: Optional[int] = None,
    month: Optional[int] = None,
    include_shift_data: bool = True,
) -> bytes:
    """
    Esporta daily_closures + shift_closures (merged) in formato Excel canonico.
    Se year/month specificati, filtra per periodo.
    Ritorna bytes del file Excel.
    """
    # Fix 1.11.2 (sessione 52) — WAL + synchronous + busy_timeout
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA busy_timeout=30000")

    try:
        where_clauses = []
        params = []
        ym_prefix = None

        if year and month:
            ym_prefix = f"{year:04d}-{month:02d}"
            where_clauses.append("substr(date, 1, 7) = ?")
            params.append(ym_prefix)
        elif year:
            where_clauses.append("substr(date, 1, 4) = ?")
            params.append(str(year))

        where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

        rows = _merge_shift_and_daily(conn, where_sql, params, ym_prefix)
    finally:
        conn.close()

    # Create Excel
    wb = Workbook()
    ws = wb.active

    # Sheet name
    if year and month:
        ws.title = f"{year}-{month:02d}"
    elif year:
        ws.title = str(year)
    else:
        ws.title = "Corrispettivi"

    # Headers
    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_def in enumerate(CANONICAL_COLUMNS, 1):
            val = row.get(col_def["db"]) if isinstance(row, dict) else row[col_def["db"]]
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_def["type"] == "euro":
                cell.value = float(val) if val else 0.0
                cell.number_format = EURO_FORMAT
            elif col_def["type"] == "date":
                cell.value = val
            elif col_def["type"] == "int":
                cell.value = int(val) if val else 0
            else:
                cell.value = str(val) if val else ""

            cell.border = THIN_BORDER

    nrows = len(rows)
    ncols = len(EXCEL_HEADERS)
    _style_header(ws, ncols)
    _auto_width(ws, ncols, nrows)

    # Totals row
    if nrows > 0:
        total_row = nrows + 2
        ws.cell(row=total_row, column=1, value="TOTALE").font = Font(bold=True)
        for col_idx, col_def in enumerate(CANONICAL_COLUMNS, 1):
            if col_def["type"] == "euro":
                # Sum formula
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                cell = ws.cell(row=total_row, column=col_idx)
                cell.value = f"=SUM({col_letter}2:{col_letter}{nrows + 1})"
                cell.number_format = EURO_FORMAT
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER

    # Freeze header
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════
# EXPORT DB → PDF (prospetto fiscale commercialista)
# Modulo: cassa
# ══════════════════════════════════════════════

_WEEKDAY_IT = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
_MESI_IT = ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
            "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]


def _fmt_euro_it(value) -> str:
    """Formatta un numero in stile italiano: 1.234,56 (senza simbolo €)."""
    try:
        n = float(value or 0)
    except (TypeError, ValueError):
        n = 0.0
    s = f"{n:,.2f}"  # formato US: 1,234.56
    return s.replace(",", "§").replace(".", ",").replace("§", ".")


def _scorpora_imponibile(lordo: float, aliquota_pct: int) -> float:
    """Imponibile netto scorporato da un importo lordo (IVA inclusa).

    Es. lordo 110,00 al 10% → imponibile 100,00. Arrotondamento commerciale
    (half-up) a 2 decimali, come richiesto per i documenti fiscali.
    """
    if not lordo or lordo <= 0:
        return 0.0
    div = Decimal(100 + aliquota_pct) / Decimal(100)
    imp = (Decimal(str(lordo)) / div).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return float(imp)


def _corrispettivi_pdf_css() -> str:
    """CSS compatto: comprime righe, header e box per far stare un mese
    intero (fino a 31 giorni) in una sola pagina A4."""
    return """
    @page { margin: 11mm 12mm 13mm 12mm; }
    .brand-header { padding-bottom: 5px; margin-bottom: 8px; }
    .brand-header .wordmark-svg svg { height: 22pt; }
    .brand-header .wordmark-text { font-size: 21pt; }
    .brand-header .doc-title { font-size: 12pt; }
    .gobbette-strip { margin: -6px 0 8px 0; }
    .summary-row { margin: 2pt 0 7pt 0; }
    .summary-box { padding: 3pt 9pt; margin: 0 5pt 5pt 0; }
    .summary-box .label { font-size: 6.5pt; }
    .summary-box .value { font-size: 10.5pt; }
    table.brand { margin: 4pt 0 6pt 0; }
    table.brand th { padding: 2.6pt 6pt; font-size: 7pt; }
    table.brand td { padding: 1.7pt 6pt; font-size: 8pt; line-height: 1.15; }
    h3 { font-size: 9.5pt; margin: 6pt 0 1pt 0; }
    .small { font-size: 6.8pt; line-height: 1.3; margin: 4pt 0 0 0; }
    """


def build_corrispettivi_pdf(year: int, month: int) -> bytes:
    """
    Genera il PDF del prospetto fiscale dei corrispettivi di un mese, pensato
    per il controllo del commercialista.

    Sorgente dati: fonte unita `_merge_shift_and_daily` — chiusure turno
    (`shift_closures`, operative) come primaria, import Excel (`daily_closures`)
    come ripiego. Stesso pattern di export Excel e dashboard.

    Ripartizione IVA: le chiusure turno non registrano lo split 10/22. L'osteria
    fa somministrazione pura, quindi i giorni senza split vengono trattati come
    interamente IVA 10% (decisione Marco 2026-05-21).

    Per ogni giorno riporta il corrispettivo al lordo (IVA inclusa) e il suo
    scorporo in imponibile + imposta; le fatture emesse a parte; il totale.
    Chiude con i totali del mese e un riepilogo IVA per aliquota.

    NB: non riproduce il tracciato XML 7.0 dei corrispettivi telematici (formato
    di trasmissione macchina-a-macchina generato dal RT); ne riporta solo la
    sostanza utile al commercialista in forma leggibile.

    Args:
        year: anno (es. 2026)
        month: mese 1-12

    Returns:
        bytes del PDF brandizzato (mattone M.B `pdf_brand`).

    Raises:
        ValueError: se non esistono righe per il periodo richiesto.
    """
    from datetime import datetime as _dt
    from app.services.pdf_brand import wrappa_html_brand

    ym_prefix = f"{year:04d}-{month:02d}"

    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout=30000")
    try:
        rows = _merge_shift_and_daily(
            conn, "WHERE substr(date, 1, 7) = ?", [ym_prefix], ym_prefix
        )
    finally:
        conn.close()

    if not rows:
        raise ValueError(
            f"Nessun corrispettivo registrato per {_MESI_IT[month]} {year}."
        )

    # Totali per aliquota (per il riepilogo IVA) + totali generali.
    lordo10 = imp10_tot = iva10_tot = 0.0
    lordo22 = imp22_tot = iva22_tot = 0.0
    tot_fatt = tot_gen = 0.0
    giorni_con_incasso = 0
    body_rows = []
    note_rows = []  # (data_label, testo_nota) per la tabella note in coda

    for r in rows:
        try:
            d_obj = _dt.strptime(r["date"], "%Y-%m-%d").date()
        except (TypeError, ValueError):
            d_obj = None
        data_label = d_obj.strftime("%d/%m/%Y") if d_obj else (r.get("date") or "")
        giorno_label = r.get("weekday") or (_WEEKDAY_IT[d_obj.weekday()] if d_obj else "")

        # Raccogli le note del giorno (incluse quelle dei giorni chiusi)
        nota = (r.get("note") or "").strip()
        if nota:
            note_rows.append((data_label, nota))

        corr = float(r.get("corrispettivi") or 0)
        i10 = float(r.get("iva_10") or 0)
        i22 = float(r.get("iva_22") or 0)
        fatt = float(r.get("fatture") or 0)

        if r.get("is_closed"):
            body_rows.append(
                f"<tr><td>{data_label}</td><td>{giorno_label}</td>"
                f"<td colspan='5' class='text-muted' style='text-align:center'>— chiuso —</td></tr>"
            )
            continue

        # Ripartizione IVA: le chiusure turno non registrano lo split 10/22.
        # Somministrazione pura → giorni senza split = interamente IVA 10%
        # (decisione Marco 2026-05-21).
        if corr > 0 and i10 == 0 and i22 == 0:
            i10 = corr

        # Scorporo: dal lordo (IVA inclusa) → imponibile netto + imposta.
        imp10 = _scorpora_imponibile(i10, 10)
        iva10 = round(i10 - imp10, 2)
        imp22 = _scorpora_imponibile(i22, 22)
        iva22 = round(i22 - imp22, 2)

        day_lordo = i10 + i22
        day_imponibile = round(imp10 + imp22, 2)
        day_imposta = round(iva10 + iva22, 2)
        day_totale = day_lordo + fatt

        lordo10 += i10
        imp10_tot += imp10
        iva10_tot += iva10
        lordo22 += i22
        imp22_tot += imp22
        iva22_tot += iva22
        tot_fatt += fatt
        tot_gen += day_totale
        if day_lordo > 0:
            giorni_con_incasso += 1

        body_rows.append(
            f"<tr><td>{data_label}</td><td>{giorno_label}</td>"
            f"<td class='num'>{_fmt_euro_it(day_lordo)}</td>"
            f"<td class='num'>{_fmt_euro_it(day_imponibile)}</td>"
            f"<td class='num'>{_fmt_euro_it(day_imposta)}</td>"
            f"<td class='num'>{_fmt_euro_it(fatt)}</td>"
            f"<td class='num'>{_fmt_euro_it(day_totale)}</td></tr>"
        )

    tot_lordo = lordo10 + lordo22
    tot_imponibile = imp10_tot + imp22_tot
    tot_imposta = iva10_tot + iva22_tot

    # Riga totali in tbody (non in tfoot: tfoot verrebbe ripetuto su ogni pagina).
    body_rows.append(
        f"<tr class='tot-row'><td colspan='2'>TOTALE {_MESI_IT[month].upper()} {year}</td>"
        f"<td class='num'>{_fmt_euro_it(tot_lordo)}</td>"
        f"<td class='num'>{_fmt_euro_it(tot_imponibile)}</td>"
        f"<td class='num'>{_fmt_euro_it(tot_imposta)}</td>"
        f"<td class='num'>{_fmt_euro_it(tot_fatt)}</td>"
        f"<td class='num'>{_fmt_euro_it(tot_gen)}</td></tr>"
    )

    summary = (
        "<div class='summary-row'>"
        f"<div class='summary-box'><div class='label'>Corrispettivi lordo</div>"
        f"<div class='value'>&euro; {_fmt_euro_it(tot_lordo)}</div></div>"
        f"<div class='summary-box'><div class='label'>Imponibile</div>"
        f"<div class='value'>&euro; {_fmt_euro_it(tot_imponibile)}</div></div>"
        f"<div class='summary-box'><div class='label'>IVA (imposta)</div>"
        f"<div class='value'>&euro; {_fmt_euro_it(tot_imposta)}</div></div>"
        f"<div class='summary-box'><div class='label'>Fatture emesse</div>"
        f"<div class='value'>&euro; {_fmt_euro_it(tot_fatt)}</div></div>"
        "</div>"
    )

    tabella = (
        "<table class='brand'>"
        "<thead><tr>"
        "<th>Data</th><th>Giorno</th>"
        "<th class='num'>Corrispettivo lordo</th>"
        "<th class='num'>Imponibile 10%</th>"
        "<th class='num'>IVA 10%</th>"
        "<th class='num'>Fatture</th>"
        "<th class='num'>Totale</th>"
        "</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody>"
        "</table>"
    )

    # Riepilogo IVA per aliquota — la base su cui il commercialista liquida l'IVA.
    riep_rows = []
    if lordo10 > 0 or (lordo22 == 0):
        riep_rows.append(
            f"<tr><td>10%</td><td class='num'>{_fmt_euro_it(lordo10)}</td>"
            f"<td class='num'>{_fmt_euro_it(imp10_tot)}</td>"
            f"<td class='num'>{_fmt_euro_it(iva10_tot)}</td></tr>"
        )
    if lordo22 > 0:
        riep_rows.append(
            f"<tr><td>22%</td><td class='num'>{_fmt_euro_it(lordo22)}</td>"
            f"<td class='num'>{_fmt_euro_it(imp22_tot)}</td>"
            f"<td class='num'>{_fmt_euro_it(iva22_tot)}</td></tr>"
        )
    riep_rows.append(
        f"<tr class='tot-row'><td>Totale</td><td class='num'>{_fmt_euro_it(tot_lordo)}</td>"
        f"<td class='num'>{_fmt_euro_it(tot_imponibile)}</td>"
        f"<td class='num'>{_fmt_euro_it(tot_imposta)}</td></tr>"
    )

    riepilogo = (
        "<h3>Riepilogo IVA</h3>"
        "<table class='brand'>"
        "<thead><tr><th>Aliquota</th>"
        "<th class='num'>Corrispettivi lordo</th>"
        "<th class='num'>Imponibile</th>"
        "<th class='num'>Imposta</th></tr></thead>"
        f"<tbody>{''.join(riep_rows)}</tbody>"
        "</table>"
        "<p class='small'>Corrispettivi al lordo (IVA inclusa); imponibile e imposta ne "
        "sono lo scorporo. Fatture emesse indicate a parte, al lordo. &laquo;Totale&raquo; "
        "= corrispettivo lordo + fatture. Aliquota 10% (somministrazione di alimenti e "
        "bevande); i giorni dalle chiusure turno, privi di split IVA, sono trattati al 10%. "
        f"Giorni con incasso nel mese: {giorni_con_incasso}.</p>"
    )

    # Tabella note — solo se ci sono note nel mese. Va dopo il riepilogo IVA.
    note_section = ""
    if note_rows:
        import html as _html
        righe_note = "".join(
            f"<tr><td style='white-space:nowrap'>{d}</td>"
            f"<td>{_html.escape(n)}</td></tr>"
            for d, n in note_rows
        )
        note_section = (
            "<h3>Note</h3>"
            "<table class='brand'>"
            "<thead><tr><th style='width:18%'>Data</th><th>Nota</th></tr></thead>"
            f"<tbody>{righe_note}</tbody>"
            "</table>"
            "<p class='small'>Note libere registrate nelle chiusure (P = pranzo, "
            "C = cena).</p>"
        )

    return wrappa_html_brand(
        titolo="Corrispettivi — Controllo Commercialista",
        sottotitolo=f"{_MESI_IT[month]} {year} &mdash; IVA 10%",
        body_html=summary + tabella + riepilogo + note_section,
        orientamento="portrait",
        css_extra=_corrispettivi_pdf_css(),
    )


# ══════════════════════════════════════════════
# TEMPLATE VUOTO
# ══════════════════════════════════════════════

def generate_template() -> bytes:
    """
    Genera un Excel template vuoto con:
    - Header canonici
    - Descrizione campi nella riga 2
    - Formattazione pronta
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    # Headers
    ncols = len(EXCEL_HEADERS)
    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    _style_header(ws, ncols)

    # Description row (row 2, italic, gray)
    desc_font = Font(italic=True, color="888888", size=9)
    for col_idx, col_def in enumerate(CANONICAL_COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_def["description"])
        cell.font = desc_font
        cell.alignment = Alignment(wrap_text=True)

    # Example row (row 3)
    example_data = {
        "date": "2026-01-15",
        "weekday": "Giovedì",
        "corrispettivi": 2450.00,
        "iva_10": 1200.00,
        "iva_22": 1250.00,
        "fatture": 150.00,
        "corrispettivi_tot": 2600.00,
        "contanti_finali": 800.00,
        "pos_bpm": 1200.00,
        "pos_sella": 350.00,
        "theforkpay": 100.00,
        "other_e_payments": 50.00,
        "bonifici": 100.00,
        "mance": 25.00,
        "note": "Esempio — eliminare questa riga",
        "is_closed": 0,
    }
    example_font = Font(color="999999", size=10)
    for col_idx, col_def in enumerate(CANONICAL_COLUMNS, 1):
        val = example_data.get(col_def["db"], "")
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.font = example_font
        if col_def["type"] == "euro":
            cell.number_format = EURO_FORMAT

    # Auto width
    _auto_width(ws, ncols, 3)

    # Instructions sheet
    ws_info = wb.create_sheet("Istruzioni")
    ws_info.cell(row=1, column=1, value="TRGB Gestionale — Template Corrispettivi").font = Font(bold=True, size=14)
    ws_info.cell(row=3, column=1, value="Istruzioni:").font = Font(bold=True)

    instructions = [
        "1. Compilare il foglio 'Template' con i dati giornalieri dei corrispettivi.",
        "2. Il campo 'Data' (YYYY-MM-DD) è obbligatorio.",
        "3. I campi numerici accettano sia formato italiano (1.234,56) che internazionale (1234.56).",
        "4. 'Corrispettivi Tot' = Corrispettivi + Fatture. Se omesso, viene calcolato automaticamente.",
        "5. 'Giorno' viene calcolato automaticamente dalla data se omesso.",
        "6. 'Chiuso' = 1 per segnare un giorno di chiusura (ferie, festivi).",
        "7. La riga di esempio (riga 3) va eliminata prima dell'import.",
        "8. Per importare, usare Vendite → Impostazioni → Import/Export.",
        "",
        "Campi calcolati automaticamente (non presenti nel template):",
        "• Totale Incassi = Contanti + POS BPM + POS Sella + TheForkPay + Stripe/PayPal + Bonifici",
        "• Differenza Cassa = Totale Incassi - Corrispettivi Tot",
    ]
    for i, line in enumerate(instructions, 4):
        ws_info.cell(row=i, column=1, value=line)

    ws_info.column_dimensions["A"].width = 80

    # Field reference
    ws_ref = wb.create_sheet("Campi")
    ws_ref.cell(row=1, column=1, value="Campo Excel").font = Font(bold=True)
    ws_ref.cell(row=1, column=2, value="Campo DB").font = Font(bold=True)
    ws_ref.cell(row=1, column=3, value="Tipo").font = Font(bold=True)
    ws_ref.cell(row=1, column=4, value="Obbligatorio").font = Font(bold=True)
    ws_ref.cell(row=1, column=5, value="Descrizione").font = Font(bold=True)

    for i, col_def in enumerate(CANONICAL_COLUMNS, 2):
        ws_ref.cell(row=i, column=1, value=col_def["excel"])
        ws_ref.cell(row=i, column=2, value=col_def["db"])
        ws_ref.cell(row=i, column=3, value=col_def["type"])
        ws_ref.cell(row=i, column=4, value="Sì" if col_def["required"] else "No")
        ws_ref.cell(row=i, column=5, value=col_def["description"])

    for col in ["A", "B", "C", "D", "E"]:
        ws_ref.column_dimensions[col].width = 20

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
