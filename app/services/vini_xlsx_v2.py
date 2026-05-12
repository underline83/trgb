# Modulo: vini
"""
Service: Import/Export Vini v2 (sessione 2026-05-12, V-H.J).

Genera template Excel pulito per il nuovo formato, parsa import dal nuovo
formato, esporta tutti i vini nello stesso formato (così template e backup
sono speculari: scarica → modifica → reimporta).

Sostituisce completamente la vecchia logica `vini_model.normalize_dataframe`
+ `import_excel_to_cantina` + `export_excel` (eredità Excel originale).

PRINCIPI:
- DB è la source of truth, Excel è solo I/O leggibile umano.
- Chiave d'unicità per import = `id` (vini_magazzino.id, auto-increment DB).
- Import idempotente: se ID esiste → SKIP (no overwrite). Se ID vuoto → INSERT.
- Flag SI/NO leggibili nel file Excel, conversione 0/1 automatica al DB.
- Template/export hanno le STESSE colonne (round-trip pulito).
"""

from __future__ import annotations
from io import BytesIO
from typing import Any, Dict, List, Tuple
import sqlite3

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models import vini_magazzino_db as mag_db


# ============================================================
# COSTANTI DI VALIDAZIONE (single source of truth)
# Promosse qui da vini_model.py (eliminato in V-H.J).
# ============================================================

TIPOLOGIA_VALIDE = [
    "GRANDI FORMATI",
    "BOLLICINE",
    "BIANCHI",
    "ROSATI",
    "ROSSI",
    "PASSITI E VINI DA MEDITAZIONE",
    "VINI ANALCOLICI",
    "ERRORE",
]

FORMATO_VALIDI = [
    "MN", "QP", "ME", "DM", "CL", "BT", "BN", "MG", "MJ",
    "JB", "RH", "JBX", "MS", "SM", "BZ", "NB", "ML", "PR", "MZ",
]

STATO_VENDITA_VALIDI = ["N", "T", "V", "F", "S", "C"]
STATO_RIORDINO_VALIDI = ["D", "0", "A", "X"]
STATO_CONSERVAZIONE_VALIDI = ["1", "2", "3"]


# ============================================================
# SCHEMA TEMPLATE: ordine + tipo + obbligatorio
# Una sola fonte di verità per template, import e export.
# ============================================================

# Tipo: 'int_id'   → id DB (auto-increment, vuoto su nuovo)
#       'str_req'  → stringa obbligatoria
#       'str_opt'  → stringa opzionale
#       'num'      → numero (float)
#       'int'      → numero intero
#       'qta'      → intero giacenza (default 0)
#       'flag_yn'  → SI/NO leggibile (convertito a 0/1 lato DB)
#       'enum'     → valore in lista controllata
#       'isodate'  → data ISO YYYY-MM-DD (solo export)

TEMPLATE_COLUMNS: List[Dict[str, Any]] = [
    # SISTEMA
    {"col": "ID",                   "field": "id",                    "tipo": "int_id",   "obbligatorio": False, "hint": "vuoto = nuovo vino; pieno = vino esistente (verrà saltato)"},
    # ANAGRAFICA
    {"col": "TIPOLOGIA",            "field": "TIPOLOGIA",             "tipo": "enum",     "obbligatorio": True,  "hint": "BOLLICINE / BIANCHI / ROSSI / ROSATI / GRANDI FORMATI / PASSITI E VINI DA MEDITAZIONE / VINI ANALCOLICI", "options": TIPOLOGIA_VALIDE},
    {"col": "DESCRIZIONE",          "field": "DESCRIZIONE",           "tipo": "str_req",  "obbligatorio": True,  "hint": "nome vino"},
    {"col": "PRODUTTORE",           "field": "PRODUTTORE",            "tipo": "str_opt",  "obbligatorio": False, "hint": ""},
    {"col": "ANNATA",               "field": "ANNATA",                "tipo": "str_opt",  "obbligatorio": False, "hint": "es. 2021"},
    {"col": "FORMATO",              "field": "FORMATO",               "tipo": "enum",     "obbligatorio": False, "hint": "BT=Bottiglia 0.75L, MG=Magnum, ME=Mezza, DM=Demie 0.5L, ecc.", "options": FORMATO_VALIDI},
    {"col": "NAZIONE",              "field": "NAZIONE",               "tipo": "str_req",  "obbligatorio": True,  "hint": "es. Italia, Francia, Germania"},
    {"col": "REGIONE",              "field": "REGIONE",               "tipo": "str_opt",  "obbligatorio": False, "hint": ""},
    {"col": "DENOMINAZIONE",        "field": "DENOMINAZIONE",         "tipo": "str_opt",  "obbligatorio": False, "hint": "es. Barolo DOCG"},
    {"col": "VITIGNI",              "field": "VITIGNI",               "tipo": "str_opt",  "obbligatorio": False, "hint": "es. Nebbiolo, Sangiovese"},
    {"col": "GRADO_ALCOLICO",       "field": "GRADO_ALCOLICO",        "tipo": "num",      "obbligatorio": False, "hint": "% volume, es. 13.5"},
    {"col": "DISTRIBUTORE",         "field": "DISTRIBUTORE",          "tipo": "str_opt",  "obbligatorio": False, "hint": ""},
    {"col": "RAPPRESENTANTE",       "field": "RAPPRESENTANTE",        "tipo": "str_opt",  "obbligatorio": False, "hint": ""},
    # PREZZI
    {"col": "PREZZO_CARTA",         "field": "PREZZO_CARTA",          "tipo": "num",      "obbligatorio": False, "hint": "prezzo bottiglia in carta cliente"},
    {"col": "EURO_LISTINO",         "field": "EURO_LISTINO",          "tipo": "num",      "obbligatorio": False, "hint": "costo fornitore"},
    {"col": "SCONTO",               "field": "SCONTO",                "tipo": "num",      "obbligatorio": False, "hint": ""},
    {"col": "PREZZO_CALICE",        "field": "PREZZO_CALICE",         "tipo": "num",      "obbligatorio": False, "hint": "vuoto = auto-calcolo PREZZO_CARTA/5"},
    {"col": "NOTE_PREZZO",          "field": "NOTE_PREZZO",           "tipo": "str_opt",  "obbligatorio": False, "hint": ""},
    # FLAG
    {"col": "CARTA",                "field": "CARTA",                 "tipo": "flag_yn",  "obbligatorio": False, "hint": "SI/NO — pubblicato in carta cliente"},
    {"col": "IPRATICO",             "field": "IPRATICO",              "tipo": "flag_yn",  "obbligatorio": False, "hint": "SI/NO — esportato verso iPratico"},
    {"col": "BIOLOGICO",            "field": "BIOLOGICO",             "tipo": "flag_yn",  "obbligatorio": False, "hint": "SI/NO"},
    {"col": "VENDITA_CALICE",       "field": "VENDITA_CALICE",        "tipo": "flag_yn",  "obbligatorio": False, "hint": "SI/NO — abilitato per vendita al calice"},
    {"col": "ABBINAMENTI",          "field": "ABBINAMENTI",           "tipo": "str_opt",  "obbligatorio": False, "hint": "abbinamenti consigliati (mostrato in carta cliente per i calici)"},
    # STATI
    {"col": "STATO_VENDITA",        "field": "STATO_VENDITA",         "tipo": "enum",     "obbligatorio": False, "hint": "N=Non vendere, T=Cautela, V=Vendere, F=Spingere, S=Aggressivo, C=Controllare", "options": STATO_VENDITA_VALIDI},
    {"col": "STATO_RIORDINO",       "field": "STATO_RIORDINO",        "tipo": "enum",     "obbligatorio": False, "hint": "D=Da ordinare, 0=Ordinato, A=Annata esaurita, X=Non ricomprare", "options": STATO_RIORDINO_VALIDI},
    {"col": "STATO_CONSERVAZIONE",  "field": "STATO_CONSERVAZIONE",   "tipo": "enum",     "obbligatorio": False, "hint": "1=Difficile, 2=Buona, 3=Perfetta", "options": STATO_CONSERVAZIONE_VALIDI},
    {"col": "NOTE_STATO",           "field": "NOTE_STATO",            "tipo": "str_opt",  "obbligatorio": False, "hint": ""},
    # LOCAZIONI E GIACENZE INIZIALI
    {"col": "FRIGORIFERO",          "field": "FRIGORIFERO",           "tipo": "str_opt",  "obbligatorio": False, "hint": "nome locazione frigo (vedi foglio Locazioni)"},
    {"col": "QTA_FRIGO",            "field": "QTA_FRIGO",             "tipo": "qta",      "obbligatorio": False, "hint": "bottiglie in frigo"},
    {"col": "LOCAZIONE_1",          "field": "LOCAZIONE_1",           "tipo": "str_opt",  "obbligatorio": False, "hint": ""},
    {"col": "QTA_LOC1",             "field": "QTA_LOC1",              "tipo": "qta",      "obbligatorio": False, "hint": ""},
    {"col": "LOCAZIONE_2",          "field": "LOCAZIONE_2",           "tipo": "str_opt",  "obbligatorio": False, "hint": ""},
    {"col": "QTA_LOC2",             "field": "QTA_LOC2",              "tipo": "qta",      "obbligatorio": False, "hint": ""},
    {"col": "LOCAZIONE_3",          "field": "LOCAZIONE_3",           "tipo": "str_opt",  "obbligatorio": False, "hint": ""},
    {"col": "QTA_LOC3",             "field": "QTA_LOC3",              "tipo": "qta",      "obbligatorio": False, "hint": ""},
    # NOTE
    {"col": "NOTE",                 "field": "NOTE",                  "tipo": "str_opt",  "obbligatorio": False, "hint": "note operative interne"},
]


# Helper conversione SI/NO ↔ 0/1
def _yn_to_int(v: Any):
    if v is None or v == "":
        return None
    s = str(v).strip().upper()
    if s in ("SI", "1", "TRUE", "YES"):
        return 1
    if s in ("NO", "0", "FALSE", "N"):
        return 0
    return None


def _int_to_yn(v: Any) -> str:
    if v == 1:
        return "SI"
    if v == 0:
        return "NO"
    return ""


def _safe_int(v: Any):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _safe_float(v: Any):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ============================================================
# GENERAZIONE TEMPLATE / EXPORT
# ============================================================

HEADER_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
HEADER_REQ_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="8D4E00")


def _get_locazioni_config() -> Dict[str, List[str]]:
    """
    Legge i nomi locazioni dalla tabella `locazioni_config` che vive nel DB
    `vini_magazzino.sqlite3` (schema in `vini_magazzino_db.py:358`), NON in
    `vini_settings.sqlite3`. Fix audit V-H.J.
    """
    try:
        conn = mag_db.get_magazzino_connection()
        cur = conn.cursor()
        row = cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='locazioni_config'"
        ).fetchone()
        if not row:
            conn.close()
            return {"frigorifero": [], "locazione_1": [], "locazione_2": [], "locazione_3": []}
        result = {"frigorifero": [], "locazione_1": [], "locazione_2": [], "locazione_3": []}
        for campo in result.keys():
            rows = cur.execute(
                "SELECT nome FROM locazioni_config WHERE campo = ? ORDER BY ordine",
                (campo,),
            ).fetchall()
            result[campo] = [r[0] for r in rows]
        conn.close()
        return result
    except Exception:
        return {"frigorifero": [], "locazione_1": [], "locazione_2": [], "locazione_3": []}


def _scrivi_foglio_vini(ws, righe_dati: List[Dict[str, Any]] = None) -> None:
    """
    Scrive il foglio 'Vini' con header colorato + righe dati (vuote o popolate).
    Se righe_dati è None o vuoto, scrive 1 riga esempio.
    """
    # Header
    for idx, c in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=c["col"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_REQ_FILL if c["obbligatorio"] else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Larghezza colonna auto-stimata
        ws.column_dimensions[get_column_letter(idx)].width = max(len(c["col"]) + 2, 14)

    # Freeze header
    ws.freeze_panes = "A2"

    # Riga dati
    if not righe_dati:
        # Esempio template
        esempio = {
            "TIPOLOGIA": "ROSSI",
            "DESCRIZIONE": "Barolo Esempio",
            "PRODUTTORE": "Cantina Esempio",
            "ANNATA": "2020",
            "FORMATO": "BT",
            "NAZIONE": "Italia",
            "REGIONE": "Piemonte",
            "DENOMINAZIONE": "Barolo DOCG",
            "VITIGNI": "Nebbiolo",
            "GRADO_ALCOLICO": 14.0,
            "PREZZO_CARTA": 45,
            "EURO_LISTINO": 18,
            "CARTA": "SI",
            "IPRATICO": "NO",
            "BIOLOGICO": "NO",
            "VENDITA_CALICE": "NO",
            "STATO_VENDITA": "V",
            "QTA_FRIGO": 3,
            "FRIGORIFERO": "Cantina",
            "NOTE": "<-- riga di esempio, eliminala prima di importare",
        }
        for idx, c in enumerate(TEMPLATE_COLUMNS, start=1):
            val = esempio.get(c["field"], "")
            cell = ws.cell(row=2, column=idx, value=val)
            if c["field"] == "NOTE" and esempio.get("NOTE"):
                cell.font = Font(italic=True, color="808080")
    else:
        for r_idx, vino in enumerate(righe_dati, start=2):
            for c_idx, c in enumerate(TEMPLATE_COLUMNS, start=1):
                field = c["field"]
                raw = vino.get(field)
                if c["tipo"] == "flag_yn":
                    val = _int_to_yn(raw)
                else:
                    val = raw if raw is not None else ""
                ws.cell(row=r_idx, column=c_idx, value=val)


def _scrivi_foglio_riferimento(ws) -> None:
    """Foglio 'Riferimento valori' con tutti i codici validi."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    row = 1
    ws.cell(row=row, column=1, value="Riferimento valori validi").font = TITLE_FONT
    row += 2

    blocchi = [
        ("TIPOLOGIA", TIPOLOGIA_VALIDE),
        ("FORMATO", FORMATO_VALIDI),
        ("STATO_VENDITA", [
            "N — Non vendere",
            "T — Vendere con cautela",
            "V — Vendere (normale)",
            "F — Spingere",
            "S — Vendere aggressivo",
            "C — Controllare",
        ]),
        ("STATO_RIORDINO", [
            "D — Da ordinare",
            "0 — Ordinato (in arrivo)",
            "A — Annata esaurita",
            "X — Non ricomprare (fuori catalogo)",
        ]),
        ("STATO_CONSERVAZIONE", [
            "1 — Difficile / vendere subito",
            "2 — Buona / vendere",
            "3 — Perfetta / non urgente",
        ]),
        ("Flag SI/NO", [
            "SI — il flag è attivo",
            "NO — il flag non è attivo",
            "(vuoto)  — interpretato come NO",
        ]),
    ]
    for titolo, valori in blocchi:
        cell = ws.cell(row=row, column=1, value=titolo)
        cell.font = Font(bold=True, size=12, color="8D4E00")
        row += 1
        for v in valori:
            ws.cell(row=row, column=1, value=str(v))
            row += 1
        row += 1


def _scrivi_foglio_locazioni(ws, loc_cfg: Dict[str, List[str]]) -> None:
    """Foglio 'Locazioni' con i nomi configurati nel locale."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30

    ws.cell(row=1, column=1, value="Locazioni configurate nel locale").font = TITLE_FONT

    headers = ["FRIGORIFERO", "LOCAZIONE_1", "LOCAZIONE_2", "LOCAZIONE_3"]
    keys = ["frigorifero", "locazione_1", "locazione_2", "locazione_3"]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    max_rows = max([len(loc_cfg.get(k, [])) for k in keys] + [1])
    for r in range(max_rows):
        for c_idx, k in enumerate(keys, start=1):
            vals = loc_cfg.get(k, [])
            if r < len(vals):
                ws.cell(row=4 + r, column=c_idx, value=vals[r])

    nota_row = 5 + max_rows
    ws.cell(row=nota_row, column=1,
            value=("Usa esattamente questi nomi nelle colonne FRIGORIFERO / LOCAZIONE_1 / "
                   "LOCAZIONE_2 / LOCAZIONE_3 del foglio 'Vini'.")
    ).font = Font(italic=True, color="808080")


def _scrivi_foglio_istruzioni(ws) -> None:
    """Foglio 'Istruzioni' testuale con la guida."""
    ws.column_dimensions["A"].width = 100

    righe = [
        ("Istruzioni per la compilazione", TITLE_FONT),
        ("", None),
        ("Questo file è il modello ufficiale per inserire vini in TRGB Gestionale.", None),
        ("Compila il foglio 'Vini' e importa il file da: Impostazioni Vini → Importa.", None),
        ("", None),
        ("Campi OBBLIGATORI (colonne con sfondo arancione scuro):", Font(bold=True, size=11)),
        ("• TIPOLOGIA — scegli da foglio 'Riferimento valori'", None),
        ("• DESCRIZIONE — nome del vino", None),
        ("• NAZIONE — es. Italia, Francia, Germania, Austria", None),
        ("", None),
        ("Campi opzionali utili:", Font(bold=True, size=11)),
        ("• ID: lascia vuoto per un vino nuovo. Se inserisci un ID già esistente nel sistema,", None),
        ("  la riga viene SALTATA (nessuna sovrascrittura). Per modificare un vino esistente,", None),
        ("  usa la scheda del vino dentro il gestionale.", None),
        ("• PREZZO_CALICE: lascia vuoto per il calcolo automatico (PREZZO_CARTA / 5,", None),
        ("  arrotondato a step 0.50 — soglie configurabili da Impostazioni → Widget e soglie).", None),
        ("", None),
        ("Flag SI/NO (CARTA, IPRATICO, BIOLOGICO, VENDITA_CALICE):", Font(bold=True, size=11)),
        ("• Scrivi 'SI' o 'NO' (case insensitive). Vuoto = NO.", None),
        ("• CARTA = SI → il vino entra nella carta cliente.", None),
        ("• VENDITA_CALICE = SI → vino disponibile al calice.", None),
        ("• BIOLOGICO = SI → badge biologico in scheda.", None),
        ("• IPRATICO = SI → vino esportabile verso iPratico (cassa).", None),
        ("", None),
        ("Stati codificati (vedi foglio 'Riferimento valori' per la legenda completa):", Font(bold=True, size=11)),
        ("• STATO_VENDITA: N/T/V/F/S/C", None),
        ("• STATO_RIORDINO: D/0/A/X (Non ricomprare = X)", None),
        ("• STATO_CONSERVAZIONE: 1/2/3", None),
        ("", None),
        ("Locazioni iniziali (foglio 'Locazioni' per i nomi configurati):", Font(bold=True, size=11)),
        ("• FRIGORIFERO/QTA_FRIGO — bottiglie nel frigo principale", None),
        ("• LOCAZIONE_1/QTA_LOC1, LOCAZIONE_2/QTA_LOC2, LOCAZIONE_3/QTA_LOC3 — cantina e magazzini", None),
        ("• QTA_TOTALE NON va inserito: è calcolato automaticamente come somma delle 4 locazioni.", None),
        ("", None),
        ("Esempio compilato:", Font(bold=True, size=11)),
        ("Vedi la prima riga del foglio 'Vini'. Eliminala prima di compilare i tuoi dati.", None),
        ("", None),
        ("Round-trip (export ↔ import):", Font(bold=True, size=11)),
        ("Esporta tutti i vini da Impostazioni → Esporta tutto, modifica nel foglio 'Vini',", None),
        ("reimporta. I vini con ID già presente vengono saltati (no overwrite); le righe", None),
        ("nuove (ID vuoto) vengono inserite. Per modifiche puntuali usa sempre la UI.", None),
    ]
    for r_idx, (txt, fnt) in enumerate(righe, start=1):
        cell = ws.cell(row=r_idx, column=1, value=txt)
        if fnt:
            cell.font = fnt
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def generate_template_xlsx(include_esempio: bool = True) -> bytes:
    """
    Genera il file template `.xlsx` vuoto (con o senza riga esempio).
    Fogli: Vini, Locazioni, Riferimento valori, Istruzioni.
    """
    wb = Workbook()
    ws_vini = wb.active
    ws_vini.title = "Vini"
    _scrivi_foglio_vini(ws_vini, righe_dati=None if include_esempio else [])

    ws_loc = wb.create_sheet("Locazioni")
    _scrivi_foglio_locazioni(ws_loc, _get_locazioni_config())

    ws_ref = wb.create_sheet("Riferimento valori")
    _scrivi_foglio_riferimento(ws_ref)

    ws_ist = wb.create_sheet("Istruzioni")
    _scrivi_foglio_istruzioni(ws_ist)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_export_xlsx() -> bytes:
    """
    Esporta tutti i vini del DB nel formato template. Round-trip pulito.
    """
    conn = mag_db.get_magazzino_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    rows = cur.execute(
        f"""
        SELECT {', '.join(c['field'] for c in TEMPLATE_COLUMNS)}
        FROM vini_magazzino
        ORDER BY TIPOLOGIA, NAZIONE, REGIONE, PRODUTTORE, DESCRIZIONE
        """
    ).fetchall()
    conn.close()

    vini = [dict(r) for r in rows]

    wb = Workbook()
    ws_vini = wb.active
    ws_vini.title = "Vini"
    _scrivi_foglio_vini(ws_vini, righe_dati=vini)

    ws_loc = wb.create_sheet("Locazioni")
    _scrivi_foglio_locazioni(ws_loc, _get_locazioni_config())

    ws_ref = wb.create_sheet("Riferimento valori")
    _scrivi_foglio_riferimento(ws_ref)

    ws_ist = wb.create_sheet("Istruzioni")
    _scrivi_foglio_istruzioni(ws_ist)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ============================================================
# IMPORT
# ============================================================

class ImportResult(dict):
    """Risultato dell'import: contatori + dettaglio errori."""


def parse_import_xlsx(file_bytes: bytes) -> ImportResult:
    """
    Parsa un file Excel nel formato template e inserisce i vini nuovi.

    Strategia:
      - Riga con ID vuoto       → INSERT nuovo vino
      - Riga con ID popolato:
          - se ID esiste in DB  → SKIP (no overwrite)
          - se ID non esiste    → ERRORE riga (dati ambigui)
      - Validazione minima per riga: TIPOLOGIA + DESCRIZIONE + NAZIONE obbligatori
      - Valori invalidi (es. TIPOLOGIA non in lista) → ERRORE riga
      - Flag SI/NO → 0/1 al DB
    """
    result: ImportResult = ImportResult({
        "inseriti": 0,
        "saltati_esistenti": 0,
        "errori": 0,
        "dettaglio_errori": [],
    })

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        result["errori"] = 1
        result["dettaglio_errori"].append({"riga": 0, "errore": f"File non valido: {e}"})
        return result

    if "Vini" not in wb.sheetnames:
        result["errori"] = 1
        result["dettaglio_errori"].append({"riga": 0, "errore": "Foglio 'Vini' mancante. Usa il template ufficiale."})
        return result

    ws = wb["Vini"]
    # Header: prima riga
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_to_idx = {h: i for i, h in enumerate(header) if h}

    # Verifica header minimo
    for c in TEMPLATE_COLUMNS:
        if c["obbligatorio"] and c["col"] not in col_to_idx:
            result["errori"] = 1
            result["dettaglio_errori"].append({"riga": 1, "errore": f"Colonna obbligatoria mancante: {c['col']}"})
            return result

    # Carica ID esistenti per check rapido
    conn = mag_db.get_magazzino_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    existing_ids = {r[0] for r in cur.execute("SELECT id FROM vini_magazzino").fetchall()}
    conn.close()

    # Itera righe dati (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip righe completamente vuote
        if not any(v not in (None, "") for v in row):
            continue

        def _val(col_name):
            i = col_to_idx.get(col_name)
            return row[i] if i is not None and i < len(row) else None

        # Skip riga esempio (la riconosciamo dal commento nelle NOTE)
        note = _val("NOTE")
        if note and "riga di esempio" in str(note).lower():
            continue

        # ID check
        id_raw = _safe_int(_val("ID"))
        if id_raw is not None:
            if id_raw in existing_ids:
                result["saltati_esistenti"] += 1
                continue
            else:
                result["errori"] += 1
                result["dettaglio_errori"].append({
                    "riga": row_idx,
                    "errore": f"ID {id_raw} specificato ma non esiste nel sistema. Lascia ID vuoto per nuovo vino.",
                })
                continue

        # Costruisco il dict dei campi
        data: Dict[str, Any] = {}
        riga_errore = None

        for c in TEMPLATE_COLUMNS:
            if c["field"] == "id":
                continue  # gestito sopra
            raw = _val(c["col"])
            tipo = c["tipo"]

            if tipo == "str_req":
                if raw is None or str(raw).strip() == "":
                    riga_errore = f"Campo obbligatorio '{c['col']}' vuoto"
                    break
                data[c["field"]] = str(raw).strip()
            elif tipo == "str_opt":
                data[c["field"]] = str(raw).strip() if raw is not None and str(raw).strip() != "" else None
            elif tipo == "enum":
                if raw is None or str(raw).strip() == "":
                    if c["obbligatorio"]:
                        riga_errore = f"Campo obbligatorio '{c['col']}' vuoto"
                        break
                    data[c["field"]] = None
                else:
                    val = str(raw).strip()
                    options = c.get("options", [])
                    if val not in options:
                        riga_errore = f"Valore '{val}' non valido per '{c['col']}'. Validi: {', '.join(options)}"
                        break
                    data[c["field"]] = val
            elif tipo == "flag_yn":
                conv = _yn_to_int(raw)
                data[c["field"]] = conv if conv is not None else 0
            elif tipo in ("num",):
                data[c["field"]] = _safe_float(raw)
            elif tipo in ("int", "qta"):
                v = _safe_int(raw)
                data[c["field"]] = v if v is not None else (0 if tipo == "qta" else None)

        if riga_errore:
            result["errori"] += 1
            result["dettaglio_errori"].append({"riga": row_idx, "errore": riga_errore})
            continue

        # ORIGINE automatica per import
        data["ORIGINE"] = "MANUALE"

        # Insert
        try:
            new_id = mag_db.create_vino(data)
            existing_ids.add(new_id)
            result["inseriti"] += 1
        except Exception as e:
            result["errori"] += 1
            result["dettaglio_errori"].append({"riga": row_idx, "errore": f"Insert fallito: {e}"})

    return result
