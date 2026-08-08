# Modulo: clienti
# Classificazione: [core] — il parser e' generico (colonne configurabili),
# le regole di scarto sono decisioni di Marco documentate qui sotto.

# @version: v1.0-giftcard-import
# -*- coding: utf-8 -*-
"""
Import gift card dal foglio Excel storico.

CONTESTO (2026-08-08): fino a oggi i buoni erano tenuti in un Excel
(`gift-card-lista.xlsx`, foglio GIFT-CARD, 174 righe dal dic 2021).
Il file e' compilato a mano in cinque anni, quindi e' disomogeneo:
date mancanti o impossibili (`29/02/2023`), importi solo dentro la
descrizione (`deg 130`), codici ripetuti, una data scritta `20/'5/2'23`,
la colonna Utente che contiene a volte un nome, a volte un telefono,
a volte una data.

Questo modulo NON scrive su DB: legge, normalizza e restituisce un piano
(`RigaImport` importabili + scarti motivati). Chi chiama decide se
applicarlo. Cosi' la UI puo' mostrare l'anteprima prima di toccare i dati,
e la logica e' testabile senza database.

──────────────────────────────────────────────────────────────
REGOLE DI IMPORT (decise da Marco, 2026-08-08)
──────────────────────────────────────────────────────────────
1. ANNO DAL CODICE. Il codice segue lo schema <lettera>1<AA>-<progressivo>:
   `A125-330` = serie 2025, progressivo 330. L'anno del CODICE vince sulla
   colonna Data, perche' e' quello il criterio con cui i buoni sono stati
   numerati: la serie A124 e' stata aperta a dicembre 2023 per i regali di
   Natale, quindi 26 buoni hanno data 2023 ma appartengono alla stagione
   2024. Sui codici fuori schema (serie vecchie `N###`, `M###`, `A000-*`,
   `####-####`) si usa l'anno della colonna Data.
2. SOGLIA 2024. Tutto cio' che e' anteriore al 2024 resta fuori: sono buoni
   di 4-5 anni fa, restano consultabili nell'Excel.
3. IMPORTO OBBLIGATORIO. Si prende dalla colonna Importo; se vuota si prova
   a dedurlo dalla descrizione (`deg 130`, `VALORE 100€`). Se non si deduce,
   la riga NON entra: una gift card senza valore non e' verificabile al
   banco e sporcherebbe il totale del valore in circolazione.
4. CODICI DOPPI. A parita' di codice normalizzato vince la riga con
   l'importo PIU' ALTO; l'altra finisce nelle note. E' la scelta prudente
   verso il cliente che si presenta con quel buono in mano.
5. SENZA SCADENZA. I buoni importati entrano con `data_scadenza = NULL`,
   come erano nell'Excel. La scadenza di default vale solo per i nuovi.
"""

from __future__ import annotations

import datetime
import logging
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger("trgb.clienti.giftcard.import")

ANNO_MINIMO = 2024

# <lettera>1<AA>-<progressivo> → A125-330, B126-350, a123-198
RE_CODICE_SERIE = re.compile(r"^([A-Za-z])1(\d{2})-(\d+)")

# Importo dentro la descrizione: "deg 130", "VALORE 100€", "2 deg da 60"
RE_IMPORTO_DESCR = re.compile(
    r"(\d{2,4})\s*(?:€|EUR)"
    r"|(?:€|EUR)\s*(\d{2,4})"
    r"|\b(?:deg|degustazion\w*|valore|sconto|buono|da)\D{0,12}?(\d{2,4})\b",
    re.IGNORECASE,
)
IMPORTO_MIN, IMPORTO_MAX = 20, 1000

# Nomi noti nella colonna Utente. Quello che non e' un nome (telefoni,
# date, "RIF. 6B7F622AE5CB") finisce in nota invece di sparire.
RE_SOLO_CIFRE = re.compile(r"^\+?\d[\d\s./-]*$")

INTESTAZIONI = {
    "scontrino": "codice",
    "data": "data",
    "importo": "importo",
    "descrizione": "descrizione",
    "utente": "utente",
    "usato": "usato",
    "id": "id",
}


@dataclass
class RigaImport:
    riga_excel: int
    codice: str
    tipo: str
    importo: float
    descrizione: Optional[str]
    stato: str
    data_emissione: str
    data_utilizzo: Optional[str]
    emessa_da: Optional[str]
    note: Optional[str]
    avvisi: List[str] = field(default_factory=list)


@dataclass
class RigaScartata:
    riga_excel: int
    codice: str
    motivo: str
    dettaglio: str = ""


@dataclass
class PianoImport:
    importabili: List[RigaImport] = field(default_factory=list)
    scartate: List[RigaScartata] = field(default_factory=list)

    @property
    def totale_valore_attive(self) -> float:
        return round(sum(r.importo for r in self.importabili if r.stato == "attiva"), 2)

    def riepilogo(self) -> Dict[str, Any]:
        attive = [r for r in self.importabili if r.stato == "attiva"]
        motivi: Dict[str, int] = {}
        for s in self.scartate:
            motivi[s.motivo] = motivi.get(s.motivo, 0) + 1
        return {
            "importabili": len(self.importabili),
            "attive": len(attive),
            "usate": len(self.importabili) - len(attive),
            "valore_attive": self.totale_valore_attive,
            "scartate": len(self.scartate),
            "scartate_per_motivo": motivi,
            "con_avvisi": sum(1 for r in self.importabili if r.avvisi),
        }


# ─────────────────────────────────────────────────────────────
# Normalizzatori
# ─────────────────────────────────────────────────────────────

def normalizza_codice(codice: str) -> str:
    """Chiave di confronto: solo alfanumerici maiuscoli."""
    return "".join(ch for ch in (codice or "").upper() if ch.isalnum())


def anno_dal_codice(codice: str) -> Optional[int]:
    m = RE_CODICE_SERIE.match((codice or "").strip())
    if not m:
        return None
    anno = 2000 + int(m.group(2))
    # Guardia contro falsi positivi tipo "A199-..." o serie future assurde
    if 2015 <= anno <= datetime.date.today().year + 1:
        return anno
    return None


def parse_data(valore: Any) -> Tuple[Optional[datetime.date], Optional[str]]:
    """
    Ritorna (data, avviso). Tollera datetime, date, e le stringhe scritte a
    mano nel foglio. `29/02/2023` non esiste (2023 non e' bisestile): non la
    inventiamo, torna None con avviso.
    """
    if valore is None or valore == "":
        return None, None
    if isinstance(valore, datetime.datetime):
        return valore.date(), None
    if isinstance(valore, datetime.date):
        return valore, None

    testo = str(valore).strip()
    # Ripulisce apici e spazi sparsi: "20/'5/2'23" → "20/5/223" → gestito sotto
    pulito = testo.replace("'", "").replace(" ", "")
    numeri = re.findall(r"\d+", pulito)
    if len(numeri) >= 3:
        g, m, a = numeri[0], numeri[1], numeri[2]
        # "2'23" diventa "223": se l'anno ha 3 cifre e inizia per 2, e' un 20xx
        if len(a) == 3 and a.startswith("2"):
            a = "20" + a[1:]
        elif len(a) == 2:
            a = "20" + a
        try:
            d = datetime.date(int(a), int(m), int(g))
            return d, f"data '{testo}' interpretata come {d.isoformat()}"
        except ValueError:
            return None, f"data '{testo}' non valida, ignorata"
    return None, f"data '{testo}' non interpretabile, ignorata"


def deduci_importo(descrizione: Any) -> Optional[float]:
    """Importo nascosto nella descrizione. Prende il valore piu' alto
    plausibile: 'DEG 5 PIU 70€ ABBINAMENTO' → 70, non 5."""
    if not descrizione:
        return None
    trovati = [
        int(x)
        for tup in RE_IMPORTO_DESCR.findall(str(descrizione))
        for x in tup
        if x
    ]
    validi = [v for v in trovati if IMPORTO_MIN <= v <= IMPORTO_MAX]
    return float(max(validi)) if validi else None


def _pulisci_utente(valore: Any) -> Tuple[Optional[str], Optional[str]]:
    """
    (emessa_da, nota). La colonna Utente contiene nomi ma anche telefoni,
    date e codici di riferimento: solo i nomi diventano `emessa_da`, il
    resto viene conservato come nota invece di essere buttato.
    """
    if valore is None or valore == "":
        return None, None
    if isinstance(valore, (datetime.datetime, datetime.date)):
        return None, f"colonna Utente conteneva una data: {valore}"
    testo = str(valore).strip()
    if not testo:
        return None, None
    if RE_SOLO_CIFRE.match(testo):
        return None, f"riferimento in colonna Utente: {testo}"
    if testo.upper().startswith("RIF."):
        return None, f"colonna Utente: {testo}"
    return testo, None


# ─────────────────────────────────────────────────────────────
# Parser principale
# ─────────────────────────────────────────────────────────────

def _mappa_colonne(header: Tuple) -> Dict[str, int]:
    mappa: Dict[str, int] = {}
    for i, cella in enumerate(header):
        if cella is None:
            continue
        chiave = INTESTAZIONI.get(str(cella).strip().lower())
        if chiave and chiave not in mappa:
            mappa[chiave] = i
    return mappa


def costruisci_piano(
    righe: List[Tuple],
    anno_minimo: int = ANNO_MINIMO,
) -> PianoImport:
    """
    `righe` = tutte le righe del foglio, header compreso (come le da'
    openpyxl con values_only=True). Non tocca il DB.
    """
    piano = PianoImport()
    if not righe:
        return piano

    col = _mappa_colonne(righe[0])
    mancanti = {"codice", "data", "importo"} - set(col)
    if mancanti:
        raise ValueError(
            f"Colonne non trovate nel foglio: {', '.join(sorted(mancanti))}. "
            f"Attese le intestazioni Scontrino / Data / Importo."
        )

    def val(riga: Tuple, chiave: str):
        i = col.get(chiave)
        return riga[i] if i is not None and i < len(riga) else None

    # Colonne senza intestazione: nel foglio storico contengono note sparse
    # (un telefono, un nome, un appunto). Le raccogliamo invece di perderle.
    indici_noti = set(col.values())
    indici_extra = [i for i in range(len(righe[0])) if i not in indici_noti]

    candidate: List[RigaImport] = []

    for numero, riga in enumerate(righe[1:], start=2):
        if not any(c not in (None, "") for c in riga):
            continue

        codice = str(val(riga, "codice") or "").strip()
        if not codice:
            piano.scartate.append(RigaScartata(numero, "", "codice mancante"))
            continue

        avvisi: List[str] = []
        note: List[str] = []

        # ── importo (regola 3) ──
        grezzo = val(riga, "importo")
        descrizione = val(riga, "descrizione")
        if isinstance(grezzo, (int, float)):
            importo = float(grezzo)
        else:
            importo = deduci_importo(descrizione)
            if importo is not None:
                avvisi.append(f"importo {importo:.0f}€ dedotto dalla descrizione")
        if importo is None:
            piano.scartate.append(RigaScartata(
                numero, codice, "senza importo",
                f"descrizione: {str(descrizione or '')[:40] or '(vuota)'}",
            ))
            continue

        # ── anno (regola 1) ──
        data_em, avviso_data = parse_data(val(riga, "data"))
        if avviso_data:
            avvisi.append(avviso_data)
        anno_cod = anno_dal_codice(codice)
        anno = anno_cod or (data_em.year if data_em else None)
        if anno is None:
            piano.scartate.append(RigaScartata(
                numero, codice, "anno non determinabile",
                "ne' dal codice ne' dalla data",
            ))
            continue
        if anno_cod and data_em and data_em.year != anno_cod:
            avvisi.append(
                f"il codice dice {anno_cod}, la data dice {data_em.year}: "
                f"vale il codice"
            )

        # ── soglia (regola 2) ──
        if anno < anno_minimo:
            piano.scartate.append(RigaScartata(
                numero, codice, f"anteriore al {anno_minimo}", f"serie {anno}",
            ))
            continue

        # ── data di emissione ──
        if data_em is None:
            data_em = datetime.date(anno, 1, 1)
            avvisi.append(f"data di emissione non nota, messo 01/01/{anno} (anno del codice)")
        data_emissione = data_em.isoformat()

        # ── stato ──
        data_uso, avviso_uso = parse_data(val(riga, "usato"))
        usato_grezzo = val(riga, "usato")
        if data_uso:
            stato, data_utilizzo = "usata", data_uso.isoformat()
            if avviso_uso:
                avvisi.append(avviso_uso)
        elif usato_grezzo not in (None, ""):
            # C'e' scritto qualcosa che non e' una data: la card e' stata
            # usata, ma non sappiamo quando. Meglio "usata senza data" che
            # rimetterla in circolo per errore.
            stato, data_utilizzo = "usata", None
            avvisi.append(f"segnata usata ma la data non e' leggibile ({usato_grezzo!r})")
        else:
            stato, data_utilizzo = "attiva", None

        # ── autore + note ──
        emessa_da, nota_utente = _pulisci_utente(val(riga, "utente"))
        if nota_utente:
            note.append(nota_utente)
        for i in indici_extra:
            if i < len(riga) and riga[i] not in (None, ""):
                note.append(str(riga[i]).strip())

        descrizione_pulita = str(descrizione).strip() if descrizione else None
        if isinstance(descrizione, (datetime.datetime, datetime.date)):
            note.append(f"colonna Descrizione conteneva una data: {descrizione}")
            descrizione_pulita = None

        note.append(f"Importata dall'Excel storico (riga {numero})")

        candidate.append(RigaImport(
            riga_excel=numero,
            codice=codice,
            # Il buono storico descrive quasi sempre un'esperienza ("2 deg
            # da 5 portate") ma ha anche un valore incassato: teniamo
            # entrambi, il PDF mostrera' solo la descrizione.
            tipo="esperienza" if descrizione_pulita else "valore",
            importo=importo,
            descrizione=descrizione_pulita,
            stato=stato,
            data_emissione=data_emissione,
            data_utilizzo=data_utilizzo,
            emessa_da=emessa_da,
            note=" · ".join(note) if note else None,
            avvisi=avvisi,
        ))

    # ── codici doppi (regola 4): vince l'importo piu' alto ──
    per_codice: Dict[str, List[RigaImport]] = {}
    for r in candidate:
        per_codice.setdefault(normalizza_codice(r.codice), []).append(r)

    for _, gruppo in per_codice.items():
        if len(gruppo) == 1:
            piano.importabili.append(gruppo[0])
            continue
        gruppo.sort(key=lambda r: (-r.importo, r.riga_excel))
        vincitrice, perdenti = gruppo[0], gruppo[1:]
        dettagli = "; ".join(
            f"riga {p.riga_excel}: {p.importo:.0f}€ {p.descrizione or ''}".strip()
            for p in perdenti
        )
        vincitrice.avvisi.append(
            f"codice doppio nell'Excel: tenuta questa ({vincitrice.importo:.0f}€), "
            f"scartata {dettagli}"
        )
        vincitrice.note = f"{vincitrice.note} · Riga doppia scartata → {dettagli}"
        piano.importabili.append(vincitrice)
        for p in perdenti:
            piano.scartate.append(RigaScartata(
                p.riga_excel, p.codice, "codice doppio",
                f"tenuta la riga {vincitrice.riga_excel} da {vincitrice.importo:.0f}€",
            ))

    piano.importabili.sort(key=lambda r: r.riga_excel)
    piano.scartate.sort(key=lambda r: r.riga_excel)
    return piano


def leggi_xlsx(contenuto: bytes, foglio: Optional[str] = None) -> List[Tuple]:
    """Estrae le righe dal file. Se `foglio` non e' indicato usa il primo
    che contiene una colonna 'Scontrino', altrimenti il primo del file."""
    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(contenuto), data_only=True, read_only=True)
    ws = None
    if foglio and foglio in wb.sheetnames:
        ws = wb[foglio]
    else:
        for candidato in wb.worksheets:
            intestazioni = next(candidato.iter_rows(values_only=True), ())
            if any(str(c).strip().lower() == "scontrino" for c in intestazioni if c):
                ws = candidato
                break
        ws = ws or wb.worksheets[0]
    return list(ws.iter_rows(values_only=True))
