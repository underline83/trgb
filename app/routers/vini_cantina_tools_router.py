# @version: v3.1-unified-loader
# -*- coding: utf-8 -*-
"""
Tre Gobbi — Router Cantina Tools
File: app/routers/vini_cantina_tools_router.py

v3.0: eliminato sync-from-excel (vecchio DB vini.sqlite3 non più usato)

Strumenti:

1. POST /vini/cantina-tools/import-excel
   → Import diretto di un file Excel nel DB cantina (senza passare
     per vini.sqlite3). Usa la stessa logica di normalizzazione.

2. GET /vini/cantina-tools/export-excel
   → Scarica un .xlsx dal DB cantina, formato compatibile con l'Excel
     storico di lavoro.

3. GET /vini/cantina-tools/carta-cantina
   → Genera la carta vini HTML leggendo dal DB cantina
4. GET /vini/cantina-tools/carta-cantina/pdf
   → Genera il PDF della carta leggendo dal DB cantina
5. GET /vini/cantina-tools/carta-cantina/docx
   → Genera il DOCX della carta leggendo dal DB cantina

Tutti gli endpoint richiedono autenticazione; import solo admin.
"""

from __future__ import annotations

import os
import re
import tempfile
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional
from itertools import groupby

import pandas as pd
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Query, Request, status
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from weasyprint import HTML, CSS

from app.services.auth_service import get_current_user, decode_access_token
from app.models import vini_magazzino_db as mag_db
from app.models.vini_model import normalize_dataframe
from app.services.carta_vini_service import (
    build_carta_body_html,
    build_carta_body_html_htmlsafe,
    build_carta_toc_html,
    build_carta_docx,
    resolve_regione,
)
from app.repositories.vini_repository import load_vini_ordinati, _load_ordinamenti
from app.models.vini_settings import _TIPOLOGIA_MAP


router = APIRouter(
    prefix="/vini/cantina-tools",
    tags=["Vini Cantina Tools"],
)

# PATH DI BASE
BASE_DIR = Path(__file__).resolve().parents[2]
STATIC_DIR = BASE_DIR / "static"
CSS_HTML = STATIC_DIR / "css" / "carta_html.css"
CSS_PDF = STATIC_DIR / "css" / "carta_pdf.css"
LOGO_PATH = STATIC_DIR / "img" / "logo_tregobbi.png"


# ---------------------------------------------------------
# HELPER: verifica ruolo admin
# ---------------------------------------------------------
def _require_admin(current_user: Any):
    role = None
    if isinstance(current_user, dict):
        role = current_user.get("role")
    elif hasattr(current_user, "role"):
        role = current_user.role
    if role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo admin può eseguire questa operazione.",
        )


def _get_user_from_query_token(token: Optional[str] = Query(None)) -> Any:
    """
    Autenticazione opzionale via query parameter ?token=...
    Utile per i download via window.open() dove non si può passare l'header.
    """
    if not token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token mancante. Passa ?token=... nell'URL.",
        )
    return decode_access_token(token)


def _get_username(current_user: Any) -> str:
    if isinstance(current_user, dict):
        return current_user.get("username") or current_user.get("sub") or "unknown"
    for attr in ("username", "sub"):
        if hasattr(current_user, attr):
            val = getattr(current_user, attr)
            if val:
                return str(val)
    return "unknown"




# =============================================================
# 0. RESET DATABASE CANTINA
# =============================================================
@router.post("/reset-database", summary="Azzera completamente il DB cantina")
def reset_database(
    current_user: Any = Depends(get_current_user),
):
    """
    Elimina TUTTI i dati dal DB cantina: vini, movimenti e note.
    Operazione irreversibile — solo admin.
    """
    _require_admin(current_user)

    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()

    # Conta prima di cancellare (per il report)
    n_vini = cur.execute("SELECT COUNT(*) FROM vini_magazzino").fetchone()[0]
    n_mov = cur.execute("SELECT COUNT(*) FROM vini_magazzino_movimenti").fetchone()[0]
    n_note = cur.execute("SELECT COUNT(*) FROM vini_magazzino_note").fetchone()[0]

    # Svuota le tabelle (ordine: figlie prima, poi padre)
    cur.execute("DELETE FROM vini_magazzino_note;")
    cur.execute("DELETE FROM vini_magazzino_movimenti;")
    cur.execute("DELETE FROM vini_magazzino;")

    # Reset autoincrement
    cur.execute("DELETE FROM sqlite_sequence WHERE name IN ('vini_magazzino', 'vini_magazzino_movimenti', 'vini_magazzino_note');")

    conn.commit()
    conn.close()

    return {
        "status": "ok",
        "msg": f"Database cantina azzerato: {n_vini} vini, {n_mov} movimenti, {n_note} note eliminati.",
        "eliminati": {"vini": n_vini, "movimenti": n_mov, "note": n_note},
    }


# =============================================================
# 1. IMPORT DIRETTO EXCEL → CANTINA
# =============================================================
@router.post("/import-excel", summary="Import diretto Excel → Cantina")
async def import_excel_to_cantina(
    file: UploadFile = File(...),
    current_user: Any = Depends(get_current_user),
):
    """
    Importa un file Excel direttamente nel DB cantina (senza passare
    per vini.sqlite3). Usa normalize_dataframe per compatibilità.
    I vini con id già presente vengono aggiornati (upsert su id_excel).
    I vini nuovi vengono inseriti con giacenze dall'Excel.
    """
    _require_admin(current_user)

    # Salva file temporaneo
    suffix = Path(file.filename or "upload.xlsx").suffix
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        content = await file.read()
        tmp.write(content)
        tmp.close()

        df = pd.read_excel(tmp.name, sheet_name="VINI")
        df = df.dropna(how="all")
        df = normalize_dataframe(df)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Errore lettura Excel: {e}",
        )
    finally:
        os.unlink(tmp.name)

    # Pre-carica mappa di vini esistenti per match veloce
    conn_mag = mag_db.get_magazzino_connection()
    cur_mag = conn_mag.cursor()
    # Mappa: chiave naturale → id cantina (per match robusto)
    natural_key_map = {}
    for vrow in cur_mag.execute(
        "SELECT id, DESCRIZIONE, PRODUTTORE, ANNATA, FORMATO FROM vini_magazzino;"
    ):
        nk = (
            (vrow["DESCRIZIONE"] or "").strip().upper(),
            (vrow["PRODUTTORE"] or "").strip().upper(),
            (vrow["ANNATA"] or "").strip(),
            (vrow["FORMATO"] or "").strip(),
        )
        natural_key_map[nk] = vrow["id"]
    conn_mag.close()

    inseriti = 0
    aggiornati = 0
    errori = []

    for ridx, row in df.iterrows():
        try:
            data = {
                "TIPOLOGIA": row.get("TIPOLOGIA"),
                "NAZIONE": row.get("NAZIONE"),
                "REGIONE": row.get("REGIONE"),
                "DESCRIZIONE": row.get("DESCRIZIONE"),
                "DENOMINAZIONE": row.get("DENOMINAZIONE"),
                "ANNATA": row.get("ANNATA"),
                "FORMATO": row.get("FORMATO"),
                "PRODUTTORE": row.get("PRODUTTORE"),
                "DISTRIBUTORE": row.get("DISTRIBUTORE"),
                "PREZZO_CARTA": row.get("PREZZO"),
                "EURO_LISTINO": row.get("EURO_LISTINO"),
                "SCONTO": row.get("SCONTO"),
                "CARTA": row.get("CARTA"),
                "IPRATICO": row.get("IPRATICO"),
                "FRIGORIFERO": row.get("FRIGORIFERO"),
                "LOCAZIONE_1": row.get("LOCAZIONE_1"),
                "LOCAZIONE_2": row.get("LOCAZIONE_2"),
                "ORIGINE": "EXCEL",
            }

            # Pulizia None/NaN
            for k, v in data.items():
                if pd.isna(v) if isinstance(v, float) else False:
                    data[k] = None

            # Se DESCRIZIONE è necessaria
            if not data.get("DESCRIZIONE"):
                continue
            if not data.get("TIPOLOGIA"):
                data["TIPOLOGIA"] = "ERRORE"
            if not data.get("NAZIONE"):
                data["NAZIONE"] = "SCONOSCIUTA"

            # Cerca match per chiave naturale (DESCRIZIONE+PRODUTTORE+ANNATA+FORMATO)
            nk = (
                (data["DESCRIZIONE"] or "").strip().upper(),
                (data.get("PRODUTTORE") or "").strip().upper(),
                (data.get("ANNATA") or "").strip(),
                (data.get("FORMATO") or "").strip(),
            )
            existing_id = natural_key_map.get(nk)

            if existing_id:
                # Aggiorna il vino esistente (senza toccare giacenze)
                update_data = {
                    k: v for k, v in data.items()
                    if k not in ("QTA_FRIGO", "QTA_LOC1", "QTA_LOC2", "QTA_TOTALE")
                    and v is not None
                }
                mag_db.update_vino(existing_id, update_data)
                aggiornati += 1
            else:
                # Nuovo: importa con giacenze
                data["QTA_FRIGO"] = int(row.get("N_FRIGO", 0) or 0)
                data["QTA_LOC1"] = int(row.get("N_LOC1", 0) or 0)
                data["QTA_LOC2"] = int(row.get("N_LOC2", 0) or 0)
                data["QTA_TOTALE"] = int(row.get("QTA", 0) or 0)
                # Serve almeno una locazione per create_vino
                if not any([data.get("FRIGORIFERO"), data.get("LOCAZIONE_1"), data.get("LOCAZIONE_2")]):
                    data["FRIGORIFERO"] = "Frigo"
                new_id = mag_db.create_vino(data)
                # Aggiungi alla mappa per evitare doppi inserimenti nello stesso import
                natural_key_map[nk] = new_id
                inseriti += 1

        except Exception as e:
            desc = row.get("DESCRIZIONE") or ""
            prod = row.get("PRODUTTORE") or ""
            errori.append(f"riga {ridx}: {desc} ({prod}): {e}")

    return {
        "status": "ok",
        "righe_excel": len(df),
        "inseriti": inseriti,
        "aggiornati": aggiornati,
        "errori": errori,
        "msg": f"Import completato: {inseriti} nuovi, {aggiornati} aggiornati"
        + (f", {len(errori)} errori" if errori else ""),
    }


# =============================================================
# 2b. PULIZIA DUPLICATI
# =============================================================
@router.post("/cleanup-duplicates", summary="Rimuovi vini duplicati dalla cantina")
def cleanup_duplicates(
    dry_run: bool = Query(
        True,
        description="Se true, mostra solo i duplicati senza eliminarli",
    ),
    current_user: Any = Depends(get_current_user),
):
    """
    Trova e rimuove vini duplicati nella cantina.
    Duplicati = stessa DESCRIZIONE + PRODUTTORE + ANNATA + FORMATO (case-insensitive).
    Mantiene il record con id più basso (il primo inserito).
    """
    _require_admin(current_user)

    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()

    # Trova gruppi di duplicati
    groups = cur.execute("""
        SELECT
            UPPER(TRIM(COALESCE(DESCRIZIONE, ''))) AS dk,
            UPPER(TRIM(COALESCE(PRODUTTORE, '')))  AS pk,
            TRIM(COALESCE(ANNATA, ''))              AS ak,
            TRIM(COALESCE(FORMATO, ''))             AS fk,
            COUNT(*) AS cnt,
            GROUP_CONCAT(id, ',') AS ids
        FROM vini_magazzino
        GROUP BY dk, pk, ak, fk
        HAVING cnt > 1
        ORDER BY cnt DESC, dk
    """).fetchall()

    duplicati = []
    ids_to_delete = []

    for g in groups:
        ids_list = sorted(int(x) for x in g["ids"].split(","))
        keep_id = ids_list[0]
        delete_ids = ids_list[1:]
        ids_to_delete.extend(delete_ids)
        duplicati.append({
            "descrizione": g["dk"],
            "produttore": g["pk"],
            "annata": g["ak"],
            "formato": g["fk"],
            "copie": g["cnt"],
            "keep_id": keep_id,
            "delete_ids": delete_ids,
        })

    eliminati = 0
    if not dry_run and ids_to_delete:
        placeholders = ",".join("?" * len(ids_to_delete))
        cur.execute(f"DELETE FROM vini_magazzino WHERE id IN ({placeholders})", ids_to_delete)
        conn.commit()
        eliminati = len(ids_to_delete)

    conn.close()

    return {
        "status": "ok",
        "dry_run": dry_run,
        "gruppi_duplicati": len(duplicati),
        "vini_da_eliminare": len(ids_to_delete),
        "eliminati": eliminati,
        "duplicati": duplicati[:50],  # limita output
        "msg": (
            f"Trovati {len(duplicati)} gruppi ({len(ids_to_delete)} copie). "
            + (f"Eliminati {eliminati} duplicati." if not dry_run else "Dry-run: nessuna modifica.")
        ),
    }


# =============================================================
# 3. EXPORT CANTINA → EXCEL
# =============================================================
@router.get("/export-excel", summary="Esporta cantina in Excel")
def export_cantina_excel(
    current_user: Any = Depends(_get_user_from_query_token),
):
    """
    Genera un .xlsx con tutti i vini dal DB cantina,
    in formato compatibile con l'Excel storico di lavoro.
    """
    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()
    rows = cur.execute(
        """
        SELECT *
        FROM vini_magazzino
        ORDER BY TIPOLOGIA, NAZIONE, REGIONE, PRODUTTORE, DESCRIZIONE, ANNATA
        """
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "VINI"

    # Header
    headers = [
        "ID", "TIPOLOGIA", "NAZIONE", "REGIONE",
        "CARTA", "DESCRIZIONE", "ANNATA", "PRODUTTORE",
        "DENOMINAZIONE", "FORMATO",
        "PREZZO", "LISTINO", "SCONTO",
        "FRIGORIFERO", "N", "LOCAZIONE 1", "N.1", "LOCAZIONE 2", "N.2",
        "Q.TA", "DISTRIBUTORE", "IPRATICO", "ORIGINE",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="6B2D5B", end_color="6B2D5B", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Dati
    for row_idx, r in enumerate(rows, 2):
        values = [
            r["id"],
            r["TIPOLOGIA"],
            r["NAZIONE"],
            r["REGIONE"],
            r["CARTA"],
            r["DESCRIZIONE"],
            r["ANNATA"],
            r["PRODUTTORE"],
            r["DENOMINAZIONE"],
            r["FORMATO"],
            r["PREZZO_CARTA"],
            r["EURO_LISTINO"],
            r["SCONTO"],
            r["FRIGORIFERO"],
            r["QTA_FRIGO"],
            r["LOCAZIONE_1"],
            r["QTA_LOC1"],
            r["LOCAZIONE_2"],
            r["QTA_LOC2"],
            r["QTA_TOTALE"],
            r["DISTRIBUTORE"],
            r["IPRATICO"],
            r["ORIGINE"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Freeze header
    ws.freeze_panes = "A2"

    # Salva
    out_path = STATIC_DIR / "cantina_export.xlsx"
    wb.save(str(out_path))

    return FileResponse(
        str(out_path),
        filename=f"cantina_vini_{datetime.now().strftime('%Y%m%d')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =============================================================
# 4. CARTA DA CANTINA — HTML
# =============================================================
@router.get("/carta-cantina", summary="Carta vini HTML da DB cantina")
def carta_cantina_html():
    """
    Genera la carta dei vini leggendo dal DB magazzino.
    Usa load_vini_ordinati() condiviso con /vini/carta.
    """
    rows = load_vini_ordinati()
    body_html = build_carta_body_html_htmlsafe(rows)

    css_content = ""
    if CSS_HTML.exists():
        css_content = f"<style>{CSS_HTML.read_text()}</style>"

    html_doc = f"""<!DOCTYPE html>
<html lang="it">
<head>
    <meta charset="utf-8">
    <title>OSTERIA TRE GOBBI — CARTA DEI VINI (da Cantina)</title>
    {css_content}
</head>
<body>
    <div class="carta-container">
        {body_html}
    </div>
</body>
</html>"""

    return HTMLResponse(content=html_doc)


# =============================================================
# 5. CARTA DA CANTINA — PDF
# =============================================================
@router.get("/carta-cantina/pdf", summary="Carta vini PDF da DB cantina")
def carta_cantina_pdf(
    current_user: Any = Depends(_get_user_from_query_token),
):
    """
    Genera il PDF della carta vini dal DB cantina.
    """
    rows = load_vini_ordinati()
    data_oggi = datetime.now().strftime("%d/%m/%Y")

    # Front page — identico al vecchio sistema (stesse classi CSS)
    frontespizio = f"""
    <div class="front-page">
        <img src="file://{LOGO_PATH}" class="front-logo">
        <div class="front-title">CARTA VINI</div>
        <div class="front-subtitle">Aggiornata al {data_oggi}</div>
    </div>
    """

    toc_html = build_carta_toc_html(rows)
    body_html = build_carta_body_html(rows)
    carta = f"<div class='carta-body'>{body_html}</div>"

    full_html = f"""
    <html>
    <head>
        <meta charset='utf-8'>
        <link rel='stylesheet' href='/static/css/carta_pdf.css'>
    </head>
    <body>
        {frontespizio}
        {toc_html}
        {carta}
    </body>
    </html>
    """

    out_path = STATIC_DIR / "carta_vini_cantina.pdf"

    HTML(string=full_html, base_url=str(BASE_DIR)).write_pdf(
        str(out_path),
        stylesheets=[CSS(filename=str(CSS_PDF))],
    )

    return FileResponse(
        str(out_path),
        filename=f"carta_vini_cantina_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type="application/pdf",
    )


# =============================================================
# 6. CARTA DA CANTINA — DOCX
# =============================================================
@router.get("/carta-cantina/docx", summary="Carta vini Word da DB cantina")
def carta_cantina_docx(
    current_user: Any = Depends(_get_user_from_query_token),
):
    """
    Genera il DOCX della carta vini dal DB magazzino.
    Usa build_carta_docx() condiviso con /vini/carta/docx.
    """
    rows = load_vini_ordinati()
    doc = build_carta_docx(rows, logo_path=LOGO_PATH)

    out_path = STATIC_DIR / "carta_vini_cantina.docx"
    doc.save(str(out_path))

    return FileResponse(
        str(out_path),
        filename=f"carta_vini_cantina_{datetime.now().strftime('%Y%m%d')}.docx",
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


# =============================================================
# 7. INVENTARIO PDF — TUTTI I VINI
# =============================================================
def _load_all_vini_inventario(
    solo_giacenza: bool = False,
    tipologia: Optional[str] = None,
    nazione: Optional[str] = None,
    regione: Optional[str] = None,
    produttore: Optional[str] = None,
    annata: Optional[str] = None,
    formato: Optional[str] = None,
    carta: Optional[str] = None,
    stato_vendita: Optional[str] = None,
    stato_riordino: Optional[str] = None,
    stato_conservazione: Optional[str] = None,
    discontinuato: Optional[str] = None,
    qta_min: Optional[int] = None,
    qta_max: Optional[int] = None,
    prezzo_min: Optional[float] = None,
    prezzo_max: Optional[float] = None,
    text: Optional[str] = None,
    locazione: Optional[str] = None,
    frigo_nome: Optional[str] = None,
    frigo_spazio: Optional[str] = None,
    loc1_nome: Optional[str] = None,
    loc1_spazio: Optional[str] = None,
    loc2_nome: Optional[str] = None,
    loc2_spazio: Optional[str] = None,
    loc3_nome: Optional[str] = None,
    loc3_spazio: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Carica vini dal DB cantina per report inventario con filtri componibili.
    Ordinamento identico alla carta vini (tipologia_order, nazioni_order,
    regioni_order dalle impostazioni, poi produttore, descrizione, annata).
    """
    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()

    conditions: List[str] = []
    params: List[Any] = []

    if solo_giacenza:
        conditions.append("QTA_TOTALE > 0")
    if tipologia:
        conditions.append("TIPOLOGIA = ?")
        params.append(tipologia)
    if nazione:
        conditions.append("NAZIONE = ?")
        params.append(nazione)
    if regione:
        conditions.append("REGIONE = ?")
        params.append(regione)
    if produttore:
        conditions.append("PRODUTTORE = ?")
        params.append(produttore)
    if annata:
        conditions.append("ANNATA = ?")
        params.append(annata)
    if formato:
        conditions.append("FORMATO = ?")
        params.append(formato)
    if carta:
        conditions.append("CARTA = ?")
        params.append(carta)
    if stato_vendita:
        conditions.append("STATO_VENDITA = ?")
        params.append(stato_vendita)
    if stato_riordino:
        conditions.append("STATO_RIORDINO = ?")
        params.append(stato_riordino)
    if stato_conservazione:
        conditions.append("STATO_CONSERVAZIONE = ?")
        params.append(stato_conservazione)
    if discontinuato:
        conditions.append("DISCONTINUATO = ?")
        params.append(discontinuato)
    if qta_min is not None:
        conditions.append("QTA_TOTALE >= ?")
        params.append(qta_min)
    if qta_max is not None:
        conditions.append("QTA_TOTALE <= ?")
        params.append(qta_max)
    if prezzo_min is not None:
        conditions.append("PREZZO_CARTA >= ?")
        params.append(prezzo_min)
    if prezzo_max is not None:
        conditions.append("PREZZO_CARTA <= ?")
        params.append(prezzo_max)
    if text:
        conditions.append(
            "(DESCRIZIONE LIKE ? OR PRODUTTORE LIKE ? OR DENOMINAZIONE LIKE ?)"
        )
        like = f"%{text}%"
        params.extend([like, like, like])
    if locazione:
        # formato legacy: "frigo:Frigo 1 - Fila 3" oppure "loc1:Cantina - Scaffale 2"
        if ":" in locazione:
            loc_type, loc_val = locazione.split(":", 1)
            col_map = {"frigo": "FRIGORIFERO", "loc1": "LOCAZIONE_1", "loc2": "LOCAZIONE_2"}
            col_name = col_map.get(loc_type)
            if col_name:
                conditions.append(f"{col_name} = ?")
                params.append(loc_val)

    # Filtri locazione gerarchici (nome + spazio opzionale)
    _hier_filters = [
        ("FRIGORIFERO", frigo_nome, frigo_spazio),
        ("LOCAZIONE_1", loc1_nome, loc1_spazio),
        ("LOCAZIONE_2", loc2_nome, loc2_spazio),
        ("LOCAZIONE_3", loc3_nome, loc3_spazio),
    ]
    for col, nome, spazio in _hier_filters:
        if nome:
            if spazio:
                conditions.append(f"{col} = ?")
                params.append(f"{nome} - {spazio}")
            else:
                # Locazione senza sotto-spazi: match esatto O con prefisso
                conditions.append(f"({col} = ? OR {col} LIKE ?)")
                params.extend([nome, f"{nome} - %"])

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    rows = cur.execute(f"""
        SELECT
            id, TIPOLOGIA, NAZIONE, REGIONE,
            id,
            DESCRIZIONE, DENOMINAZIONE, PRODUTTORE,
            ANNATA, FORMATO,
            PREZZO_CARTA, EURO_LISTINO,
            FRIGORIFERO, QTA_FRIGO,
            LOCAZIONE_1, QTA_LOC1,
            LOCAZIONE_2, QTA_LOC2,
            LOCAZIONE_3, QTA_LOC3,
            QTA_TOTALE,
            CARTA, STATO_VENDITA, STATO_RIORDINO, STATO_CONSERVAZIONE,
            DISCONTINUATO
        FROM vini_magazzino
        {where}
    """, params).fetchall()
    conn.close()

    vini = [dict(r) for r in rows]

    # Normalizza tipologie vecchie → nuove
    for r in vini:
        r["TIPOLOGIA"] = _TIPOLOGIA_MAP.get(r["TIPOLOGIA"], r["TIPOLOGIA"])

    # Ordinamento carta vini: usa le tabelle di impostazioni
    tip_map, naz_map, reg_map = _load_ordinamenti()

    def sort_key(r):
        return (
            tip_map.get(r["TIPOLOGIA"], 9999),
            naz_map.get(r["NAZIONE"], 9999),
            reg_map.get(r.get("REGIONE"), 9999),
            (r["PRODUTTORE"] or "").upper(),
            (r["DESCRIZIONE"] or "").upper(),
            r["ANNATA"] or "",
        )

    vini.sort(key=sort_key)
    return vini


def _inventario_css() -> str:
    """CSS inline per i report inventario — stile leggero, poco inchiostro."""
    return """
    @page {
        size: A4 landscape;
        margin: 12mm 10mm 12mm 10mm;
        @bottom-center {
            content: "Pagina " counter(page) " di " counter(pages);
            font-size: 8px;
            color: #999;
        }
    }
    body {
        font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
        font-size: 9px;
        color: #333;
        margin: 0;
        padding: 0;
    }
    .header {
        text-align: center;
        margin-bottom: 10px;
        border-bottom: 1px solid #999;
        padding-bottom: 6px;
    }
    .header h1 {
        font-size: 16px;
        color: #333;
        margin: 0 0 3px 0;
        font-weight: bold;
    }
    .header .subtitle {
        font-size: 9px;
        color: #777;
    }
    table {
        width: 100%;
        border-collapse: collapse;
        margin-bottom: 8px;
        page-break-inside: auto;
    }
    tr { page-break-inside: avoid; }
    th {
        background: none;
        color: #333;
        font-weight: bold;
        padding: 4px 5px;
        text-align: left;
        font-size: 8px;
        text-transform: uppercase;
        border-bottom: 2px solid #333;
    }
    td {
        padding: 3px 5px;
        border-bottom: 1px solid #ccc;
        vertical-align: top;
    }
    .num { text-align: right; }
    .qta-zero { color: #bbb; }
    .qta-pos { font-weight: bold; }
    .totale-row {
        font-weight: bold;
        border-top: 2px solid #333;
    }
    .totale-row td { padding: 5px; }
    .loc-header {
        font-size: 13px;
        font-weight: bold;
        color: #333;
        margin: 16px 0 6px 0;
        padding: 4px 0;
        border-bottom: 1px solid #999;
        page-break-after: avoid;
    }
    .summary-box {
        display: inline-block;
        border: 1px solid #ccc;
        padding: 5px 14px;
        margin: 2px 6px;
        text-align: center;
    }
    .summary-box .label { font-size: 8px; color: #777; }
    .summary-box .value { font-size: 13px; font-weight: bold; color: #333; }
    .summary-row { text-align: center; margin: 8px 0 12px 0; }
    """


def _fmt_qta(q: int) -> str:
    """Formatta quantita con classe CSS."""
    if q == 0:
        return '<span class="qta-zero">-</span>'
    return f'<span class="qta-pos">{q}</span>'


def _build_inventario_table(vini: List[Dict], show_locations: bool = True) -> str:
    """Costruisce tabella HTML per inventario."""
    loc_cols = ""
    if show_locations:
        loc_cols = """
            <th>Frigo</th><th class="num">Qta</th>
            <th>Loc. 1</th><th class="num">Qta</th>
            <th>Loc. 2</th><th class="num">Qta</th>
            <th>Loc. 3</th><th class="num">Qta</th>
        """

    html = f"""
    <table>
        <thead><tr>
            <th class="num">ID</th>
            <th>Tipologia</th>
            <th>Nazione</th>
            <th>Regione</th>
            <th>Produttore</th>
            <th>Descrizione</th>
            <th>Annata</th>
            <th>Formato</th>
            <th class="num">Prezzo</th>
            <th>S.Vend</th>
            <th>S.Riord</th>
            <th>S.Cons</th>
            {loc_cols}
            <th class="num">Totale</th>
        </tr></thead>
        <tbody>
    """

    tot_bottiglie = 0
    for v in vini:
        qta = v.get("QTA_TOTALE") or 0
        tot_bottiglie += qta
        prezzo = v.get("PREZZO_CARTA")
        prezzo_str = f"&euro; {float(prezzo):.2f}" if prezzo else ""

        loc_tds = ""
        if show_locations:
            loc_tds = f"""
                <td>{v.get('FRIGORIFERO') or ''}</td>
                <td class="num">{_fmt_qta(v.get('QTA_FRIGO') or 0)}</td>
                <td>{v.get('LOCAZIONE_1') or ''}</td>
                <td class="num">{_fmt_qta(v.get('QTA_LOC1') or 0)}</td>
                <td>{v.get('LOCAZIONE_2') or ''}</td>
                <td class="num">{_fmt_qta(v.get('QTA_LOC2') or 0)}</td>
                <td>{v.get('LOCAZIONE_3') or ''}</td>
                <td class="num">{_fmt_qta(v.get('QTA_LOC3') or 0)}</td>
            """

        html += f"""
        <tr>
            <td class="num">{v.get('id') or ''}</td>
            <td>{v.get('TIPOLOGIA') or ''}</td>
            <td>{v.get('NAZIONE') or ''}</td>
            <td>{v.get('REGIONE') or ''}</td>
            <td>{v.get('PRODUTTORE') or ''}</td>
            <td>{v.get('DESCRIZIONE') or ''}</td>
            <td>{v.get('ANNATA') or ''}</td>
            <td>{v.get('FORMATO') or ''}</td>
            <td class="num">{prezzo_str}</td>
            <td>{v.get('STATO_VENDITA') or ''}</td>
            <td>{v.get('STATO_RIORDINO') or ''}</td>
            <td>{v.get('STATO_CONSERVAZIONE') or ''}</td>
            {loc_tds}
            <td class="num">{_fmt_qta(qta)}</td>
        </tr>
        """

    colspan = 13 + 3 + (8 if show_locations else 0)  # 13 = ID + 12 campi base, +3 stati
    html += f"""
        <tr class="totale-row">
            <td colspan="{colspan - 1}">TOTALE: {len(vini)} vini</td>
            <td class="num">{tot_bottiglie}</td>
        </tr>
        </tbody>
    </table>
    """
    return html


@router.get("/inventario/pdf", summary="PDF inventario completo")
def inventario_pdf(
    current_user: Any = Depends(_get_user_from_query_token),
):
    """
    Genera PDF con tutti i vini in cantina — inventario completo.
    """
    vini = _load_all_vini_inventario(solo_giacenza=False)
    data_oggi = datetime.now().strftime("%d/%m/%Y %H:%M")
    tot_bott = sum(v.get("QTA_TOTALE") or 0 for v in vini)
    tot_con_giacenza = sum(1 for v in vini if (v.get("QTA_TOTALE") or 0) > 0)

    table_html = _build_inventario_table(vini, show_locations=True)

    full_html = f"""
    <html><head><meta charset="utf-8">
    <style>{_inventario_css()}</style>
    </head><body>
        <div class="header">
            <h1>INVENTARIO CANTINA &mdash; TUTTI I VINI</h1>
            <div class="subtitle">Osteria Tre Gobbi &mdash; Generato il {data_oggi}</div>
        </div>
        <div class="summary-row">
            <div class="summary-box">
                <div class="label">Referenze</div>
                <div class="value">{len(vini)}</div>
            </div>
            <div class="summary-box">
                <div class="label">Con giacenza</div>
                <div class="value">{tot_con_giacenza}</div>
            </div>
            <div class="summary-box">
                <div class="label">Bottiglie totali</div>
                <div class="value">{tot_bott}</div>
            </div>
        </div>
        {table_html}
    </body></html>
    """

    out_path = STATIC_DIR / "inventario_completo.pdf"
    HTML(string=full_html).write_pdf(
        str(out_path),
        stylesheets=[CSS(string=_inventario_css())],
    )

    return FileResponse(
        str(out_path),
        filename=f"inventario_completo_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type="application/pdf",
    )


# =============================================================
# 8. INVENTARIO PDF — SOLO CON GIACENZA
# =============================================================
@router.get("/inventario/giacenza/pdf", summary="PDF inventario con giacenza")
def inventario_giacenza_pdf(
    current_user: Any = Depends(_get_user_from_query_token),
):
    """
    Genera PDF con solo i vini che hanno giacenza > 0.
    """
    vini = _load_all_vini_inventario(solo_giacenza=True)
    data_oggi = datetime.now().strftime("%d/%m/%Y %H:%M")
    tot_bott = sum(v.get("QTA_TOTALE") or 0 for v in vini)

    table_html = _build_inventario_table(vini, show_locations=True)

    full_html = f"""
    <html><head><meta charset="utf-8">
    <style>{_inventario_css()}</style>
    </head><body>
        <div class="header">
            <h1>INVENTARIO CANTINA &mdash; VINI CON GIACENZA</h1>
            <div class="subtitle">Osteria Tre Gobbi &mdash; Generato il {data_oggi}</div>
        </div>
        <div class="summary-row">
            <div class="summary-box">
                <div class="label">Referenze con giacenza</div>
                <div class="value">{len(vini)}</div>
            </div>
            <div class="summary-box">
                <div class="label">Bottiglie totali</div>
                <div class="value">{tot_bott}</div>
            </div>
        </div>
        {table_html}
    </body></html>
    """

    out_path = STATIC_DIR / "inventario_giacenza.pdf"
    HTML(string=full_html).write_pdf(
        str(out_path),
        stylesheets=[CSS(string=_inventario_css())],
    )

    return FileResponse(
        str(out_path),
        filename=f"inventario_giacenza_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type="application/pdf",
    )


# =============================================================
# 9. INVENTARIO PDF — PER LOCAZIONE
# =============================================================
@router.get("/inventario/locazioni/pdf", summary="PDF inventario per locazione")
def inventario_locazioni_pdf(
    current_user: Any = Depends(_get_user_from_query_token),
):
    """
    Genera PDF con vini raggruppati e separati per locazione fisica.
    Ogni locazione ha la sua sezione con i vini presenti li.
    """
    vini = _load_all_vini_inventario(solo_giacenza=True)
    data_oggi = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Raggruppa per locazione
    locazioni: Dict[str, List[Dict]] = {}

    for v in vini:
        # Frigorifero
        qta_f = v.get("QTA_FRIGO") or 0
        if qta_f > 0:
            loc_name = v.get("FRIGORIFERO") or "Frigorifero"
            loc_key = f"FRIGO:{loc_name}"
            locazioni.setdefault(loc_key, []).append({**v, "_qta_loc": qta_f})

        # Locazione 1
        qta_1 = v.get("QTA_LOC1") or 0
        if qta_1 > 0:
            loc_name = v.get("LOCAZIONE_1") or "Locazione 1"
            loc_key = f"LOC1:{loc_name}"
            locazioni.setdefault(loc_key, []).append({**v, "_qta_loc": qta_1})

        # Locazione 2
        qta_2 = v.get("QTA_LOC2") or 0
        if qta_2 > 0:
            loc_name = v.get("LOCAZIONE_2") or "Locazione 2"
            loc_key = f"LOC2:{loc_name}"
            locazioni.setdefault(loc_key, []).append({**v, "_qta_loc": qta_2})

        # Locazione 3
        qta_3 = v.get("QTA_LOC3") or 0
        if qta_3 > 0:
            loc_name = v.get("LOCAZIONE_3") or "Locazione 3"
            loc_key = f"LOC3:{loc_name}"
            locazioni.setdefault(loc_key, []).append({**v, "_qta_loc": qta_3})

    # Costruisci HTML per ogni locazione
    sections_html = ""
    grand_total = 0

    # Ordina per nome locazione alfabetico; nomi che iniziano con ( o cifra vanno in fondo
    def _loc_sort_key(k):
        name = k.split(":", 1)[1].lower() if ":" in k else k.lower()
        if name and not name[0].isalpha():
            return (1, name)  # non-alfa in fondo
        return (0, name)

    for loc_key in sorted(locazioni.keys(), key=_loc_sort_key):
        loc_vini = locazioni[loc_key]
        loc_label = loc_key.split(":", 1)[1] if ":" in loc_key else loc_key
        tot_loc = sum(v["_qta_loc"] for v in loc_vini)
        grand_total += tot_loc

        sections_html += f'<div class="loc-header">{loc_label} — {len(loc_vini)} vini, {tot_loc} bottiglie</div>'
        sections_html += """
        <table>
            <thead><tr>
                <th class="num">ID</th>
                <th>Tipologia</th>
                <th>Produttore</th>
                <th>Descrizione</th>
                <th>Annata</th>
                <th>Formato</th>
                <th class="num">Prezzo</th>
                <th class="num">Qta in loc.</th>
                <th class="num">Qta totale</th>
            </tr></thead>
            <tbody>
        """

        for v in loc_vini:
            prezzo = v.get("PREZZO_CARTA")
            prezzo_str = f"&euro; {float(prezzo):.2f}" if prezzo else ""
            sections_html += f"""
            <tr>
                <td class="num">{v.get('id') or ''}</td>
                <td>{v.get('TIPOLOGIA') or ''}</td>
                <td>{v.get('PRODUTTORE') or ''}</td>
                <td>{v.get('DESCRIZIONE') or ''}</td>
                <td>{v.get('ANNATA') or ''}</td>
                <td>{v.get('FORMATO') or ''}</td>
                <td class="num">{prezzo_str}</td>
                <td class="num">{_fmt_qta(v['_qta_loc'])}</td>
                <td class="num">{_fmt_qta(v.get('QTA_TOTALE') or 0)}</td>
            </tr>
            """

        sections_html += f"""
            <tr class="totale-row">
                <td colspan="7">Totale {loc_label}</td>
                <td class="num">{tot_loc}</td>
                <td></td>
            </tr>
            </tbody>
        </table>
        """

    tot_bott = sum(v.get("QTA_TOTALE") or 0 for v in vini)

    full_html = f"""
    <html><head><meta charset="utf-8">
    <style>{_inventario_css()}</style>
    </head><body>
        <div class="header">
            <h1>INVENTARIO CANTINA &mdash; PER LOCAZIONE</h1>
            <div class="subtitle">Osteria Tre Gobbi &mdash; Generato il {data_oggi}</div>
        </div>
        <div class="summary-row">
            <div class="summary-box">
                <div class="label">Locazioni</div>
                <div class="value">{len(locazioni)}</div>
            </div>
            <div class="summary-box">
                <div class="label">Referenze con giacenza</div>
                <div class="value">{len(vini)}</div>
            </div>
            <div class="summary-box">
                <div class="label">Bottiglie totali</div>
                <div class="value">{tot_bott}</div>
            </div>
        </div>
        {sections_html}
    </body></html>
    """

    out_path = STATIC_DIR / "inventario_locazioni.pdf"
    HTML(string=full_html).write_pdf(
        str(out_path),
        stylesheets=[CSS(string=_inventario_css())],
    )

    return FileResponse(
        str(out_path),
        filename=f"inventario_locazioni_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type="application/pdf",
    )


# =============================================================
# 10. INVENTARIO PDF — FILTRATO (filtri componibili)
# =============================================================
def _build_filtri_subtitle(kwargs: Dict[str, Any]) -> str:
    """Costruisce stringa leggibile dei filtri attivi."""
    labels = {
        "tipologia": "Tipologia",
        "nazione": "Nazione",
        "regione": "Regione",
        "produttore": "Produttore",
        "annata": "Annata",
        "formato": "Formato",
        "carta": "In carta",
        "stato_vendita": "Stato vendita",
        "stato_riordino": "Stato riordino",
        "discontinuato": "Discontinuato",
        "text": "Testo",
    }
    parts = []
    for k, label in labels.items():
        v = kwargs.get(k)
        if v:
            parts.append(f"{label}: {v}")
    # Range quantita
    qmin = kwargs.get("qta_min")
    qmax = kwargs.get("qta_max")
    if qmin is not None and qmax is not None:
        parts.append(f"Qta: {qmin}-{qmax}")
    elif qmin is not None:
        parts.append(f"Qta >= {qmin}")
    elif qmax is not None:
        parts.append(f"Qta <= {qmax}")
    # Range prezzo
    pmin = kwargs.get("prezzo_min")
    pmax = kwargs.get("prezzo_max")
    if pmin is not None and pmax is not None:
        parts.append(f"Prezzo: {pmin:.2f}-{pmax:.2f}")
    elif pmin is not None:
        parts.append(f"Prezzo >= {pmin:.2f}")
    elif pmax is not None:
        parts.append(f"Prezzo <= {pmax:.2f}")
    if kwargs.get("solo_giacenza"):
        parts.append("Solo con giacenza")
    # Locazione legacy
    if kwargs.get("locazione"):
        parts.append(f"Locazione: {kwargs['locazione']}")
    # Locazioni gerarchiche
    for label, nome_key, spazio_key in [
        ("Frigo", "frigo_nome", "frigo_spazio"),
        ("Loc 1", "loc1_nome", "loc1_spazio"),
        ("Loc 2", "loc2_nome", "loc2_spazio"),
        ("Loc 3", "loc3_nome", "loc3_spazio"),
    ]:
        nome = kwargs.get(nome_key)
        spazio = kwargs.get(spazio_key)
        if nome:
            val = f"{nome} - {spazio}" if spazio else nome
            parts.append(f"{label}: {val}")
    return " | ".join(parts) if parts else "Nessun filtro"


@router.get("/inventario/filtrato/pdf", summary="PDF inventario con filtri componibili")
def inventario_filtrato_pdf(
    current_user: Any = Depends(_get_user_from_query_token),
    tipologia: Optional[str] = Query(None),
    nazione: Optional[str] = Query(None),
    regione: Optional[str] = Query(None),
    produttore: Optional[str] = Query(None),
    annata: Optional[str] = Query(None),
    formato: Optional[str] = Query(None),
    carta: Optional[str] = Query(None),
    stato_vendita: Optional[str] = Query(None),
    stato_riordino: Optional[str] = Query(None),
    stato_conservazione: Optional[str] = Query(None),
    discontinuato: Optional[str] = Query(None),
    qta_min: Optional[int] = Query(None),
    qta_max: Optional[int] = Query(None),
    prezzo_min: Optional[float] = Query(None),
    prezzo_max: Optional[float] = Query(None),
    solo_giacenza: bool = Query(False),
    text: Optional[str] = Query(None),
    locazione: Optional[str] = Query(None),
    frigo_nome: Optional[str] = Query(None),
    frigo_spazio: Optional[str] = Query(None),
    loc1_nome: Optional[str] = Query(None),
    loc1_spazio: Optional[str] = Query(None),
    loc2_nome: Optional[str] = Query(None),
    loc2_spazio: Optional[str] = Query(None),
    loc3_nome: Optional[str] = Query(None),
    loc3_spazio: Optional[str] = Query(None),
):
    """
    Genera PDF inventario con filtri combinabili via query string.
    Tutti i filtri sono opzionali e componibili tra loro.
    """
    filter_kwargs = dict(
        solo_giacenza=solo_giacenza,
        tipologia=tipologia,
        nazione=nazione,
        regione=regione,
        produttore=produttore,
        annata=annata,
        formato=formato,
        carta=carta,
        stato_vendita=stato_vendita,
        stato_riordino=stato_riordino,
        stato_conservazione=stato_conservazione,
        discontinuato=discontinuato,
        qta_min=qta_min,
        qta_max=qta_max,
        prezzo_min=prezzo_min,
        prezzo_max=prezzo_max,
        text=text,
        locazione=locazione,
        frigo_nome=frigo_nome,
        frigo_spazio=frigo_spazio,
        loc1_nome=loc1_nome,
        loc1_spazio=loc1_spazio,
        loc2_nome=loc2_nome,
        loc2_spazio=loc2_spazio,
        loc3_nome=loc3_nome,
        loc3_spazio=loc3_spazio,
    )

    vini = _load_all_vini_inventario(**filter_kwargs)
    data_oggi = datetime.now().strftime("%d/%m/%Y %H:%M")
    tot_bott = sum(v.get("QTA_TOTALE") or 0 for v in vini)
    filtri_str = _build_filtri_subtitle(filter_kwargs)

    table_html = _build_inventario_table(vini, show_locations=True)

    full_html = f"""
    <html><head><meta charset="utf-8">
    <style>{_inventario_css()}</style>
    </head><body>
        <div class="header">
            <h1>INVENTARIO CANTINA &mdash; FILTRATO</h1>
            <div class="subtitle">Osteria Tre Gobbi &mdash; Generato il {data_oggi}</div>
            <div class="subtitle" style="margin-top:3px;">{filtri_str}</div>
        </div>
        <div class="summary-row">
            <div class="summary-box">
                <div class="label">Referenze</div>
                <div class="value">{len(vini)}</div>
            </div>
            <div class="summary-box">
                <div class="label">Bottiglie totali</div>
                <div class="value">{tot_bott}</div>
            </div>
        </div>
        {table_html}
    </body></html>
    """

    out_path = STATIC_DIR / "inventario_filtrato.pdf"
    HTML(string=full_html).write_pdf(
        str(out_path),
        stylesheets=[CSS(string=_inventario_css())],
    )

    return FileResponse(
        str(out_path),
        filename=f"inventario_filtrato_{datetime.now().strftime('%Y%m%d')}.pdf",
        media_type="application/pdf",
    )


# =============================================================
# 10b. ENDPOINT: valori distinti per popolare i filtri frontend
# =============================================================
@router.get("/inventario/filtri-options", summary="Valori distinti per filtri inventario")
def inventario_filtri_options(
    current_user: Any = Depends(get_current_user),
):
    """Restituisce i valori distinti per popolare le select dei filtri."""
    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()

    def distinct(col):
        rows = cur.execute(
            f"SELECT DISTINCT {col} FROM vini_magazzino "
            f"WHERE {col} IS NOT NULL AND {col} != '' ORDER BY {col}"
        ).fetchall()
        return [r[0] for r in rows]

    result = {
        "tipologie": distinct("TIPOLOGIA"),
        "nazioni": distinct("NAZIONE"),
        "regioni": distinct("REGIONE"),
        "produttori": distinct("PRODUTTORE"),
        "annate": distinct("ANNATA"),
        "formati": distinct("FORMATO"),
        "stati_vendita": distinct("STATO_VENDITA"),
        "stati_riordino": distinct("STATO_RIORDINO"),
        "stati_conservazione": distinct("STATO_CONSERVAZIONE"),
    }
    conn.close()
    return result


# =============================================================
# GESTIONE LOCAZIONI FISICHE (generico per tutti i campi)
# =============================================================

import json as _json

# Mappa campo frontend → colonna DB
LOCATION_FIELDS = {
    "frigorifero": {
        "column": "FRIGORIFERO",
        "qta_column": "QTA_FRIGO",
        "label": "Frigorifero",
    },
    "locazione_1": {
        "column": "LOCAZIONE_1",
        "qta_column": "QTA_LOC1",
        "label": "Locazione 1",
    },
    "locazione_2": {
        "column": "LOCAZIONE_2",
        "qta_column": "QTA_LOC2",
        "label": "Locazione 2",
    },
    "locazione_3": {
        "column": "LOCAZIONE_3",
        "qta_column": "QTA_LOC3",
        "label": "Locazione 3",
    },
}


def _config_campo(campo: str) -> str:
    """Locazione 2 condivide la config di Locazione 1."""
    return "locazione_1" if campo == "locazione_2" else campo


def _load_locazioni_config(campo: str) -> list[dict]:
    """Carica la configurazione locazioni per un campo dal DB."""
    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()
    rows = cur.execute(
        "SELECT id, nome, spazi, ordine, tipo, righe, colonne FROM locazioni_config "
        "WHERE campo = ? ORDER BY ordine, id",
        (campo,),
    ).fetchall()
    conn.close()
    result = []
    for row in rows:
        tipo = row["tipo"] if row["tipo"] else "standard"
        righe = row["righe"] if row["righe"] else None
        colonne = row["colonne"] if row["colonne"] else None
        spazi = _json.loads(row["spazi"]) if row["spazi"] else []
        result.append({
            "id": row["id"], "nome": row["nome"], "spazi": spazi,
            "ordine": row["ordine"], "tipo": tipo,
            "righe": righe, "colonne": colonne,
        })
    return result


def _build_options_from_config(campo: str) -> list[str]:
    """Costruisce la lista di opzioni valide per un campo leggendo la config dal DB.
    Se una locazione non ha sotto-spazi (spazi vuoto), il valore è solo il nome.
    Se tipo='matrice', genera le coordinate (R,C) automaticamente da righe×colonne."""
    items = _load_locazioni_config(campo)
    opts = []
    for item in items:
        if item.get("tipo") == "matrice" and item.get("righe") and item.get("colonne"):
            # Genera coordinate matrice: (col,riga) — colonna prima, riga dopo
            for r in range(1, item["righe"] + 1):
                for c in range(1, item["colonne"] + 1):
                    opts.append(f"{item['nome']} - ({c},{r})")
        elif item["spazi"]:
            for spazio in item["spazi"]:
                opts.append(f"{item['nome']} - {spazio}")
        else:
            # Locazione con spazio unico — nessun sotto-spazio
            opts.append(item["nome"])
    return opts


def _suggest_value(val: str, valid_options: set, campo_items: list[dict]) -> str | None:
    """Prova a suggerire il valore normalizzato per un valore di locazione."""
    if not val or not val.strip():
        return None
    val = val.strip()

    if val in valid_options:
        return val

    # Pattern comuni per frigoriferi
    m = re.match(r"^[Ff]rigo[\s-]*(\d+)\s*[-\s]+[Ff]?i?l?a?\s*(\d+)$", val)
    if m:
        candidate = f"Frigo {m.group(1)} - Fila {m.group(2)}"
        if candidate in valid_options:
            return candidate

    m = re.match(r"^[Ff]rigo[\s-]*(\d+)[\s-]+(\d+)$", val)
    if m:
        candidate = f"Frigo {m.group(1)} - Fila {m.group(2)}"
        if candidate in valid_options:
            return candidate

    # Formato numerico: "1-3"
    m = re.match(r"^(\d+)[\s-]+(\d+)$", val)
    if m:
        for item in campo_items:
            candidate = f"{item['nome']} - Fila {m.group(2)}"
            if candidate in valid_options:
                return candidate

    return None


# -----------------------------------------------
# CRUD configurazione locazioni
# -----------------------------------------------

@router.get("/locazioni-config", summary="Config locazioni fisiche")
async def get_locazioni_config(current_user=Depends(get_current_user)):
    """Ritorna la struttura delle locazioni fisiche per tutti i campi."""
    result = {"fields": {k: v["label"] for k, v in LOCATION_FIELDS.items()}}
    for campo_key in LOCATION_FIELDS:
        # Locazione 2 condivide la stessa configurazione di Locazione 1
        if campo_key == "locazione_2":
            result[campo_key] = _load_locazioni_config("locazione_1")
        else:
            result[campo_key] = _load_locazioni_config(campo_key)
    result["opzioni_frigo"] = _build_options_from_config("frigorifero")
    opzioni_loc = _build_options_from_config("locazione_1")
    result["opzioni_locazione_1"] = opzioni_loc
    result["opzioni_locazione_2"] = opzioni_loc  # stesse opzioni di loc1
    result["opzioni_locazione_3"] = _build_options_from_config("locazione_3")
    return result


@router.post("/locazioni-config/{campo}", summary="Salva configurazione locazione")
async def save_locazione_config(
    campo: str,
    request: Request,
    current_user=Depends(get_current_user),
):
    """
    Salva/aggiorna la configurazione di un singolo elemento locazione.
    Body: { "nome": "Frigo 1", "spazi": ["Fila 1", "Fila 2", ...], "ordine": 0 }
    Se presente "id", aggiorna. Altrimenti crea nuovo.
    """
    _require_admin(current_user)

    if campo not in LOCATION_FIELDS:
        raise HTTPException(400, f"Campo non valido. Usa: {', '.join(LOCATION_FIELDS.keys())}")

    body = await request.json()
    nome = body.get("nome", "").strip()
    spazi = body.get("spazi", [])
    ordine = body.get("ordine", 0)
    item_id = body.get("id")
    tipo = body.get("tipo", "standard")  # "standard" o "matrice"
    righe = body.get("righe")            # solo per tipo=matrice
    colonne = body.get("colonne")        # solo per tipo=matrice

    if not nome:
        raise HTTPException(400, "Il nome è obbligatorio.")

    if tipo == "matrice":
        if not righe or not colonne or righe < 1 or colonne < 1:
            raise HTTPException(400, "Per una matrice, righe e colonne devono essere >= 1.")
        # Per le matrici, gli spazi vengono generati automaticamente — non serve salvarli
        spazi = []

    now = datetime.now().isoformat(timespec="seconds")
    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()

    if item_id:
        cur.execute(
            "UPDATE locazioni_config SET nome = ?, spazi = ?, ordine = ?, tipo = ?, righe = ?, colonne = ?, updated_at = ? "
            "WHERE id = ? AND campo = ?",
            (nome, _json.dumps(spazi), ordine, tipo, righe, colonne, now, item_id, campo),
        )
    else:
        cur.execute(
            "INSERT INTO locazioni_config (campo, nome, spazi, ordine, tipo, righe, colonne, created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (campo, nome, _json.dumps(spazi), ordine, tipo, righe, colonne, now, now),
        )

    conn.commit()
    conn.close()
    return {"ok": True, "msg": f"Locazione '{nome}' salvata per {LOCATION_FIELDS[campo]['label']}."}


@router.delete("/locazioni-config/{campo}/{item_id}", summary="Elimina configurazione locazione")
async def delete_locazione_config(
    campo: str,
    item_id: int,
    current_user=Depends(get_current_user),
):
    """Elimina un elemento dalla configurazione locazioni."""
    _require_admin(current_user)

    if campo not in LOCATION_FIELDS:
        raise HTTPException(400, f"Campo non valido.")

    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM locazioni_config WHERE id = ? AND campo = ?", (item_id, campo))
    conn.commit()
    conn.close()
    return {"ok": True, "msg": "Locazione eliminata."}


# -----------------------------------------------
# Estrazione valori e normalizzazione
# -----------------------------------------------

@router.get("/locazioni-valori/{campo}", summary="Valori distinti per campo locazione")
async def get_locazioni_valori(
    campo: str,
    current_user=Depends(get_current_user),
):
    """Estrae valori distinti con conteggio e suggerimento normalizzazione."""
    if campo not in LOCATION_FIELDS:
        raise HTTPException(400, f"Campo non valido. Usa: {', '.join(LOCATION_FIELDS.keys())}")

    col = LOCATION_FIELDS[campo]["column"]
    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()
    rows = cur.execute(
        f"SELECT {col}, COUNT(*) as cnt FROM vini_magazzino "
        f"WHERE {col} IS NOT NULL AND {col} != '' "
        f"GROUP BY {col} ORDER BY cnt DESC"
    ).fetchall()
    conn.close()

    valid_options = set(_build_options_from_config(_config_campo(campo)))
    campo_items = _load_locazioni_config(_config_campo(campo))

    valori = []
    for row in rows:
        val = row[0]
        cnt = row[1]
        entry = {"valore": val, "conteggio": cnt, "suggerimento": None, "ok": False}
        if val in valid_options:
            entry["ok"] = True
        elif valid_options:
            entry["suggerimento"] = _suggest_value(val, valid_options, campo_items)
        valori.append(entry)

    return {
        "campo": campo,
        "label": LOCATION_FIELDS[campo]["label"],
        "totale_distinti": len(valori),
        "totale_record": sum(v["conteggio"] for v in valori),
        "valori": valori,
        "opzioni_valide": sorted(valid_options),
    }


@router.post("/locazioni-normalizza", summary="Applica mapping normalizzazione locazioni")
async def applica_normalizzazione_locazioni(
    request: Request,
    current_user=Depends(get_current_user),
):
    """
    Applica mapping di normalizzazione.
    Body: { "campo": "frigorifero", "mapping": {"Frigo-1-3": "Frigo 1 - Fila 3", ...} }
    """
    _require_admin(current_user)

    body = await request.json()
    campo = body.get("campo")
    mapping = body.get("mapping", {})

    if campo not in LOCATION_FIELDS:
        raise HTTPException(400, f"Campo non valido. Usa: {', '.join(LOCATION_FIELDS.keys())}")

    if not mapping:
        return {"modificati": 0, "msg": "Nessun mapping da applicare."}

    col = LOCATION_FIELDS[campo]["column"]
    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()
    totale = 0

    svuotati = 0
    for old_val, new_val in mapping.items():
        if new_val is None:
            continue  # None = non toccare
        # Stringa vuota = svuota la locazione
        actual_val = new_val.strip() if new_val else None
        result = cur.execute(
            f"UPDATE vini_magazzino SET {col} = ? WHERE {col} = ?",
            (actual_val, old_val),
        )
        totale += result.rowcount
        if not actual_val:
            svuotati += result.rowcount

    conn.commit()
    conn.close()

    parts = [f"Aggiornati {totale} record"]
    if svuotati:
        parts.append(f"di cui {svuotati} svuotati")
    return {
        "campo": campo,
        "modificati": totale,
        "svuotati": svuotati,
        "msg": f"{' '.join(parts)} nel campo {LOCATION_FIELDS[campo]['label']}.",
    }


@router.get("/locazioni-vini/{campo}", summary="Vini per valore locazione")
async def get_vini_per_locazione(
    campo: str,
    valore: str = Query(..., description="Valore locazione da cercare"),
    current_user=Depends(get_current_user),
):
    """Ritorna i vini che hanno un certo valore nel campo locazione specificato."""
    if campo not in LOCATION_FIELDS:
        raise HTTPException(400, f"Campo non valido.")

    col = LOCATION_FIELDS[campo]["column"]
    qta_col = LOCATION_FIELDS[campo]["qta_column"]
    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()
    rows = cur.execute(
        f"SELECT id, DESCRIZIONE, PRODUTTORE, ANNATA, FORMATO, {col}, {qta_col} "
        f"FROM vini_magazzino WHERE {col} = ? ORDER BY DESCRIZIONE",
        (valore,),
    ).fetchall()
    conn.close()

    return {
        "campo": campo,
        "valore": valore,
        "vini": [
            {
                "id": r["id"],
                "descrizione": r["DESCRIZIONE"],
                "produttore": r["PRODUTTORE"],
                "annata": r["ANNATA"],
                "formato": r["FORMATO"],
                "locazione": r[col],
                "quantita": r[qta_col],
            }
            for r in rows
        ],
    }


@router.post("/locazioni-check-giacenze", summary="Verifica giacenze prima di svuotare")
async def check_giacenze_locazione(
    request: Request,
    current_user=Depends(get_current_user),
):
    """
    Dato un elenco di valori da svuotare, ritorna i vini con giacenza > 0.
    Body: { "campo": "frigorifero", "valori": ["Frigo-1-1", "Frigo"] }
    """
    body = await request.json()
    campo = body.get("campo")
    valori = body.get("valori", [])

    if campo not in LOCATION_FIELDS:
        raise HTTPException(400, "Campo non valido.")
    if not valori:
        return {"vini_con_giacenza": []}

    col = LOCATION_FIELDS[campo]["column"]
    qta_col = LOCATION_FIELDS[campo]["qta_column"]
    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()

    placeholders = ",".join(["?"] * len(valori))
    rows = cur.execute(
        f"SELECT id, DESCRIZIONE, PRODUTTORE, ANNATA, {col}, {qta_col} "
        f"FROM vini_magazzino WHERE {col} IN ({placeholders}) "
        f"AND {qta_col} IS NOT NULL AND {qta_col} > 0 "
        f"ORDER BY {qta_col} DESC",
        valori,
    ).fetchall()
    conn.close()

    return {
        "vini_con_giacenza": [
            {
                "id": r["id"],
                "descrizione": r["DESCRIZIONE"],
                "produttore": r["PRODUTTORE"],
                "annata": r["ANNATA"],
                "locazione": r[col],
                "quantita": r[qta_col],
            }
            for r in rows
        ],
        "totale": len(rows),
    }


@router.post("/locazioni-vino-update", summary="Aggiorna locazione singolo vino")
async def update_vino_locazione(
    request: Request,
    current_user=Depends(get_current_user),
):
    """
    Aggiorna il valore di locazione per un singolo vino.
    Body: { "campo": "frigorifero", "vino_id": 123, "nuovo_valore": "Frigo 1 - Fila 3" }
    """
    _require_admin(current_user)

    body = await request.json()
    campo = body.get("campo")
    vino_id = body.get("vino_id")
    nuovo_valore = body.get("nuovo_valore", "").strip()

    if campo not in LOCATION_FIELDS:
        raise HTTPException(400, f"Campo non valido.")
    if not vino_id:
        raise HTTPException(400, "vino_id obbligatorio.")

    col = LOCATION_FIELDS[campo]["column"]
    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()
    cur.execute(
        f"UPDATE vini_magazzino SET {col} = ? WHERE id = ?",
        (nuovo_valore if nuovo_valore else None, vino_id),
    )
    conn.commit()
    conn.close()

    return {"ok": True, "msg": f"Locazione aggiornata per vino #{vino_id}."}


# -----------------------------------------------
# MATRICE — gestione celle
# -----------------------------------------------

@router.get("/matrice/stato", summary="Stato completo della matrice")
async def get_matrice_stato(current_user=Depends(get_current_user)):
    """Ritorna la griglia con tutte le celle occupate e la config dimensioni."""
    return mag_db.matrice_get_stato()


@router.get("/matrice/celle/{vino_id}", summary="Celle matrice per un vino")
async def get_matrice_celle_vino(
    vino_id: int,
    current_user=Depends(get_current_user),
):
    """Ritorna le celle assegnate a un vino specifico."""
    return {"vino_id": vino_id, "celle": mag_db.matrice_get_celle_vino(vino_id)}


@router.post("/matrice/assegna", summary="Assegna cella a un vino")
async def matrice_assegna(
    request: Request,
    current_user=Depends(get_current_user),
):
    """Body: { vino_id, riga, colonna }"""
    body = await request.json()
    vino_id = body.get("vino_id")
    riga = body.get("riga")
    colonna = body.get("colonna")

    if not vino_id or riga is None or colonna is None:
        raise HTTPException(400, "vino_id, riga e colonna sono obbligatori.")

    try:
        result = mag_db.matrice_assegna_cella(vino_id, riga, colonna)
    except ValueError as e:
        raise HTTPException(409, str(e))

    # Ritorna anche il vino aggiornato per sincronizzare il frontend
    updated = mag_db.get_vino_by_id(vino_id)
    result["vino"] = dict(updated) if updated else None
    return result


@router.post("/matrice/rimuovi", summary="Rimuovi cella da un vino")
async def matrice_rimuovi(
    request: Request,
    current_user=Depends(get_current_user),
):
    """Body: { vino_id, riga, colonna }"""
    body = await request.json()
    vino_id = body.get("vino_id")
    riga = body.get("riga")
    colonna = body.get("colonna")

    if not vino_id or riga is None or colonna is None:
        raise HTTPException(400, "vino_id, riga e colonna sono obbligatori.")

    result = mag_db.matrice_rimuovi_cella(vino_id, riga, colonna)

    updated = mag_db.get_vino_by_id(vino_id)
    result["vino"] = dict(updated) if updated else None
    return result


@router.post("/matrice/set-celle", summary="Imposta tutte le celle per un vino")
async def matrice_set_celle(
    request: Request,
    current_user=Depends(get_current_user),
):
    """Body: { vino_id, celle: [{riga, colonna}, ...] }"""
    body = await request.json()
    vino_id = body.get("vino_id")
    celle = body.get("celle", [])

    if not vino_id:
        raise HTTPException(400, "vino_id è obbligatorio.")

    result = mag_db.matrice_set_celle_vino(vino_id, celle)

    updated = mag_db.get_vino_by_id(vino_id)
    result["vino"] = dict(updated) if updated else None
    return result


@router.get("/matrice/recalc-preview", summary="Anteprima migrazione coordinate matrice")
async def matrice_recalc_preview(current_user=Depends(get_current_user)):
    """Mostra prima/dopo per tutti i vini con celle matrice, senza modificare nulla."""
    _require_admin(current_user)
    return mag_db.matrice_recalc_preview()


@router.post("/matrice/recalc-all", summary="Ricalcola LOCAZIONE_3 per tutti i vini con celle matrice")
async def matrice_recalc_all(current_user=Depends(get_current_user)):
    """Migrazione: ricalcola il campo LOCAZIONE_3 con il formato (col,riga) per tutti i vini."""
    _require_admin(current_user)
    count = mag_db.matrice_recalc_all()
    return {"ok": True, "vini_aggiornati": count}


@router.get("/matrice/old-values", summary="Mostra valori matrice in TUTTE le locazioni")
async def matrice_old_values(current_user=Depends(get_current_user)):
    """Debug: cerca valori contenenti 'matrice' o coordinate in tutte le locazioni."""
    import re
    conn = mag_db.get_magazzino_connection()
    cur = conn.cursor()
    rows = cur.execute(
        "SELECT id, DESCRIZIONE, FRIGORIFERO, QTA_FRIGO, "
        "LOCAZIONE_1, QTA_LOC1, LOCAZIONE_2, QTA_LOC2, "
        "LOCAZIONE_3, QTA_LOC3 FROM vini_magazzino "
        "ORDER BY id"
    ).fetchall()
    existing = set(r[0] for r in cur.execute("SELECT DISTINCT vino_id FROM matrice_celle").fetchall())
    # Conta celle per vino
    celle_count = {}
    for r2 in cur.execute("SELECT vino_id, COUNT(*) as n FROM matrice_celle GROUP BY vino_id").fetchall():
        celle_count[r2[0]] = r2[1]
    conn.close()

    results = []
    loc_fields = {
        "FRIGORIFERO": "QTA_FRIGO",
        "LOCAZIONE_1": "QTA_LOC1",
        "LOCAZIONE_2": "QTA_LOC2",
        "LOCAZIONE_3": "QTA_LOC3",
    }
    for r in rows:
        trovati = {}
        for campo, qta_campo in loc_fields.items():
            val = r[campo]
            if not val:
                continue
            # Cerca "matrice" nel testo O coordinate con parentesi O coordinate senza parentesi (es. "6,7")
            is_matrice = "matrice" in val.lower() or re.search(r'\(?\d+\s*,\s*\d+\)?', val)
            if is_matrice:
                # Prova a estrarre coordinate: prima con parentesi, poi senza
                coords = re.findall(r'\((\d+)\s*,\s*(\d+)\)', val)
                if not coords:
                    # Senza parentesi: "Matrice 6,7" o "Matrice - 6,7" o solo "6,7"
                    coords = re.findall(r'(\d+)\s*,\s*(\d+)', val)
                trovati[campo] = {
                    "valore": val,
                    "qta": r[qta_campo],
                    "coordinate": [f"({a},{b})" for a, b in coords],
                }
        if trovati:
            results.append({
                "id": r["id"],
                "descrizione": r["DESCRIZIONE"],
                "campi_con_matrice": trovati,
                "ha_celle_matrice": r["id"] in existing,
                "celle_in_tabella": celle_count.get(r["id"], 0),
            })
    return results


@router.post("/matrice/import-old", summary="Importa vecchi valori matrice da tutte le locazioni in matrice_celle")
async def matrice_import_old(current_user=Depends(get_current_user)):
    """Migrazione: cerca coordinate matrice in tutte le locazioni, le importa in matrice_celle,
    e pulisce i vecchi campi."""
    _require_admin(current_user)
    result = mag_db.matrice_import_from_all_locations()
    return result
