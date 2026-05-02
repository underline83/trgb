# @version: v2.0-ipratico-trgb-priority
# Router iPratico Products — import/export Excel prodotti, mapping ↔ vini TRGB
# Il codice 4 cifre nel Name iPratico corrisponde DIRETTAMENTE a vini_magazzino.id
# TRGB ha priorità: se i dati cambiano su TRGB, l'export aggiorna iPratico.
"""
Endpoints:
  POST /vini/ipratico/upload          Upload Excel export iPratico, parse Bottiglie, match per ID diretto
  GET  /vini/ipratico/mappings        Lista mapping corrente (con dati iPratico + TRGB)
  PUT  /vini/ipratico/mappings/{id}   Aggiorna mapping manuale (assegna/rimuovi vino_id)
  PUT  /vini/ipratico/ignore/{id}     Segna prodotto come ignorato (esiste solo su iPratico)
  POST /vini/ipratico/export          Genera Excel aggiornato con QTA, testi, prezzi TRGB + vini mancanti
  GET  /vini/ipratico/missing         Vini TRGB non presenti nell'export iPratico (da aggiungere)
  GET  /vini/ipratico/sync-log        Storico sincronizzazioni
  GET  /vini/ipratico/stats           Riepilogo veloce
"""
from __future__ import annotations

import io
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List as TList

router = APIRouter(prefix="/vini/ipratico", tags=["ipratico-products"])

from app.utils.locale_data import locale_data_path

# R6.5 — path tenant-aware. Modulo: vini (sub: ipratico products).
DB_MAG = locale_data_path("vini_magazzino.sqlite3")
DB_FC = locale_data_path("foodcost.db")  # migration tables live here

# NB: ipratico_uploads è una cartella di upload utente, fuori scope R6.5.
# TODO Modulo K: spostare sotto TRGB_UPLOADS_DIR/<locale>/ipratico_uploads.
UPLOAD_DIR = Path("app/data/ipratico_uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


# ─── helpers ────────────────────────────────────────────────────────
def _fc_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_FC)
    conn.row_factory = sqlite3.Row
    return conn


def _mag_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_MAG)
    conn.row_factory = sqlite3.Row
    return conn


_RE_ID = re.compile(r"^(\d{4})")


def _load_export_defaults() -> dict[str, str]:
    """Carica i default export dalla tabella ipratico_export_defaults → {field_name: field_value}."""
    fc = _fc_conn()
    try:
        rows = fc.execute("SELECT field_name, field_value FROM ipratico_export_defaults").fetchall()
        return {r["field_name"]: r["field_value"] for r in rows}
    except Exception:
        return {}
    finally:
        fc.close()


def _extract_wine_id(name: str) -> Optional[int]:
    """Estrae il codice 4 cifre dal campo Name iPratico → corrisponde a vini_magazzino.id."""
    m = _RE_ID.match(name.strip())
    return int(m.group(1)) if m else None


def _build_ipratico_name(wine: dict) -> str:
    """Costruisce il campo Name iPratico dai dati TRGB.
    Formato: {ID:04d} {DESCRIZIONE} ({FORMATO}) {ANNATA} {PRODUTTORE}
    """
    parts = [str(wine["id"]).zfill(4)]

    desc = (wine.get("DESCRIZIONE") or "").strip()
    if desc:
        parts.append(desc)

    denom = (wine.get("DENOMINAZIONE") or "").strip()
    if denom:
        parts.append(f"({denom})")

    fmt = (wine.get("FORMATO") or "").strip()
    if fmt and fmt.upper() != "BT":  # BT = bottiglia standard, non serve
        parts.append(f"({fmt})")

    annata = (wine.get("ANNATA") or "").strip()
    if annata:
        parts.append(annata)

    prod = (wine.get("PRODUTTORE") or "").strip()
    if prod:
        parts.append(prod)

    return " ".join(parts)


# ─── Upload & Match diretto per ID ────────────────────────────────
@router.post("/upload")
async def upload_ipratico_export(file: UploadFile = File(...)):
    """Upload export Excel iPratico, parse Bottiglie, match diretto per ID."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Il file deve essere in formato .xlsx o .xls")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl non installato sul server")

    content = await file.read()

    # Save file
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    saved_path = UPLOAD_DIR / f"ipratico_{ts}.xlsx"
    saved_path.write_bytes(content)

    # Parse Excel
    try:
        import pandas as pd
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Errore lettura Excel: {e}")

    if "Category" not in df.columns or "Name" not in df.columns:
        raise HTTPException(400, "Colonne 'Category' e 'Name' non trovate nel file")

    bottiglie = df[df["Category"] == "Bottiglie"].copy()
    if bottiglie.empty:
        raise HTTPException(400, "Nessun prodotto con categoria 'Bottiglie' trovato")

    # Load vini_magazzino IDs per verifica esistenza
    mag_ids = set()
    try:
        mconn = _mag_conn()
        for r in mconn.execute("SELECT id FROM vini_magazzino").fetchall():
            mag_ids.add(r["id"])
        mconn.close()
    except Exception:
        pass

    fc = _fc_conn()

    # Pulisci mapping precedenti e ricostruisci da zero
    fc.execute("DELETE FROM ipratico_product_map")
    fc.commit()

    n_matched = 0
    n_unmatched = 0

    for _, row in bottiglie.iterrows():
        ipratico_uuid = str(row.get("Id", ""))
        name = str(row.get("Name", ""))
        wine_id = _extract_wine_id(name)

        # Match diretto: wine_id = vini_magazzino.id
        vino_id = wine_id if wine_id and wine_id in mag_ids else None
        status = "auto" if vino_id else "unmatched"

        if vino_id:
            n_matched += 1
        else:
            n_unmatched += 1

        fc.execute(
            """INSERT INTO ipratico_product_map
               (ipratico_uuid, ipratico_wine_id, ipratico_name, ipratico_category,
                vino_id, match_status)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (ipratico_uuid, str(wine_id).zfill(4) if wine_id else None,
             name, "Bottiglie", vino_id, status),
        )

    fc.commit()

    # Log sync
    fc.execute(
        """INSERT INTO ipratico_sync_log (direction, filename, n_matched, n_unmatched)
           VALUES ('import', ?, ?, ?)""",
        (file.filename, n_matched, n_unmatched),
    )
    fc.commit()
    fc.close()

    return {
        "total_products": len(df),
        "total_bottiglie": len(bottiglie),
        "matched": n_matched,
        "unmatched": n_unmatched,
        "filename": file.filename,
    }


# ─── Mappings CRUD ─────────────────────────────────────────────────
@router.get("/mappings")
def get_mappings(
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
):
    """Lista mapping iPratico ↔ TRGB con dati arricchiti."""
    fc = _fc_conn()
    rows = fc.execute(
        "SELECT * FROM ipratico_product_map ORDER BY ipratico_wine_id"
    ).fetchall()
    fc.close()

    # Load TRGB wine data per arricchire
    mag_map = {}
    try:
        mconn = _mag_conn()
        for r in mconn.execute(
            "SELECT id, DESCRIZIONE, DENOMINAZIONE, ANNATA, PRODUTTORE, "
            "FORMATO, QTA_TOTALE, PREZZO_CARTA FROM vini_magazzino"
        ).fetchall():
            mag_map[r["id"]] = dict(r)
        mconn.close()
    except Exception:
        pass

    result = []
    for r in rows:
        d = dict(r)
        vid = d.get("vino_id")
        trgb = mag_map.get(vid, {})
        d["trgb_produttore"] = trgb.get("PRODUTTORE", "")
        d["trgb_descrizione"] = trgb.get("DESCRIZIONE", "")
        d["trgb_annata"] = trgb.get("ANNATA", "")
        d["trgb_qta"] = trgb.get("QTA_TOTALE", 0)
        d["trgb_prezzo"] = trgb.get("PREZZO_CARTA", 0)
        d["trgb_formato"] = trgb.get("FORMATO", "")

        # Filter
        if status and d["match_status"] != status:
            continue
        if search:
            s = search.lower()
            searchable = f"{d.get('ipratico_name', '')} {d.get('trgb_produttore', '')} {d.get('trgb_descrizione', '')}".lower()
            if s not in searchable:
                continue

        result.append(d)

    return result


class MappingUpdate(BaseModel):
    vino_id: Optional[int] = None


@router.put("/mappings/{map_id}")
def update_mapping(map_id: int, body: MappingUpdate):
    """Aggiorna mapping: assegna o rimuovi collegamento a vino TRGB."""
    fc = _fc_conn()
    existing = fc.execute("SELECT id FROM ipratico_product_map WHERE id = ?", (map_id,)).fetchone()
    if not existing:
        fc.close()
        raise HTTPException(404, "Mapping non trovato")

    status = "manual" if body.vino_id is not None else "unmatched"

    fc.execute(
        "UPDATE ipratico_product_map SET vino_id = ?, match_status = ?, updated_at = datetime('now') WHERE id = ?",
        (body.vino_id, status, map_id),
    )
    fc.commit()
    fc.close()
    return {"ok": True, "map_id": map_id, "vino_id": body.vino_id, "match_status": status}


@router.put("/ignore/{map_id}")
def ignore_mapping(map_id: int):
    """Segna un prodotto iPratico come 'ignorato' (esiste solo su iPratico, nessun corrispondente TRGB)."""
    fc = _fc_conn()
    existing = fc.execute("SELECT id, match_status FROM ipratico_product_map WHERE id = ?", (map_id,)).fetchone()
    if not existing:
        fc.close()
        raise HTTPException(404, "Mapping non trovato")

    # Toggle: se già ignorato, torna a unmatched
    new_status = "unmatched" if existing["match_status"] == "ignored" else "ignored"
    fc.execute(
        "UPDATE ipratico_product_map SET match_status = ?, updated_at = datetime('now') WHERE id = ?",
        (new_status, map_id),
    )
    fc.commit()
    fc.close()
    return {"ok": True, "map_id": map_id, "match_status": new_status}


# ─── TRGB wines list per match manuale ─────────────────────────────
@router.get("/trgb-wines")
def get_trgb_wines(search: Optional[str] = Query(None)):
    """Lista vini TRGB per dropdown selezione match manuale."""
    wines = []
    try:
        mconn = _mag_conn()
        rows = mconn.execute(
            "SELECT id, DESCRIZIONE, DENOMINAZIONE, ANNATA, PRODUTTORE, FORMATO, QTA_TOTALE, PREZZO_CARTA "
            "FROM vini_magazzino ORDER BY PRODUTTORE, ANNATA"
        ).fetchall()
        wines = [dict(r) for r in rows]
        mconn.close()
    except Exception:
        pass

    if search:
        s = search.lower()
        wines = [w for w in wines if
                 s in (w.get("PRODUTTORE") or "").lower() or
                 s in (w.get("DESCRIZIONE") or "").lower() or
                 s in str(w.get("id", ""))]

    return wines[:200]


# ─── Export Excel ──────────────────────────────────────────────────
@router.post("/export")
async def export_ipratico(file: UploadFile = File(...)):
    """
    Riceve l'export iPratico originale, aggiorna:
    1. Warehouse_quantity (TRGB → iPratico)
    2. Name — ricostruito da TRGB se cambiato (TRGB ha priorità)
    3. Prezzi Ristorante table = PREZZO_CARTA
    4. Aggiunge righe per vini TRGB mancanti su iPratico
    Ritorna l'Excel modificato pronto per import in iPratico.
    """
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Build header map
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx

    name_col = headers.get("Name")
    cat_col = headers.get("Category")
    qty_col = headers.get("Warehouse_quantity")

    if not name_col or not cat_col:
        raise HTTPException(400, "Colonne Name/Category non trovate")

    # Load TRGB data completi da magazzino
    mag_data = {}
    try:
        mconn = _mag_conn()
        for r in mconn.execute(
            "SELECT id, DESCRIZIONE, DENOMINAZIONE, ANNATA, PRODUTTORE, FORMATO, "
            "QTA_TOTALE, PREZZO_CARTA FROM vini_magazzino"
        ).fetchall():
            mag_data[r["id"]] = dict(r)
        mconn.close()
    except Exception:
        pass

    # Price columns (Ristorante table price = PREZZO_CARTA)
    price_table_1_col = headers.get("Price_table_1")

    n_updated_qty = 0
    n_updated_price = 0
    n_updated_name = 0
    n_matched = 0
    existing_wine_ids = set()  # Track wine IDs already in the file

    for row_idx in range(2, ws.max_row + 1):
        cat = ws.cell(row=row_idx, column=cat_col).value
        if cat != "Bottiglie":
            continue

        name = str(ws.cell(row=row_idx, column=name_col).value or "")
        wine_id = _extract_wine_id(name)
        if wine_id:
            existing_wine_ids.add(wine_id)

        if not wine_id or wine_id not in mag_data:
            continue

        n_matched += 1
        trgb = mag_data[wine_id]

        # 1. Update Name (TRGB priority — ricostruisci da TRGB)
        new_name = _build_ipratico_name(trgb)
        old_name = ws.cell(row=row_idx, column=name_col).value or ""
        if new_name != old_name:
            ws.cell(row=row_idx, column=name_col).value = new_name
            n_updated_name += 1

        # 2. Update Warehouse_quantity (TRGB → iPratico)
        if qty_col:
            trgb_qty = trgb.get("QTA_TOTALE", 0) or 0
            old_qty = ws.cell(row=row_idx, column=qty_col).value
            if trgb_qty != old_qty:
                ws.cell(row=row_idx, column=qty_col).value = trgb_qty
                n_updated_qty += 1

        # 3. Update price Ristorante table = PREZZO_CARTA
        if price_table_1_col:
            trgb_price = trgb.get("PREZZO_CARTA")
            old_price = ws.cell(row=row_idx, column=price_table_1_col).value
            if trgb_price and trgb_price > 0 and trgb_price != old_price:
                ws.cell(row=row_idx, column=price_table_1_col).value = trgb_price
                n_updated_price += 1

    # 4. Aggiungi vini TRGB mancanti su iPratico come nuove righe
    #    Campi default letti dalla tabella ipratico_export_defaults (configurabili da frontend)
    n_added = 0
    defaults = _load_export_defaults()

    # Campi prezzo: tutti e 12 a PREZZO_CARTA
    price_fields = (
        "Price_table", "Price_counter", "Price_takeaway", "Price_delivery",
        "Price_table_1", "Price_counter_1", "Price_takeaway_1", "Price_delivery_1",
        "Price_table_2", "Price_counter_2", "Price_takeaway_2", "Price_delivery_2",
    )

    for wine_id, trgb in mag_data.items():
        if wine_id in existing_wine_ids:
            continue
        new_row = ws.max_row + 1
        n_added += 1
        price = trgb.get("PREZZO_CARTA") or 0

        # Campi obbligatori fissi
        ws.cell(row=new_row, column=cat_col).value = "Bottiglie"
        ws.cell(row=new_row, column=name_col).value = _build_ipratico_name(trgb)

        # Campi default dalla tabella configurabile
        for field_name, field_value in defaults.items():
            col = headers.get(field_name)
            if col:
                ws.cell(row=new_row, column=col).value = field_value

        # Prezzi — tutti e 12 campi uguali a PREZZO_CARTA
        for price_col_name in price_fields:
            if headers.get(price_col_name) and price > 0:
                ws.cell(row=new_row, column=headers[price_col_name]).value = price

        # Warehouse_quantity
        if qty_col:
            ws.cell(row=new_row, column=qty_col).value = trgb.get("QTA_TOTALE", 0) or 0

    # Log
    fc = _fc_conn()
    fc.execute(
        """INSERT INTO ipratico_sync_log
           (direction, filename, n_matched, n_updated_qty, n_updated_price)
           VALUES ('export', ?, ?, ?, ?)""",
        (file.filename, n_matched, n_updated_qty, n_updated_price),
    )
    fc.commit()
    fc.close()

    # Return modified Excel
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="ipratico_sync_{ts}.xlsx"',
            "X-Updated-Qty": str(n_updated_qty),
            "X-Updated-Price": str(n_updated_price),
            "X-Updated-Name": str(n_updated_name),
            "X-Total-Matched": str(n_matched),
            "X-Added-Missing": str(n_added),
        },
    )


# ─── Mancanti: vini TRGB non presenti su iPratico ─────────────────
@router.get("/missing")
def get_missing_wines(search: Optional[str] = Query(None)):
    """
    Vini presenti in vini_magazzino ma NON nell'export iPratico (da aggiungere).
    Confronta vini_magazzino.id con ipratico_product_map.vino_id.
    """
    # IDs iPratico già mappati (compresi ignored)
    fc = _fc_conn()
    mapped_rows = fc.execute(
        "SELECT CAST(ipratico_wine_id AS INTEGER) AS wid FROM ipratico_product_map WHERE ipratico_wine_id IS NOT NULL"
    ).fetchall()
    fc.close()
    ipratico_ids = set(r["wid"] for r in mapped_rows if r["wid"])

    # Tutti i vini TRGB
    missing = []
    try:
        mconn = _mag_conn()
        rows = mconn.execute(
            "SELECT id, DESCRIZIONE, DENOMINAZIONE, ANNATA, PRODUTTORE, FORMATO, "
            "QTA_TOTALE, PREZZO_CARTA, IPRATICO FROM vini_magazzino ORDER BY id"
        ).fetchall()
        for r in rows:
            if r["id"] not in ipratico_ids:
                d = dict(r)
                if search:
                    s = search.lower()
                    searchable = f"{d.get('DESCRIZIONE','')} {d.get('PRODUTTORE','')} {d.get('ANNATA','')}".lower()
                    if s not in searchable:
                        continue
                missing.append(d)
        mconn.close()
    except Exception:
        pass

    return missing


# ─── Sync log ─────────────────────────────────────────────────────
@router.get("/sync-log")
def get_sync_log():
    fc = _fc_conn()
    rows = fc.execute("SELECT * FROM ipratico_sync_log ORDER BY created_at DESC LIMIT 50").fetchall()
    fc.close()
    return [dict(r) for r in rows]


# ─── Stats ─────────────────────────────────────────────────────────
@router.get("/stats")
def get_ipratico_stats():
    """Riepilogo veloce: quanti match, unmatched, totali."""
    fc = _fc_conn()
    total = fc.execute("SELECT COUNT(*) FROM ipratico_product_map").fetchone()[0]
    matched = fc.execute(
        "SELECT COUNT(*) FROM ipratico_product_map WHERE vino_id IS NOT NULL AND match_status != 'ignored'"
    ).fetchone()[0]
    auto = fc.execute(
        "SELECT COUNT(*) FROM ipratico_product_map WHERE match_status = 'auto'"
    ).fetchone()[0]
    manual = fc.execute(
        "SELECT COUNT(*) FROM ipratico_product_map WHERE match_status = 'manual'"
    ).fetchone()[0]
    ignored = fc.execute(
        "SELECT COUNT(*) FROM ipratico_product_map WHERE match_status = 'ignored'"
    ).fetchone()[0]
    unmatched = fc.execute(
        "SELECT COUNT(*) FROM ipratico_product_map WHERE match_status = 'unmatched'"
    ).fetchone()[0]
    fc.close()

    # Count missing (TRGB wines not in iPratico)
    n_missing = 0
    ipratico_ids = set()
    fc2 = _fc_conn()
    for r in fc2.execute("SELECT CAST(ipratico_wine_id AS INTEGER) AS wid FROM ipratico_product_map WHERE ipratico_wine_id IS NOT NULL").fetchall():
        if r["wid"]:
            ipratico_ids.add(r["wid"])
    fc2.close()
    if ipratico_ids:
        try:
            mconn = _mag_conn()
            all_ids = set(r["id"] for r in mconn.execute("SELECT id FROM vini_magazzino").fetchall())
            n_missing = len(all_ids - ipratico_ids)
            mconn.close()
        except Exception:
            pass

    return {
        "total": total,
        "matched": matched,
        "auto": auto,
        "manual": manual,
        "ignored": ignored,
        "unmatched": unmatched,
        "missing": n_missing,
    }


# ─── Export defaults (configurabili da frontend) ──────────────────
@router.get("/export-defaults")
def get_export_defaults():
    """Lista valori di default per campi vini nuovi nell'export."""
    fc = _fc_conn()
    try:
        rows = fc.execute(
            "SELECT id, field_name, field_value, field_group, label FROM ipratico_export_defaults ORDER BY field_group, id"
        ).fetchall()
        return [dict(r) for r in rows]
    except Exception:
        return []
    finally:
        fc.close()


class DefaultUpdate(BaseModel):
    field_value: str


@router.put("/export-defaults/{default_id}")
def update_export_default(default_id: int, body: DefaultUpdate):
    """Aggiorna il valore di un campo default export."""
    fc = _fc_conn()
    existing = fc.execute("SELECT id FROM ipratico_export_defaults WHERE id = ?", (default_id,)).fetchone()
    if not existing:
        fc.close()
        raise HTTPException(404, "Default non trovato")
    fc.execute(
        "UPDATE ipratico_export_defaults SET field_value = ?, updated_at = datetime('now') WHERE id = ?",
        (body.field_value, default_id),
    )
    fc.commit()
    fc.close()
    return {"ok": True, "id": default_id, "field_value": body.field_value}
