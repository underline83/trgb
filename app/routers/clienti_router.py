# ============================================================
# FILE: app/routers/clienti_router.py
# Router Clienti CRM — TRGB Gestionale
# ============================================================

# @version: v1.0-clienti-router
# -*- coding: utf-8 -*-
"""
Router Clienti CRM — TRGB Gestionale

Funzionalità v1.0:
- Anagrafica clienti (CRUD, ricerca full-text, filtri)
- Import da TheFork (XLSX)
- Tag/categorie (CRUD + associazione)
- Note/diario interazioni (CRUD)
- Dashboard statistiche CRM
- Compleanni in arrivo

Autenticazione:
- Tutti gli endpoint richiedono utente loggato (JWT).
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import JSONResponse
from pydantic import BaseModel, Field

from app.models.clienti_db import get_clienti_conn, init_clienti_db
from app.services.auth_service import get_current_user

logger = logging.getLogger("trgb.clienti")

router = APIRouter(prefix="/clienti", tags=["Clienti"])

# Inizializza DB alla prima importazione del router
init_clienti_db()


# ============================================================
# MODELLI Pydantic
# ============================================================
class ClienteBase(BaseModel):
    titolo: Optional[str] = None
    nome: str
    cognome: str
    email: Optional[str] = None
    telefono: Optional[str] = None
    telefono2: Optional[str] = None
    data_nascita: Optional[str] = None
    lingua: Optional[str] = "it_IT"
    indirizzo: Optional[str] = None
    cap: Optional[str] = None
    citta: Optional[str] = None
    paese: Optional[str] = "Italy"
    vip: bool = False
    rank: Optional[str] = None
    promoter: bool = False
    newsletter: bool = False
    risk_level: Optional[str] = None
    pref_cibo: Optional[str] = None
    pref_bevande: Optional[str] = None
    pref_posto: Optional[str] = None
    restrizioni_dietetiche: Optional[str] = None
    allergie: Optional[str] = None
    note_thefork: Optional[str] = None
    attivo: bool = True
    origine: Optional[str] = "manuale"
    nome2: Optional[str] = None
    cognome2: Optional[str] = None


class ClienteCreate(ClienteBase):
    pass


class ClienteUpdate(ClienteBase):
    pass


class TagBase(BaseModel):
    nome: str
    colore: str = "#0d9488"
    ordine: int = 0


class NotaCreate(BaseModel):
    tipo: str = "nota"
    testo: str
    data: Optional[str] = None
    autore: Optional[str] = None


# ============================================================
# ENDPOINT: DASHBOARD CRM (DEVE stare PRIMA di /{cliente_id})
# ============================================================
@router.get("/dashboard/stats")
def dashboard_stats(current_user: Dict[str, Any] = Depends(get_current_user)):
    conn = get_clienti_conn()
    try:
        stats = {}

        # Totali
        stats["totale"] = conn.execute("SELECT COUNT(*) FROM clienti WHERE attivo = 1").fetchone()[0]
        stats["vip"] = conn.execute("SELECT COUNT(*) FROM clienti WHERE attivo = 1 AND vip = 1").fetchone()[0]
        stats["con_email"] = conn.execute("SELECT COUNT(*) FROM clienti WHERE attivo = 1 AND email IS NOT NULL AND email != ''").fetchone()[0]
        stats["con_telefono"] = conn.execute("SELECT COUNT(*) FROM clienti WHERE attivo = 1 AND telefono IS NOT NULL AND telefono != ''").fetchone()[0]
        stats["con_allergie"] = conn.execute("SELECT COUNT(*) FROM clienti WHERE attivo = 1 AND (allergie IS NOT NULL AND allergie != '')").fetchone()[0]
        stats["con_preferenze"] = conn.execute("SELECT COUNT(*) FROM clienti WHERE attivo = 1 AND (pref_cibo IS NOT NULL AND pref_cibo != '')").fetchone()[0]
        stats["con_compleanno"] = conn.execute("SELECT COUNT(*) FROM clienti WHERE attivo = 1 AND data_nascita IS NOT NULL AND data_nascita != ''").fetchone()[0]

        # Per rank
        ranks = conn.execute(
            "SELECT COALESCE(rank, 'Nessuno') as rank, COUNT(*) as n FROM clienti WHERE attivo = 1 GROUP BY rank ORDER BY n DESC"
        ).fetchall()
        stats["per_rank"] = [dict(r) for r in ranks]

        # Per lingua
        lingue = conn.execute(
            "SELECT COALESCE(lingua, 'n/d') as lingua, COUNT(*) as n FROM clienti WHERE attivo = 1 GROUP BY lingua ORDER BY n DESC LIMIT 10"
        ).fetchall()
        stats["per_lingua"] = [dict(r) for r in lingue]

        # Per tag
        tags = conn.execute(
            """
            SELECT t.nome, t.colore, COUNT(ta.cliente_id) as n
            FROM clienti_tag t
            LEFT JOIN clienti_tag_assoc ta ON ta.tag_id = t.id
            GROUP BY t.id
            ORDER BY t.ordine
            """
        ).fetchall()
        stats["per_tag"] = [dict(r) for r in tags]

        # Nuovi ultimi 30 giorni
        stats["nuovi_30gg"] = conn.execute(
            "SELECT COUNT(*) FROM clienti WHERE created_at >= date('now', '-30 days')"
        ).fetchone()[0]

        # Compleanni prossimi 7 giorni
        today = date.today()
        birthday_conditions = []
        for i in range(8):
            d = today + timedelta(days=i)
            birthday_conditions.append(f"substr(data_nascita, 1, 5) = '{d.strftime('%d/%m')}'")
        compleanni = conn.execute(
            f"""
            SELECT id, nome, cognome, data_nascita, telefono, email
            FROM clienti
            WHERE attivo = 1 AND data_nascita IS NOT NULL
              AND ({' OR '.join(birthday_conditions)})
            ORDER BY substr(data_nascita, 4, 2), substr(data_nascita, 1, 2)
            """
        ).fetchall()
        stats["compleanni_prossimi"] = [dict(r) for r in compleanni]

        return JSONResponse(stats)
    except Exception as e:
        logger.exception("Errore dashboard stats")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


# ============================================================
# ENDPOINT: TAG (DEVE stare PRIMA di /{cliente_id})
# ============================================================
@router.get("/tag/lista")
def lista_tag(current_user: Dict[str, Any] = Depends(get_current_user)):
    conn = get_clienti_conn()
    try:
        rows = conn.execute("SELECT * FROM clienti_tag ORDER BY ordine, nome").fetchall()
        return JSONResponse([dict(r) for r in rows])
    finally:
        conn.close()


@router.post("/tag")
def crea_tag(body: TagBase, current_user: Dict[str, Any] = Depends(get_current_user)):
    conn = get_clienti_conn()
    try:
        cur = conn.execute(
            "INSERT INTO clienti_tag (nome, colore, ordine) VALUES (?,?,?)",
            (body.nome, body.colore, body.ordine),
        )
        conn.commit()
        return JSONResponse({"id": cur.lastrowid, "status": "ok"})
    except Exception as e:
        logger.exception("Errore creazione tag")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


@router.delete("/tag/{tag_id}")
def elimina_tag(tag_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    conn = get_clienti_conn()
    try:
        conn.execute("DELETE FROM clienti_tag_assoc WHERE tag_id = ?", (tag_id,))
        conn.execute("DELETE FROM clienti_tag WHERE id = ?", (tag_id,))
        conn.commit()
        return JSONResponse({"status": "ok"})
    finally:
        conn.close()


# ============================================================
# ENDPOINT: EXPORT CSV (Google Contacts / Gmail compatibile)
# ============================================================
@router.get("/export/google-csv")
def export_google_csv(
    solo_attivi: bool = Query(True),
    solo_con_contatto: bool = Query(True),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Esporta i clienti in formato CSV compatibile con Google Contacts.
    Colonne: Name, Given Name, Family Name, E-mail 1 - Value,
             Phone 1 - Value, Phone 2 - Value, Birthday, Notes,
             Group Membership
    """
    import csv
    from io import StringIO
    from fastapi.responses import StreamingResponse

    conn = get_clienti_conn()
    try:
        where = []
        if solo_attivi:
            where.append("c.attivo = 1")
        if solo_con_contatto:
            where.append("(c.email IS NOT NULL AND c.email != '' OR c.telefono IS NOT NULL AND c.telefono != '')")

        where_sql = f"WHERE {' AND '.join(where)}" if where else ""

        rows = conn.execute(f"""
            SELECT c.*,
                   GROUP_CONCAT(t.nome, ' ::: ') as tag_nomi
            FROM clienti c
            LEFT JOIN clienti_tag_assoc ta ON ta.cliente_id = c.id
            LEFT JOIN clienti_tag t ON t.id = ta.tag_id
            {where_sql}
            GROUP BY c.id
            ORDER BY c.cognome, c.nome
        """).fetchall()

        output = StringIO()
        writer = csv.writer(output)

        # Header Google Contacts
        writer.writerow([
            "Name", "Given Name", "Family Name",
            "E-mail 1 - Type", "E-mail 1 - Value",
            "Phone 1 - Type", "Phone 1 - Value",
            "Phone 2 - Type", "Phone 2 - Value",
            "Birthday", "Notes",
            "Group Membership",
        ])

        for r in rows:
            nome = r["nome"] or ""
            cognome = r["cognome"] or ""
            full_name = f"{nome} {cognome}".strip()

            # Tag → Google Groups (es. "* TRGB ::: VIP ::: Abituale")
            tags = r["tag_nomi"].split(" ::: ") if r["tag_nomi"] else []
            groups = " ::: ".join(["* TRGB"] + tags) if tags else "* TRGB"

            # Note combinate
            note_parts = []
            if r["allergie"]:
                note_parts.append(f"Allergie: {r['allergie']}")
            if r["pref_cibo"]:
                note_parts.append(f"Cibo: {r['pref_cibo']}")
            if r["pref_bevande"]:
                note_parts.append(f"Bevande: {r['pref_bevande']}")
            if r["restrizioni_dietetiche"]:
                note_parts.append(f"Dieta: {r['restrizioni_dietetiche']}")
            if r["note_thefork"]:
                note_parts.append(r["note_thefork"])
            notes = " | ".join(note_parts) if note_parts else ""

            # Birthday: TheFork usa dd/mm/yyyy, Google usa yyyy-mm-dd
            bday = ""
            if r["data_nascita"]:
                try:
                    parts = str(r["data_nascita"]).split("/")
                    if len(parts) == 3:
                        bday = f"{parts[2]}-{parts[1]}-{parts[0]}"
                except Exception:
                    bday = str(r["data_nascita"])

            writer.writerow([
                full_name, nome, cognome,
                "* Other", r["email"] or "",
                "* Mobile", r["telefono"] or "",
                "Other", r["telefono2"] or "",
                bday, notes,
                groups,
            ])

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=trgb_clienti_google_{date.today().isoformat()}.csv"
            },
        )
    finally:
        conn.close()


# ============================================================
# ENDPOINT: IMPORT THEFORK XLSX (DEVE stare PRIMA di /{cliente_id})
# ============================================================
@router.post("/import/thefork")
async def import_thefork(
    file: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Importa clienti da export TheFork (XLSX).
    Usa thefork_id come chiave univoca per evitare duplicati.
    Aggiorna i record esistenti se già presenti.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl non installato sul server")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "File vuoto")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    def val(row, name):
        idx = col(name)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        return str(v).strip() if not isinstance(v, (int, float)) else v

    def clean_phone(p):
        if p is None:
            return None
        s = str(p).replace(".0", "").replace(" ", "").replace("-", "")
        if s and not s.startswith("+"):
            if s.startswith("39") and len(s) > 10:
                s = "+" + s
            elif len(s) == 10:
                s = "+39" + s
        return s if s else None

    def bool_val(v):
        if v is None:
            return 0
        return 1 if str(v).strip().lower() in ("yes", "si", "sì", "1", "true") else 0

    conn = get_clienti_conn()
    inseriti = 0
    aggiornati = 0
    errori = 0
    diff_trovati = 0

    try:
        for row in data_rows:
            try:
                tf_id = val(row, "Id")
                nome = val(row, "First Name")
                cognome = val(row, "Last Name")
                if not nome and not cognome:
                    continue

                telefono = clean_phone(val(row, "Phone number"))
                telefono2 = clean_phone(val(row, "Secondary phone number"))

                record = {
                    "thefork_id": tf_id,
                    "titolo": val(row, "Title"),
                    "nome": nome or "",
                    "cognome": cognome or "",
                    "email": val(row, "Email") if isinstance(val(row, "Email"), str) else None,
                    "telefono": telefono,
                    "telefono2": telefono2,
                    "data_nascita": val(row, "Birthday"),
                    "lingua": val(row, "Language") or "it_IT",
                    "indirizzo": val(row, "Address"),
                    "cap": str(int(val(row, "Zipcode"))) if val(row, "Zipcode") and val(row, "Zipcode") != "None" else None,
                    "citta": val(row, "City"),
                    "paese": val(row, "Country") or "Italy",
                    "vip": bool_val(val(row, "VIP")),
                    "rank": val(row, "Rank"),
                    "promoter": bool_val(val(row, "Promoter")),
                    "newsletter": bool_val(val(row, "Newsletter")),
                    "risk_level": val(row, "Risk level"),
                    "spending_behaviour": float(val(row, "Spending behaviour")) if val(row, "Spending behaviour") else None,
                    "pref_cibo": val(row, "Food preferences"),
                    "pref_bevande": val(row, "Drinks preferences"),
                    "pref_posto": str(val(row, "Seating preferences")) if val(row, "Seating preferences") else None,
                    "restrizioni_dietetiche": val(row, "Dietary restrictions"),
                    "allergie": val(row, "Allergies and intolerances"),
                    "note_thefork": val(row, "Customer Notes"),
                    "origine": "thefork",
                    "thefork_created": val(row, "Creation date"),
                    "thefork_updated": val(row, "Last update date"),
                }

                # Upsert: cerca per thefork_id diretto O tramite alias (merge)
                existing = None
                if tf_id:
                    # 1. Cerca nel campo thefork_id principale
                    existing = conn.execute(
                        "SELECT id, protetto FROM clienti WHERE thefork_id = ?", (tf_id,)
                    ).fetchone()
                    # 2. Se non trovato, cerca negli alias (clienti mergati)
                    if not existing:
                        alias_row = conn.execute(
                            "SELECT cliente_id FROM clienti_alias WHERE thefork_id = ?", (tf_id,)
                        ).fetchone()
                        if alias_row:
                            existing = conn.execute(
                                "SELECT id, protetto FROM clienti WHERE id = ?", (alias_row["cliente_id"],)
                            ).fetchone()

                if existing:
                    if existing["protetto"]:
                        # Cliente protetto: aggiorna sempre i campi TheFork-specifici
                        # + riempi campi vuoti + salva DIFFERENZE nella coda revisione
                        always_update = {
                            "rank": record["rank"],
                            "risk_level": record["risk_level"],
                            "spending_behaviour": record["spending_behaviour"],
                            "thefork_updated": record["thefork_updated"],
                        }
                        # Campi confrontabili: vuoto→riempi, diverso→salva diff
                        diffable = [
                            "email", "telefono", "telefono2", "data_nascita",
                            "indirizzo", "cap", "citta", "paese",
                            "pref_cibo", "pref_bevande", "pref_posto",
                            "restrizioni_dietetiche", "allergie", "note_thefork",
                            "nome", "cognome", "titolo", "lingua",
                        ]
                        current = conn.execute(
                            f"SELECT {', '.join(diffable)} FROM clienti WHERE id = ?",
                            (existing["id"],),
                        ).fetchone()
                        for field in diffable:
                            cur_val = current[field] if current else None
                            new_val = record.get(field)
                            cur_str = str(cur_val).strip() if cur_val else ""
                            new_str = str(new_val).strip() if new_val else ""

                            if not cur_str and new_str:
                                # Campo vuoto nel DB, pieno in TheFork → riempi
                                always_update[field] = new_val
                            elif cur_str and new_str and cur_str != new_str:
                                # Campo diverso → salva diff per revisione Marco
                                conn.execute(
                                    "DELETE FROM clienti_import_diff WHERE cliente_id = ? AND campo = ? AND stato = 'pending'",
                                    (existing["id"], field),
                                )
                                conn.execute(
                                    """INSERT INTO clienti_import_diff
                                       (cliente_id, campo, valore_crm, valore_thefork)
                                       VALUES (?, ?, ?, ?)""",
                                    (existing["id"], field, cur_str, new_str),
                                )
                                diff_trovati += 1

                        set_clause = ", ".join(f"{k}=?" for k in always_update)
                        conn.execute(
                            f"UPDATE clienti SET {set_clause} WHERE id = ?",
                            list(always_update.values()) + [existing["id"]],
                        )
                    else:
                        # Cliente non protetto: TheFork sovrascrive tutto
                        skip = {"thefork_id", "protetto"}
                        set_clause = ", ".join(f"{k}=?" for k in record if k not in skip)
                        values = [v for k, v in record.items() if k not in skip]
                        conn.execute(
                            f"UPDATE clienti SET {set_clause} WHERE id = ?",
                            values + [existing["id"]],
                        )
                    aggiornati += 1
                else:
                    cols = ", ".join(record.keys())
                    placeholders = ", ".join("?" for _ in record)
                    conn.execute(
                        f"INSERT INTO clienti ({cols}) VALUES ({placeholders})",
                        list(record.values()),
                    )
                    inseriti += 1

            except Exception as row_err:
                logger.warning(f"Errore riga import TheFork: {row_err}")
                errori += 1

        conn.commit()

        # Auto-tag VIP (auto=1 → tag automatico, rimovibile dall'import successivo)
        # Non tocca i tag manuali (auto=0) assegnati nel CRM
        conn.execute("""
            INSERT OR IGNORE INTO clienti_tag_assoc (cliente_id, tag_id, auto)
            SELECT c.id, t.id, 1
            FROM clienti c, clienti_tag t
            WHERE c.vip = 1 AND t.nome = 'VIP'
        """)
        # Rimuovi auto-tag VIP da chi non è più VIP in TheFork
        # (solo se il tag era automatico, non se aggiunto manualmente)
        conn.execute("""
            DELETE FROM clienti_tag_assoc
            WHERE auto = 1 AND tag_id = (SELECT id FROM clienti_tag WHERE nome = 'VIP')
            AND cliente_id NOT IN (SELECT id FROM clienti WHERE vip = 1)
        """)
        conn.commit()

        return JSONResponse({
            "status": "ok",
            "inseriti": inseriti,
            "aggiornati": aggiornati,
            "errori": errori,
            "diff_trovati": diff_trovati,
            "totale_righe": len(data_rows),
        })
    except Exception as e:
        logger.exception("Errore import TheFork")
        raise HTTPException(500, str(e))
    finally:
        conn.close()
        wb.close()


# ============================================================
# ENDPOINT: IMPORT PRENOTAZIONI THEFORK XLSX
# ============================================================
@router.post("/import/prenotazioni")
async def import_prenotazioni(
    file: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Importa prenotazioni da export TheFork (XLSX).
    Usa thefork_booking_id come chiave per evitare duplicati.
    Collega automaticamente al cliente tramite Customer ID.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl non installato sul server")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "File vuoto")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    def val(row, name):
        idx = col(name)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return v
        return str(v).strip()

    # Colonne con nome lungo (risposte form TheFork)
    col_degustazione = None
    col_allergie_form = None
    col_tavolo_esterno = None
    col_seggioloni = None
    for i, h in enumerate(headers):
        hl = h.lower()
        if "degustazioni" in hl and "percorsi" in hl:
            col_degustazione = i
        elif "allergie o intolleranze" in hl:
            col_allergie_form = i
        elif "tavolo esterno" in hl:
            col_tavolo_esterno = i
        elif "seggioloni" in hl:
            col_seggioloni = i

    conn = get_clienti_conn()
    inseriti = 0
    aggiornati = 0
    errori = 0
    collegati = 0

    try:
        for row in data_rows:
            try:
                booking_id = val(row, "Booking ID")
                customer_id = val(row, "Customer ID")
                data_pasto = val(row, "Meal date")
                stato = val(row, "Status")
                if not data_pasto or not stato:
                    continue

                pax = val(row, "Pax")
                if isinstance(pax, str):
                    try:
                        pax = int(pax)
                    except ValueError:
                        pax = 2

                # Importo conto: "271 EUR" → "271 EUR"
                bill = val(row, "Bill amount")
                imprint_amt = val(row, "Imprint amount")

                # Risposte form (accesso sicuro: le tuple openpyxl possono essere più corte)
                def safe_col(r, idx):
                    if idx is None or idx >= len(r):
                        return None
                    return r[idx]

                degust_raw = safe_col(row, col_degustazione)
                degust = str(degust_raw).strip() if degust_raw else None
                allergie_raw = safe_col(row, col_allergie_form)
                allergie_form = str(allergie_raw).strip() if allergie_raw else None
                tav_est_raw = safe_col(row, col_tavolo_esterno)
                try:
                    tav_est = 1 if tav_est_raw and float(tav_est_raw) == 1 else 0
                except (ValueError, TypeError):
                    tav_est = 0
                segg_raw = safe_col(row, col_seggioloni)
                segg = str(segg_raw).strip() if segg_raw else None

                # Trova cliente_id interno dal thefork_id (o alias se mergato)
                cliente_id = None
                if customer_id:
                    cli_row = conn.execute(
                        "SELECT id FROM clienti WHERE thefork_id = ?", (customer_id,)
                    ).fetchone()
                    if not cli_row:
                        # Cerca negli alias (clienti mergati)
                        alias_row = conn.execute(
                            "SELECT cliente_id FROM clienti_alias WHERE thefork_id = ?", (customer_id,)
                        ).fetchone()
                        if alias_row:
                            cli_row = conn.execute(
                                "SELECT id FROM clienti WHERE id = ?", (alias_row["cliente_id"],)
                            ).fetchone()
                    if cli_row:
                        cliente_id = cli_row["id"]
                        collegati += 1

                record = {
                    "cliente_id": cliente_id,
                    "thefork_customer_id": customer_id,
                    "thefork_booking_id": booking_id,
                    "data_pasto": str(data_pasto),
                    "ora_pasto": str(val(row, "Meal time") or ""),
                    "stato": stato,
                    "pax": pax or 2,
                    "tavolo": val(row, "Table name"),
                    "canale": val(row, "Booking channel"),
                    "occasione": val(row, "Occasions"),
                    "nota_ristorante": val(row, "Note"),
                    "nota_cliente": val(row, "Note from the customer"),
                    "data_prenotazione": val(row, "Booking taken"),
                    "prenotato_da": val(row, "Booking taken by"),
                    "importo_conto": str(bill) if bill else None,
                    "sconto": float(val(row, "Discount amount")) if val(row, "Discount amount") else None,
                    "menu_preset": val(row, "Preset menu"),
                    "offerta_speciale": 1 if val(row, "Special offer") else 0,
                    "yums": 1 if val(row, "Yums") else 0,
                    "imprint": 1 if val(row, "Booking with Imprint") else 0,
                    "importo_imprint": str(imprint_amt) if imprint_amt else None,
                    "degustazione": degust if degust and degust != "nan" else None,
                    "allergie_segnalate": allergie_form if allergie_form and allergie_form != "nan" else None,
                    "tavolo_esterno": tav_est,
                    "seggioloni": segg if segg and segg != "nan" else None,
                    "waiting_list": 1 if val(row, "Waiting list") else 0,
                }

                # Upsert su booking_id
                existing = None
                if booking_id:
                    existing = conn.execute(
                        "SELECT id FROM clienti_prenotazioni WHERE thefork_booking_id = ?",
                        (booking_id,),
                    ).fetchone()

                if existing:
                    set_clause = ", ".join(f"{k}=?" for k in record if k != "thefork_booking_id")
                    values = [v for k, v in record.items() if k != "thefork_booking_id"]
                    conn.execute(
                        f"UPDATE clienti_prenotazioni SET {set_clause} WHERE thefork_booking_id = ?",
                        values + [booking_id],
                    )
                    aggiornati += 1
                else:
                    cols = ", ".join(record.keys())
                    placeholders = ", ".join("?" for _ in record)
                    conn.execute(
                        f"INSERT INTO clienti_prenotazioni ({cols}) VALUES ({placeholders})",
                        list(record.values()),
                    )
                    inseriti += 1

            except Exception as row_err:
                logger.warning(f"Errore riga import prenotazioni: {row_err}")
                errori += 1

        conn.commit()

        return JSONResponse({
            "status": "ok",
            "inseriti": inseriti,
            "aggiornati": aggiornati,
            "errori": errori,
            "collegati_a_clienti": collegati,
            "totale_righe": len(data_rows),
        })
    except Exception as e:
        logger.exception("Errore import prenotazioni")
        raise HTTPException(500, str(e))
    finally:
        conn.close()
        wb.close()


# ============================================================
# ENDPOINT: MERGE DUPLICATI
# ============================================================
class MergeRequest(BaseModel):
    principale_id: int = Field(..., description="ID del cliente da mantenere come principale")
    secondario_id: int = Field(..., description="ID del cliente duplicato da assorbire")


@router.post("/merge")
def merge_clienti(
    req: MergeRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Unisce due clienti: il secondario viene assorbito dal principale.

    Cosa succede:
    1. Le prenotazioni del secondario passano al principale
    2. Le note del secondario passano al principale
    3. I tag del secondario vengono copiati al principale (se non già presenti)
    4. Il thefork_id del secondario va in clienti_alias (così i futuri import lo riconoscono)
    5. Il secondario viene eliminato
    6. Il principale diventa "protetto" (l'import TheFork non lo sovrascrive)
    """
    if req.principale_id == req.secondario_id:
        raise HTTPException(400, "Non puoi unire un cliente con se stesso")

    conn = get_clienti_conn()
    try:
        # Verifica che entrambi esistano
        princ = conn.execute("SELECT * FROM clienti WHERE id = ?", (req.principale_id,)).fetchone()
        sec = conn.execute("SELECT * FROM clienti WHERE id = ?", (req.secondario_id,)).fetchone()
        if not princ:
            raise HTTPException(404, f"Cliente principale {req.principale_id} non trovato")
        if not sec:
            raise HTTPException(404, f"Cliente secondario {req.secondario_id} non trovato")

        # 1. Sposta prenotazioni
        conn.execute(
            "UPDATE clienti_prenotazioni SET cliente_id = ? WHERE cliente_id = ?",
            (req.principale_id, req.secondario_id),
        )

        # 2. Sposta note
        conn.execute(
            "UPDATE clienti_note SET cliente_id = ? WHERE cliente_id = ?",
            (req.principale_id, req.secondario_id),
        )

        # 3. Copia tag (solo quelli che il principale non ha già)
        conn.execute("""
            INSERT OR IGNORE INTO clienti_tag_assoc (cliente_id, tag_id, auto)
            SELECT ?, tag_id, auto FROM clienti_tag_assoc WHERE cliente_id = ?
        """, (req.principale_id, req.secondario_id))

        # 4. Salva thefork_id del secondario come alias
        if sec["thefork_id"]:
            conn.execute("""
                INSERT OR IGNORE INTO clienti_alias (cliente_id, thefork_id, merged_from_id)
                VALUES (?, ?, ?)
            """, (req.principale_id, sec["thefork_id"], req.secondario_id))

        # Salva anche eventuali alias che il secondario aveva
        conn.execute(
            "UPDATE clienti_alias SET cliente_id = ? WHERE cliente_id = ?",
            (req.principale_id, req.secondario_id),
        )

        # 4b. Unisci campi complementari: se il principale ha un campo vuoto
        #     e il secondario ce l'ha, copialo nel principale
        fill_fields = [
            "email", "telefono", "telefono2", "data_nascita",
            "indirizzo", "cap", "citta", "paese",
            "pref_cibo", "pref_bevande", "pref_posto",
            "restrizioni_dietetiche", "allergie", "note_thefork",
        ]
        updates = []
        values = []
        for field in fill_fields:
            princ_val = princ[field]
            sec_val = sec[field]
            if (not princ_val or str(princ_val).strip() == "") and sec_val and str(sec_val).strip() != "":
                updates.append(f"{field} = ?")
                values.append(sec_val)
        if updates:
            conn.execute(
                f"UPDATE clienti SET {', '.join(updates)} WHERE id = ?",
                values + [req.principale_id],
            )

        # Rimuovi eventuali esclusioni duplicati che coinvolgono il secondario
        conn.execute(
            "DELETE FROM clienti_no_duplicato WHERE cliente_a = ? OR cliente_b = ?",
            (req.secondario_id, req.secondario_id),
        )

        # 5. Elimina il secondario
        conn.execute("DELETE FROM clienti WHERE id = ?", (req.secondario_id,))

        # 6. Segna il principale come protetto
        conn.execute("UPDATE clienti SET protetto = 1 WHERE id = ?", (req.principale_id,))

        conn.commit()

        # Conta prenotazioni totali del principale dopo il merge
        tot_pren = conn.execute(
            "SELECT COUNT(*) FROM clienti_prenotazioni WHERE cliente_id = ?",
            (req.principale_id,),
        ).fetchone()[0]

        return JSONResponse({
            "status": "ok",
            "message": f"Merge completato: {sec['cognome']} {sec['nome']} → {princ['cognome']} {princ['nome']}",
            "principale_id": req.principale_id,
            "prenotazioni_totali": tot_pren,
        })

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Errore merge clienti")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


# ============================================================
# AUTO-MERGE DUPLICATI OVVI
# ============================================================
def _find_obvious_duplicates(conn):
    """
    Trova gruppi di duplicati "ovvi" — stessa email+cognome oppure stesso telefono+cognome.
    Restituisce lista di dict: { principale, secondari, motivo, dettagli }.
    Rispetta le esclusioni in clienti_no_duplicato.
    """
    from itertools import combinations

    # Carica coppie escluse
    excl_rows = conn.execute("SELECT cliente_a, cliente_b FROM clienti_no_duplicato").fetchall()
    excluded = set(frozenset([r["cliente_a"], r["cliente_b"]]) for r in excl_rows)

    def is_excluded(ids):
        if len(ids) == 2:
            return frozenset(ids) in excluded
        return all(frozenset(p) in excluded for p in combinations(ids, 2))

    def pick_principale(ids):
        """Sceglie il principale: piu prenotazioni > protetto > ID piu basso."""
        rows = []
        for cid in ids:
            c = conn.execute("""
                SELECT c.id, c.cognome, c.nome, c.telefono, c.email, c.protetto,
                       (SELECT COUNT(*) FROM clienti_prenotazioni WHERE cliente_id = c.id) as n_pren
                FROM clienti c WHERE c.id = ?
            """, (cid,)).fetchone()
            if c:
                rows.append(dict(c))
        if not rows:
            return None, [], []
        rows.sort(key=lambda r: (-r["n_pren"], -(r["protetto"] or 0), r["id"]))
        princ = rows[0]
        secondari = [r for r in rows[1:]]
        return princ, secondari, rows

    groups = []
    seen = set()

    # Gruppo 1: stesso telefono + stesso cognome (case-insensitive)
    tel_groups = conn.execute("""
        SELECT GROUP_CONCAT(id) as ids, telefono, LOWER(cognome) as lcog
        FROM clienti
        WHERE telefono IS NOT NULL AND telefono != '' AND cognome IS NOT NULL AND cognome != ''
          AND attivo = 1
        GROUP BY telefono, LOWER(cognome)
        HAVING COUNT(*) > 1
    """).fetchall()
    for row in tel_groups:
        ids = [int(x) for x in row["ids"].split(",")]
        key = frozenset(ids)
        if key in seen or is_excluded(ids):
            continue
        seen.add(key)
        princ, secondari, dettagli = pick_principale(ids)
        if princ and secondari:
            groups.append({
                "principale": princ,
                "secondari": secondari,
                "dettagli": dettagli,
                "motivo": f"telefono ({row['telefono']}) + cognome",
            })

    # Gruppo 2: stessa email + stesso cognome (case-insensitive)
    email_groups = conn.execute("""
        SELECT GROUP_CONCAT(id) as ids, LOWER(email) as lemail, LOWER(cognome) as lcog
        FROM clienti
        WHERE email IS NOT NULL AND email != '' AND cognome IS NOT NULL AND cognome != ''
          AND attivo = 1
        GROUP BY LOWER(email), LOWER(cognome)
        HAVING COUNT(*) > 1
    """).fetchall()
    for row in email_groups:
        ids = [int(x) for x in row["ids"].split(",")]
        key = frozenset(ids)
        if key in seen or is_excluded(ids):
            continue
        seen.add(key)
        princ, secondari, dettagli = pick_principale(ids)
        if princ and secondari:
            groups.append({
                "principale": princ,
                "secondari": secondari,
                "dettagli": dettagli,
                "motivo": f"email ({row['lemail']}) + cognome",
            })

    return groups


@router.get("/merge/auto-preview")
def auto_merge_preview(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Preview dei duplicati ovvi che verrebbero auto-uniti."""
    conn = get_clienti_conn()
    try:
        groups = _find_obvious_duplicates(conn)
        result = []
        for g in groups:
            result.append({
                "motivo": g["motivo"],
                "principale": {
                    "id": g["principale"]["id"],
                    "cognome": g["principale"]["cognome"],
                    "nome": g["principale"]["nome"],
                    "prenotazioni": g["principale"]["n_pren"],
                },
                "secondari": [{
                    "id": s["id"],
                    "cognome": s["cognome"],
                    "nome": s["nome"],
                    "prenotazioni": s["n_pren"],
                } for s in g["secondari"]],
            })
        return JSONResponse({
            "gruppi": result,
            "totale_gruppi": len(result),
            "totale_secondari": sum(len(g["secondari"]) for g in result),
        })
    finally:
        conn.close()


@router.post("/merge/auto")
def auto_merge_execute(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Esegue l'auto-merge di tutti i duplicati ovvi.
    Per ogni gruppo: il principale assorbe i secondari (stessa logica di /merge).
    """
    conn = get_clienti_conn()
    try:
        groups = _find_obvious_duplicates(conn)
        merged_count = 0
        errors = []

        fill_fields = [
            "email", "telefono", "telefono2", "data_nascita",
            "indirizzo", "cap", "citta", "paese",
            "pref_cibo", "pref_bevande", "pref_posto",
            "restrizioni_dietetiche", "allergie", "note_thefork",
        ]

        for g in groups:
            pid = g["principale"]["id"]
            for sec in g["secondari"]:
                sid = sec["id"]
                try:
                    # Verifica che entrambi esistano ancora
                    princ_row = conn.execute("SELECT * FROM clienti WHERE id = ?", (pid,)).fetchone()
                    sec_row = conn.execute("SELECT * FROM clienti WHERE id = ?", (sid,)).fetchone()
                    if not princ_row or not sec_row:
                        continue

                    conn.execute("UPDATE clienti_prenotazioni SET cliente_id = ? WHERE cliente_id = ?", (pid, sid))
                    conn.execute("UPDATE clienti_note SET cliente_id = ? WHERE cliente_id = ?", (pid, sid))
                    conn.execute("""
                        INSERT OR IGNORE INTO clienti_tag_assoc (cliente_id, tag_id, auto)
                        SELECT ?, tag_id, auto FROM clienti_tag_assoc WHERE cliente_id = ?
                    """, (pid, sid))
                    if sec_row["thefork_id"]:
                        conn.execute("""
                            INSERT OR IGNORE INTO clienti_alias (cliente_id, thefork_id, merged_from_id)
                            VALUES (?, ?, ?)
                        """, (pid, sec_row["thefork_id"], sid))
                    conn.execute("UPDATE clienti_alias SET cliente_id = ? WHERE cliente_id = ?", (pid, sid))
                    # Campi complementari
                    updates, values = [], []
                    for field in fill_fields:
                        pv = princ_row[field]
                        sv = sec_row[field]
                        if (not pv or str(pv).strip() == "") and sv and str(sv).strip() != "":
                            updates.append(f"{field} = ?")
                            values.append(sv)
                    if updates:
                        conn.execute(f"UPDATE clienti SET {', '.join(updates)} WHERE id = ?", values + [pid])
                    conn.execute("DELETE FROM clienti_no_duplicato WHERE cliente_a = ? OR cliente_b = ?", (sid, sid))
                    conn.execute("DELETE FROM clienti WHERE id = ?", (sid,))
                    conn.execute("UPDATE clienti SET protetto = 1 WHERE id = ?", (pid,))
                    merged_count += 1
                except Exception as ex:
                    errors.append(f"Errore {sid}->{pid}: {str(ex)}")

        conn.commit()
        return JSONResponse({
            "status": "ok",
            "merged": merged_count,
            "gruppi": len(groups),
            "errors": errors,
        })
    except Exception as e:
        logger.exception("Errore auto-merge")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


@router.get("/duplicati/suggerimenti")
def suggerisci_duplicati(
    tipo: Optional[str] = Query(None, description="telefono, email, nome — se vuoto ritorna tutti"),
    limit: int = Query(100, ge=1, le=500),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Suggerisce possibili duplicati.
    tipo: 'telefono' | 'email' | 'nome' | None (tutti)
    """
    conn = get_clienti_conn()
    try:
        duplicati = []
        seen_id_sets = set()  # evita gruppi doppi tra criteri diversi

        # Carica le coppie escluse (marcate "non è un duplicato")
        excl_rows = conn.execute(
            "SELECT cliente_a, cliente_b FROM clienti_no_duplicato"
        ).fetchall()
        excluded_pairs = set()
        for er in excl_rows:
            excluded_pairs.add(frozenset([er["cliente_a"], er["cliente_b"]]))

        def _is_excluded(ids_list):
            """Controlla se TUTTE le coppie nel gruppo sono escluse."""
            if len(ids_list) == 2:
                return frozenset(ids_list) in excluded_pairs
            # Per gruppi > 2: escludi solo se tutte le coppie possibili sono escluse
            from itertools import combinations
            return all(frozenset(pair) in excluded_pairs for pair in combinations(ids_list, 2))

        def _build_detail(ids_list):
            """Costruisce il dettaglio clienti per un gruppo di ID."""
            detail = []
            for cid in ids_list:
                c = conn.execute(
                    "SELECT id, cognome, nome, telefono, email, thefork_id FROM clienti WHERE id = ?",
                    (cid,),
                ).fetchone()
                if not c:
                    continue
                pren_count = conn.execute(
                    "SELECT COUNT(*) FROM clienti_prenotazioni WHERE cliente_id = ?",
                    (cid,),
                ).fetchone()[0]
                detail.append({**dict(c), "prenotazioni": pren_count})
            return detail

        def _add_groups(query, params, tipo_label):
            """Esegue una query di grouping e aggiunge i risultati."""
            rows = conn.execute(query, params).fetchall()
            for r in rows:
                ids = [int(x) for x in r["ids"].split(",")]
                id_key = frozenset(ids)
                if id_key in seen_id_sets:
                    continue
                if _is_excluded(ids):
                    continue
                seen_id_sets.add(id_key)
                detail = _build_detail(ids)
                if len(detail) >= 2:
                    duplicati.append({
                        "tipo": tipo_label,
                        "match": r["match_val"],
                        "clienti": detail,
                    })

        # --- TELEFONO ---
        if not tipo or tipo == "telefono":
            _add_groups("""
                SELECT telefono as match_val, GROUP_CONCAT(id) as ids, COUNT(*) as cnt
                FROM clienti
                WHERE telefono IS NOT NULL AND telefono != '' AND attivo = 1
                GROUP BY telefono
                HAVING cnt > 1
                ORDER BY cnt DESC
                LIMIT ?
            """, (limit,), "telefono")

        # --- EMAIL ---
        if not tipo or tipo == "email":
            _add_groups("""
                SELECT LOWER(email) as match_val, GROUP_CONCAT(id) as ids, COUNT(*) as cnt
                FROM clienti
                WHERE email IS NOT NULL AND email != '' AND attivo = 1
                GROUP BY LOWER(email)
                HAVING cnt > 1
                ORDER BY cnt DESC
                LIMIT ?
            """, (limit,), "email")

        # --- NOME + COGNOME ---
        if not tipo or tipo == "nome":
            _add_groups("""
                SELECT LOWER(cognome) || ' ' || LOWER(nome) as match_val,
                       GROUP_CONCAT(id) as ids, COUNT(*) as cnt
                FROM clienti
                WHERE cognome != '' AND nome != '' AND attivo = 1
                GROUP BY LOWER(cognome), LOWER(nome)
                HAVING cnt > 1
                ORDER BY cnt DESC
                LIMIT ?
            """, (limit,), "nome")

        return JSONResponse({"duplicati": duplicati, "totale": len(duplicati)})
    finally:
        conn.close()


# ============================================================
# ENDPOINT: ESCLUDI COPPIA DA DUPLICATI ("Non è un duplicato")
# ============================================================
class NoDuplicatoRequest(BaseModel):
    ids: List[int] = Field(..., description="Lista degli ID clienti da escludere come gruppo")


@router.post("/duplicati/escludi")
def escludi_duplicato(
    req: NoDuplicatoRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Segna un gruppo di clienti come 'non duplicati'.
    Alla prossima ricerca duplicati, questo gruppo non verrà più suggerito.
    """
    if len(req.ids) < 2:
        raise HTTPException(400, "Servono almeno 2 ID")

    conn = get_clienti_conn()
    try:
        from itertools import combinations
        count = 0
        for a, b in combinations(sorted(req.ids), 2):
            conn.execute(
                "INSERT OR IGNORE INTO clienti_no_duplicato (cliente_a, cliente_b) VALUES (?, ?)",
                (a, b),
            )
            count += 1
        conn.commit()
        return JSONResponse({"status": "ok", "coppie_escluse": count})
    finally:
        conn.close()


# ============================================================
# ENDPOINT: CODA REVISIONE IMPORT DIFF
# ============================================================
@router.get("/import/diff")
def lista_import_diff(
    stato: str = Query("pending"),
    limit: int = Query(200, ge=1, le=1000),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Ritorna le differenze trovate tra CRM e TheFork durante l'import.
    Raggruppate per cliente, con nome/cognome per la UI.
    """
    conn = get_clienti_conn()
    try:
        rows = conn.execute("""
            SELECT d.*, c.nome, c.cognome, c.telefono, c.email
            FROM clienti_import_diff d
            JOIN clienti c ON c.id = d.cliente_id
            WHERE d.stato = ?
            ORDER BY d.data_import DESC
            LIMIT ?
        """, (stato, limit)).fetchall()

        # Raggruppa per cliente
        grouped = {}
        for r in rows:
            cid = r["cliente_id"]
            if cid not in grouped:
                grouped[cid] = {
                    "cliente_id": cid,
                    "nome": r["nome"],
                    "cognome": r["cognome"],
                    "telefono": r["telefono"],
                    "email": r["email"],
                    "diff": [],
                }
            grouped[cid]["diff"].append({
                "id": r["id"],
                "campo": r["campo"],
                "valore_crm": r["valore_crm"],
                "valore_thefork": r["valore_thefork"],
                "data_import": r["data_import"],
            })

        result = list(grouped.values())
        totale_diff = sum(len(g["diff"]) for g in result)

        return JSONResponse({
            "clienti": result,
            "totale_clienti": len(result),
            "totale_diff": totale_diff,
        })
    finally:
        conn.close()


@router.get("/import/diff/count")
def count_import_diff(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Conteggio rapido delle differenze pending (per badge UI)."""
    conn = get_clienti_conn()
    try:
        row = conn.execute(
            "SELECT COUNT(*) as n FROM clienti_import_diff WHERE stato = 'pending'"
        ).fetchone()
        return JSONResponse({"pending": row["n"]})
    finally:
        conn.close()


class DiffActionRequest(BaseModel):
    ids: List[int] = Field(..., description="ID delle righe clienti_import_diff")
    azione: str = Field(..., description="'applica' o 'ignora'")


@router.post("/import/diff/risolvi")
def risolvi_import_diff(
    req: DiffActionRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Applica o ignora una o più differenze import.
    - applica: aggiorna il campo nel DB clienti con il valore TheFork
    - ignora: segna la differenza come risolta senza modificare nulla
    """
    if req.azione not in ("applica", "ignora"):
        raise HTTPException(400, "Azione deve essere 'applica' o 'ignora'")

    conn = get_clienti_conn()
    try:
        applicati = 0
        ignorati = 0

        for diff_id in req.ids:
            diff_row = conn.execute(
                "SELECT * FROM clienti_import_diff WHERE id = ? AND stato = 'pending'",
                (diff_id,),
            ).fetchone()
            if not diff_row:
                continue

            if req.azione == "applica":
                # Aggiorna il campo nel cliente con il valore TheFork
                campo = diff_row["campo"]
                # Validazione: solo campi conosciuti (anti-injection)
                campi_validi = {
                    "email", "telefono", "telefono2", "data_nascita",
                    "indirizzo", "cap", "citta", "paese",
                    "pref_cibo", "pref_bevande", "pref_posto",
                    "restrizioni_dietetiche", "allergie", "note_thefork",
                    "nome", "cognome", "titolo", "lingua",
                }
                if campo not in campi_validi:
                    continue
                conn.execute(
                    f"UPDATE clienti SET {campo} = ? WHERE id = ?",
                    (diff_row["valore_thefork"], diff_row["cliente_id"]),
                )
                applicati += 1
            else:
                ignorati += 1

            # Segna come risolto
            conn.execute(
                "UPDATE clienti_import_diff SET stato = ?, risolto_at = datetime('now','localtime') WHERE id = ?",
                (req.azione, diff_id),
            )

        conn.commit()
        return JSONResponse({
            "status": "ok",
            "applicati": applicati,
            "ignorati": ignorati,
        })
    except Exception as e:
        logger.exception("Errore risoluzione diff import")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


# ============================================================
# ENDPOINT: LISTA PRENOTAZIONI (globale, con filtri)
# ============================================================
@router.get("/prenotazioni/lista")
def lista_prenotazioni(
    q: Optional[str] = None,
    stato: Optional[str] = None,
    canale: Optional[str] = None,
    data_da: Optional[str] = None,
    data_a: Optional[str] = None,
    cliente_id: Optional[int] = None,
    limit: int = Query(50, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    conn = get_clienti_conn()
    try:
        where = []
        params = []

        if stato:
            where.append("p.stato = ?")
            params.append(stato)
        if canale:
            where.append("p.canale = ?")
            params.append(canale)
        if data_da:
            where.append("p.data_pasto >= ?")
            params.append(data_da)
        if data_a:
            where.append("p.data_pasto <= ?")
            params.append(data_a)
        if cliente_id:
            where.append("p.cliente_id = ?")
            params.append(cliente_id)
        if q:
            where.append(
                "(c.nome LIKE ? OR c.cognome LIKE ? OR c.nome2 LIKE ? OR c.cognome2 LIKE ? OR p.nota_ristorante LIKE ? OR p.nota_cliente LIKE ?)"
            )
            like = f"%{q}%"
            params.extend([like] * 6)

        where_sql = " AND ".join(where) if where else "1=1"

        count_row = conn.execute(
            f"""SELECT COUNT(*) as tot
                FROM clienti_prenotazioni p
                LEFT JOIN clienti c ON c.id = p.cliente_id
                WHERE {where_sql}""",
            params,
        ).fetchone()
        totale = count_row["tot"]

        rows = conn.execute(
            f"""
            SELECT p.*,
                   c.nome as cliente_nome,
                   c.cognome as cliente_cognome,
                   c.telefono as cliente_telefono,
                   c.vip as cliente_vip
            FROM clienti_prenotazioni p
            LEFT JOIN clienti c ON c.id = p.cliente_id
            WHERE {where_sql}
            ORDER BY p.data_pasto DESC, p.ora_pasto DESC
            LIMIT ? OFFSET ?
            """,
            params + [limit, offset],
        ).fetchall()

        return JSONResponse({
            "prenotazioni": [dict(r) for r in rows],
            "totale": totale,
            "limit": limit,
            "offset": offset,
        })
    except Exception as e:
        logger.exception("Errore lista prenotazioni")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


# ============================================================
# ENDPOINT: STATS PRENOTAZIONI (per dashboard)
# ============================================================
@router.get("/prenotazioni/stats")
def prenotazioni_stats(
    anno: Optional[int] = None,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    conn = get_clienti_conn()
    try:
        year_filter = ""
        params = []
        if anno:
            year_filter = "AND substr(data_pasto, 1, 4) = ?"
            params.append(str(anno))

        stats = {}

        stats["totale"] = conn.execute(
            f"SELECT COUNT(*) FROM clienti_prenotazioni WHERE 1=1 {year_filter}", params
        ).fetchone()[0]

        # Per stato
        stati = conn.execute(
            f"SELECT stato, COUNT(*) as n FROM clienti_prenotazioni WHERE 1=1 {year_filter} GROUP BY stato ORDER BY n DESC",
            params,
        ).fetchall()
        stats["per_stato"] = [dict(r) for r in stati]

        # Per canale
        canali = conn.execute(
            f"SELECT COALESCE(canale, 'n/d') as canale, COUNT(*) as n FROM clienti_prenotazioni WHERE 1=1 {year_filter} GROUP BY canale ORDER BY n DESC",
            params,
        ).fetchall()
        stats["per_canale"] = [dict(r) for r in canali]

        # Pax medio
        pax_avg = conn.execute(
            f"SELECT AVG(pax) as avg_pax, SUM(pax) as tot_pax FROM clienti_prenotazioni WHERE stato IN ('SEATED','ARRIVED','BILL','LEFT') {year_filter}",
            params,
        ).fetchone()
        stats["pax_medio"] = round(pax_avg["avg_pax"], 1) if pax_avg["avg_pax"] else 0
        stats["pax_totale"] = pax_avg["tot_pax"] or 0

        # No-show
        stats["no_show"] = conn.execute(
            f"SELECT COUNT(*) FROM clienti_prenotazioni WHERE stato = 'NO_SHOW' {year_filter}", params
        ).fetchone()[0]

        # Cancellazioni
        stats["cancellazioni"] = conn.execute(
            f"SELECT COUNT(*) FROM clienti_prenotazioni WHERE stato = 'CANCELED' {year_filter}", params
        ).fetchone()[0]

        # Per mese (ultimi 12 mesi)
        per_mese = conn.execute(
            """
            SELECT substr(data_pasto, 1, 7) as mese, COUNT(*) as n, SUM(pax) as pax
            FROM clienti_prenotazioni
            WHERE stato IN ('SEATED','ARRIVED','BILL','LEFT')
              AND data_pasto >= date('now', '-12 months')
            GROUP BY mese ORDER BY mese
            """
        ).fetchall()
        stats["per_mese"] = [dict(r) for r in per_mese]

        # Top clienti (più prenotazioni seated)
        top = conn.execute(
            f"""
            SELECT c.id, c.nome, c.cognome, c.vip,
                   COUNT(*) as n_prenotazioni,
                   SUM(p.pax) as tot_pax
            FROM clienti_prenotazioni p
            JOIN clienti c ON c.id = p.cliente_id
            WHERE p.stato IN ('SEATED','ARRIVED','BILL','LEFT') {year_filter}
            GROUP BY c.id
            ORDER BY n_prenotazioni DESC
            LIMIT 20
            """,
            params,
        ).fetchall()
        stats["top_clienti"] = [dict(r) for r in top]

        # Anni disponibili
        anni = conn.execute(
            "SELECT DISTINCT substr(data_pasto, 1, 4) as anno FROM clienti_prenotazioni ORDER BY anno DESC"
        ).fetchall()
        stats["anni_disponibili"] = [r["anno"] for r in anni]

        return JSONResponse(stats)
    except Exception as e:
        logger.exception("Errore stats prenotazioni")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


# ============================================================
# ENDPOINT: CONTEGGIO SEGMENTI MARKETING
# ============================================================
@router.get("/segmenti/conteggi")
def segmenti_conteggi(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Conteggio rapido di ogni segmento marketing per badge/riepilogo UI.
    """
    conn = get_clienti_conn()
    try:
        STATI_OK = "('SEATED','ARRIVED','BILL','LEFT')"
        totale_attivi = conn.execute("SELECT COUNT(*) FROM clienti WHERE attivo = 1").fetchone()[0]

        # Clienti con almeno 1 prenotazione completata
        con_visite = conn.execute(f"""
            SELECT p.cliente_id,
                COUNT(*) as tot,
                SUM(CASE WHEN p.data_pasto >= date('now','-12 months') THEN 1 ELSE 0 END) as anno,
                MAX(p.data_pasto) as ultima,
                MIN(p.data_pasto) as prima
            FROM clienti_prenotazioni p
            JOIN clienti c ON c.id = p.cliente_id AND c.attivo = 1
            WHERE p.stato IN {STATI_OK} AND p.cliente_id IS NOT NULL
            GROUP BY p.cliente_id
        """).fetchall()

        counts = {"abituale": 0, "occasionale": 0, "nuovo": 0, "in_calo": 0, "perso": 0, "mai_venuto": 0}
        oggi = str(date.today())
        soglia_anno = str(date.today() - timedelta(days=365))
        soglia_3mesi = str(date.today() - timedelta(days=90))

        clienti_con_visite = set()
        for r in con_visite:
            clienti_con_visite.add(r["cliente_id"])
            anno = r["anno"] or 0
            ultima = r["ultima"] or ""
            prima = r["prima"] or ""

            if ultima < soglia_anno:
                counts["perso"] += 1
            elif prima >= soglia_3mesi and anno <= 2:
                counts["nuovo"] += 1
            elif anno >= 5:
                counts["abituale"] += 1
            elif anno >= 1:
                counts["occasionale"] += 1
            else:
                counts["perso"] += 1

        counts["mai_venuto"] = totale_attivi - len(clienti_con_visite)

        # In calo: query dedicata più precisa
        in_calo_rows = conn.execute(f"""
            SELECT COUNT(DISTINCT p.cliente_id) FROM clienti_prenotazioni p
            JOIN clienti c ON c.id = p.cliente_id AND c.attivo = 1
            WHERE p.stato IN {STATI_OK}
            GROUP BY p.cliente_id
            HAVING SUM(CASE WHEN p.data_pasto BETWEEN date('now','-18 months') AND date('now','-6 months') THEN 1 ELSE 0 END) >= 3
            AND SUM(CASE WHEN p.data_pasto >= date('now','-6 months') THEN 1 ELSE 0 END) <= 1
        """).fetchall()
        counts["in_calo"] = len(in_calo_rows)

        counts["totale_attivi"] = totale_attivi
        counts["con_email"] = conn.execute(
            "SELECT COUNT(*) FROM clienti WHERE attivo = 1 AND email IS NOT NULL AND email != ''"
        ).fetchone()[0]
        counts["con_telefono"] = conn.execute(
            "SELECT COUNT(*) FROM clienti WHERE attivo = 1 AND telefono IS NOT NULL AND telefono != ''"
        ).fetchone()[0]

        return JSONResponse(counts)
    finally:
        conn.close()


# ============================================================
# ENDPOINT: LISTA CLIENTI (con ricerca e filtri)
# ============================================================
@router.get("/")
def lista_clienti(
    q: Optional[str] = None,
    vip: Optional[bool] = None,
    tag_id: Optional[int] = None,
    rank: Optional[str] = None,
    segmento: Optional[str] = None,
    attivo: Optional[bool] = True,
    compleanno_entro_giorni: Optional[int] = None,
    con_email: Optional[bool] = None,
    con_telefono: Optional[bool] = None,
    limit: int = Query(100, ge=1, le=5000),
    offset: int = Query(0, ge=0),
    ordine: str = Query("cognome_asc"),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Lista clienti con filtri marketing.
    segmento: abituale | occasionale | nuovo | in_calo | perso | mai_venuto
    """
    conn = get_clienti_conn()
    try:
        where = []
        params = []

        if attivo is not None:
            where.append("c.attivo = ?")
            params.append(1 if attivo else 0)

        if vip is not None:
            where.append("c.vip = ?")
            params.append(1 if vip else 0)

        if rank:
            where.append("c.rank = ?")
            params.append(rank)

        if q:
            where.append(
                "(c.nome LIKE ? OR c.cognome LIKE ? OR c.email LIKE ? OR c.telefono LIKE ?"
                " OR c.note_thefork LIKE ? OR c.allergie LIKE ? OR c.pref_cibo LIKE ? OR c.pref_bevande LIKE ?"
                " OR c.nome2 LIKE ? OR c.cognome2 LIKE ?)"
            )
            like = f"%{q}%"
            params.extend([like] * 10)

        if tag_id:
            where.append("c.id IN (SELECT cliente_id FROM clienti_tag_assoc WHERE tag_id = ?)")
            params.append(tag_id)

        if con_email:
            where.append("c.email IS NOT NULL AND c.email != ''")
        if con_telefono:
            where.append("c.telefono IS NOT NULL AND c.telefono != ''")

        if compleanno_entro_giorni:
            today = date.today()
            where.append("c.data_nascita IS NOT NULL")
            date_conditions = []
            for i in range(compleanno_entro_giorni + 1):
                d = today + timedelta(days=i)
                date_conditions.append(f"substr(c.data_nascita, 1, 5) = '{d.strftime('%d/%m')}'")
            if date_conditions:
                where.append(f"({' OR '.join(date_conditions)})")

        # ── Segmenti marketing (calcolati da prenotazioni completate) ──
        # visite_anno = completate nell'ultimo anno
        # visite_semestre = completate negli ultimi 6 mesi
        # visite_anno_prec = completate tra -24 e -12 mesi
        # prima_visita_recente = prima visita completata negli ultimi 90 giorni
        STATI_OK = "('SEATED','ARRIVED','BILL','LEFT')"
        if segmento:
            seg_map = {
                "abituale": f"""c.id IN (
                    SELECT p.cliente_id FROM clienti_prenotazioni p
                    WHERE p.stato IN {STATI_OK} AND p.data_pasto >= date('now','-12 months')
                    GROUP BY p.cliente_id HAVING COUNT(*) >= 5
                )""",
                "occasionale": f"""c.id IN (
                    SELECT p.cliente_id FROM clienti_prenotazioni p
                    WHERE p.stato IN {STATI_OK} AND p.data_pasto >= date('now','-12 months')
                    GROUP BY p.cliente_id HAVING COUNT(*) BETWEEN 1 AND 4
                )""",
                "nuovo": f"""c.id IN (
                    SELECT p.cliente_id FROM clienti_prenotazioni p
                    WHERE p.stato IN {STATI_OK}
                    GROUP BY p.cliente_id
                    HAVING MIN(p.data_pasto) >= date('now','-3 months')
                )""",
                "in_calo": f"""c.id IN (
                    SELECT sub.cid FROM (
                        SELECT p.cliente_id as cid,
                            SUM(CASE WHEN p.data_pasto >= date('now','-6 months') THEN 1 ELSE 0 END) as recenti,
                            SUM(CASE WHEN p.data_pasto BETWEEN date('now','-18 months') AND date('now','-6 months') THEN 1 ELSE 0 END) as precedenti
                        FROM clienti_prenotazioni p
                        WHERE p.stato IN {STATI_OK}
                        GROUP BY p.cliente_id
                        HAVING precedenti >= 3 AND recenti <= 1
                    ) sub
                )""",
                "perso": f"""c.id IN (
                    SELECT p.cliente_id FROM clienti_prenotazioni p
                    WHERE p.stato IN {STATI_OK}
                    GROUP BY p.cliente_id
                    HAVING MAX(p.data_pasto) < date('now','-12 months')
                ) AND c.id IN (
                    SELECT DISTINCT p2.cliente_id FROM clienti_prenotazioni p2
                    WHERE p2.stato IN {STATI_OK}
                )""",
                "mai_venuto": f"""c.id NOT IN (
                    SELECT DISTINCT p.cliente_id FROM clienti_prenotazioni p
                    WHERE p.stato IN {STATI_OK} AND p.cliente_id IS NOT NULL
                )""",
            }
            if segmento in seg_map:
                where.append(seg_map[segmento])

        where_sql = " AND ".join(where) if where else "1=1"

        # Ordinamento
        order_map = {
            "cognome_asc": "c.cognome ASC, c.nome ASC",
            "cognome_desc": "c.cognome DESC, c.nome DESC",
            "recente": "c.created_at DESC",
            "ultima_modifica": "c.updated_at DESC",
            "vip_first": "c.vip DESC, c.cognome ASC",
            "n_prenotazioni_desc": "n_prenotazioni DESC, c.cognome ASC",
            "n_prenotazioni_asc": "n_prenotazioni ASC, c.cognome ASC",
            "ultima_visita_desc": "ultima_visita DESC",
            "ultima_visita_asc": "ultima_visita ASC",
        }
        order_sql = order_map.get(ordine, "c.cognome ASC, c.nome ASC")

        # Count totale
        count_row = conn.execute(
            f"SELECT COUNT(*) as tot FROM clienti c WHERE {where_sql}", params
        ).fetchone()
        totale = count_row["tot"]

        # Fetch pagina con subquery per prenotazioni + segmento calcolato
        rows = conn.execute(
            f"""
            SELECT c.*,
                   GROUP_CONCAT(t.nome, ', ') as tags,
                   (SELECT COUNT(*) FROM clienti_prenotazioni p
                    WHERE p.cliente_id = c.id AND p.stato IN {STATI_OK}) as n_prenotazioni,
                   (SELECT MAX(p.data_pasto) FROM clienti_prenotazioni p
                    WHERE p.cliente_id = c.id AND p.stato IN {STATI_OK}) as ultima_visita,
                   (SELECT COUNT(*) FROM clienti_prenotazioni p
                    WHERE p.cliente_id = c.id AND p.stato IN {STATI_OK}
                    AND p.data_pasto >= date('now','-12 months')) as visite_anno,
                   (SELECT MIN(p.data_pasto) FROM clienti_prenotazioni p
                    WHERE p.cliente_id = c.id AND p.stato IN {STATI_OK}) as prima_visita
            FROM clienti c
            LEFT JOIN clienti_tag_assoc ta ON ta.cliente_id = c.id
            LEFT JOIN clienti_tag t ON t.id = ta.tag_id
            WHERE {where_sql}
            GROUP BY c.id
            ORDER BY {order_sql}
            LIMIT ? OFFSET ?
            """,
            params + [limit, offset],
        ).fetchall()

        # Calcola segmento per ogni riga
        risultati = []
        for r in rows:
            d = dict(r)
            # Determina segmento marketing
            n_pren = d.get("n_prenotazioni") or 0
            visite_anno = d.get("visite_anno") or 0
            ultima = d.get("ultima_visita")
            prima = d.get("prima_visita")

            if n_pren == 0:
                d["segmento"] = "mai_venuto"
            elif ultima and ultima < str(date.today() - timedelta(days=365)):
                d["segmento"] = "perso"
            elif prima and prima >= str(date.today() - timedelta(days=90)) and visite_anno <= 2:
                d["segmento"] = "nuovo"
            elif visite_anno >= 5:
                d["segmento"] = "abituale"
            elif visite_anno >= 1:
                d["segmento"] = "occasionale"
            else:
                d["segmento"] = "perso"
            risultati.append(d)

        return JSONResponse({
            "clienti": risultati,
            "totale": totale,
            "limit": limit,
            "offset": offset,
        })
    except Exception as e:
        logger.exception("Errore lista clienti")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


# ============================================================
# ENDPOINT: DETTAGLIO CLIENTE (con tag e note)
# ============================================================
@router.get("/{cliente_id}")
def get_cliente(
    cliente_id: int,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    conn = get_clienti_conn()
    try:
        row = conn.execute("SELECT * FROM clienti WHERE id = ?", (cliente_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Cliente non trovato")

        cliente = dict(row)

        # Tag associati
        tags = conn.execute(
            """
            SELECT t.id, t.nome, t.colore
            FROM clienti_tag t
            JOIN clienti_tag_assoc ta ON ta.tag_id = t.id
            WHERE ta.cliente_id = ?
            ORDER BY t.ordine
            """,
            (cliente_id,),
        ).fetchall()
        cliente["tags"] = [dict(t) for t in tags]

        # Note/diario
        note = conn.execute(
            "SELECT * FROM clienti_note WHERE cliente_id = ? ORDER BY data DESC, created_at DESC",
            (cliente_id,),
        ).fetchall()
        cliente["note"] = [dict(n) for n in note]

        # Prenotazioni (ultime 50)
        prenotazioni = conn.execute(
            """SELECT * FROM clienti_prenotazioni
               WHERE cliente_id = ?
               ORDER BY data_pasto DESC, ora_pasto DESC
               LIMIT 50""",
            (cliente_id,),
        ).fetchall()
        cliente["prenotazioni"] = [dict(p) for p in prenotazioni]

        # Stats prenotazioni cliente
        pren_stats = conn.execute(
            """SELECT
                 COUNT(*) as totale,
                 SUM(CASE WHEN stato IN ('SEATED','ARRIVED','BILL','LEFT') THEN 1 ELSE 0 END) as completate,
                 SUM(CASE WHEN stato = 'NO_SHOW' THEN 1 ELSE 0 END) as no_show,
                 SUM(CASE WHEN stato = 'CANCELED' THEN 1 ELSE 0 END) as cancellate,
                 ROUND(AVG(pax), 1) as pax_medio,
                 MIN(data_pasto) as prima_visita,
                 MAX(data_pasto) as ultima_visita
               FROM clienti_prenotazioni WHERE cliente_id = ?""",
            (cliente_id,),
        ).fetchone()
        cliente["prenotazioni_stats"] = dict(pren_stats) if pren_stats else {}

        return JSONResponse(cliente)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Errore dettaglio cliente")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


# ============================================================
# ENDPOINT: CREA CLIENTE
# ============================================================
@router.post("/")
def crea_cliente(
    body: ClienteCreate,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    conn = get_clienti_conn()
    try:
        cur = conn.execute(
            """
            INSERT INTO clienti (titolo, nome, cognome, email, telefono, telefono2,
                data_nascita, lingua, indirizzo, cap, citta, paese,
                vip, rank, promoter, newsletter, risk_level,
                pref_cibo, pref_bevande, pref_posto, restrizioni_dietetiche, allergie,
                note_thefork, attivo, origine)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                body.titolo, body.nome, body.cognome, body.email,
                body.telefono, body.telefono2, body.data_nascita, body.lingua,
                body.indirizzo, body.cap, body.citta, body.paese,
                1 if body.vip else 0, body.rank,
                1 if body.promoter else 0, 1 if body.newsletter else 0,
                body.risk_level,
                body.pref_cibo, body.pref_bevande, body.pref_posto,
                body.restrizioni_dietetiche, body.allergie,
                body.note_thefork, 1 if body.attivo else 0, body.origine,
            ),
        )
        conn.commit()
        return JSONResponse({"id": cur.lastrowid, "status": "ok"})
    except Exception as e:
        logger.exception("Errore creazione cliente")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


# ============================================================
# ENDPOINT: MODIFICA CLIENTE
# ============================================================
@router.put("/{cliente_id}")
def modifica_cliente(
    cliente_id: int,
    body: ClienteUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    conn = get_clienti_conn()
    try:
        existing = conn.execute("SELECT id FROM clienti WHERE id = ?", (cliente_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Cliente non trovato")

        conn.execute(
            """
            UPDATE clienti SET
                titolo=?, nome=?, cognome=?, email=?, telefono=?, telefono2=?,
                data_nascita=?, lingua=?, indirizzo=?, cap=?, citta=?, paese=?,
                vip=?, rank=?, promoter=?, newsletter=?, risk_level=?,
                pref_cibo=?, pref_bevande=?, pref_posto=?, restrizioni_dietetiche=?, allergie=?,
                note_thefork=?, attivo=?, origine=?,
                nome2=?, cognome2=?,
                protetto=1
            WHERE id=?
            """,
            (
                body.titolo, body.nome, body.cognome, body.email,
                body.telefono, body.telefono2, body.data_nascita, body.lingua,
                body.indirizzo, body.cap, body.citta, body.paese,
                1 if body.vip else 0, body.rank,
                1 if body.promoter else 0, 1 if body.newsletter else 0,
                body.risk_level,
                body.pref_cibo, body.pref_bevande, body.pref_posto,
                body.restrizioni_dietetiche, body.allergie,
                body.note_thefork, 1 if body.attivo else 0, body.origine,
                body.nome2, body.cognome2,
                cliente_id,
            ),
        )
        conn.commit()
        return JSONResponse({"status": "ok"})
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Errore modifica cliente")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


# ============================================================
# ENDPOINT: ELIMINA CLIENTE (soft delete)
# ============================================================
@router.delete("/{cliente_id}")
def elimina_cliente(
    cliente_id: int,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    conn = get_clienti_conn()
    try:
        conn.execute("UPDATE clienti SET attivo = 0 WHERE id = ?", (cliente_id,))
        conn.commit()
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.exception("Errore eliminazione cliente")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


# ============================================================
# ENDPOINT: ASSOCIA / RIMUOVI TAG A CLIENTE
# ============================================================
@router.post("/{cliente_id}/tag/{tag_id}")
def associa_tag(
    cliente_id: int,
    tag_id: int,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    conn = get_clienti_conn()
    try:
        conn.execute(
            "INSERT OR IGNORE INTO clienti_tag_assoc (cliente_id, tag_id, auto) VALUES (?,?,0)",
            (cliente_id, tag_id),
        )
        # Se il tag era automatico, convertilo in manuale (non verrà rimosso dall'import)
        conn.execute(
            "UPDATE clienti_tag_assoc SET auto = 0 WHERE cliente_id = ? AND tag_id = ?",
            (cliente_id, tag_id),
        )
        conn.commit()
        return JSONResponse({"status": "ok"})
    finally:
        conn.close()


@router.delete("/{cliente_id}/tag/{tag_id}")
def rimuovi_tag(
    cliente_id: int,
    tag_id: int,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    conn = get_clienti_conn()
    try:
        conn.execute(
            "DELETE FROM clienti_tag_assoc WHERE cliente_id = ? AND tag_id = ?",
            (cliente_id, tag_id),
        )
        conn.commit()
        return JSONResponse({"status": "ok"})
    finally:
        conn.close()


# ============================================================
# ENDPOINT: NOTE CLIENTE
# ============================================================
@router.post("/{cliente_id}/note")
def aggiungi_nota(
    cliente_id: int,
    body: NotaCreate,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    conn = get_clienti_conn()
    try:
        data = body.data or date.today().isoformat()
        autore = body.autore or current_user.get("sub", "")
        cur = conn.execute(
            "INSERT INTO clienti_note (cliente_id, tipo, testo, data, autore) VALUES (?,?,?,?,?)",
            (cliente_id, body.tipo, body.testo, data, autore),
        )
        conn.commit()
        return JSONResponse({"id": cur.lastrowid, "status": "ok"})
    except Exception as e:
        logger.exception("Errore creazione nota")
        raise HTTPException(500, str(e))
    finally:
        conn.close()


@router.delete("/{cliente_id}/note/{nota_id}")
def elimina_nota(
    cliente_id: int,
    nota_id: int,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    conn = get_clienti_conn()
    try:
        conn.execute(
            "DELETE FROM clienti_note WHERE id = ? AND cliente_id = ?",
            (nota_id, cliente_id),
        )
        conn.commit()
        return JSONResponse({"status": "ok"})
    finally:
        conn.close()


# ============================================================
# ENDPOINT: MAILCHIMP INTEGRATION
# ============================================================
@router.get("/mailchimp/status")
def mailchimp_status(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Verifica connessione Mailchimp e ritorna info account+audience."""
    try:
        from app.services.mailchimp_service import check_connection
        return JSONResponse(check_connection())
    except Exception as e:
        return JSONResponse({"connected": False, "error": str(e)})


@router.post("/mailchimp/sync")
def mailchimp_sync(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """
    Sincronizza tutti i clienti con email+newsletter=true verso Mailchimp.
    Include: merge fields custom, tags CRM, segmenti marketing.
    """
    from app.services.mailchimp_service import sync_contacts

    conn = get_clienti_conn()
    try:
        STATI_OK = "('SEATED','ARRIVED','BILL','LEFT')"

        # Fetch clienti con email e newsletter attiva
        rows = conn.execute(f"""
            SELECT c.*,
                   GROUP_CONCAT(DISTINCT t.nome) as tags_str,
                   (SELECT COUNT(*) FROM clienti_prenotazioni p
                    WHERE p.cliente_id = c.id AND p.stato IN {STATI_OK}) as n_prenotazioni,
                   (SELECT MAX(p.data_pasto) FROM clienti_prenotazioni p
                    WHERE p.cliente_id = c.id AND p.stato IN {STATI_OK}) as ultima_visita,
                   (SELECT COUNT(*) FROM clienti_prenotazioni p
                    WHERE p.cliente_id = c.id AND p.stato IN {STATI_OK}
                    AND p.data_pasto >= date('now','-12 months')) as visite_anno,
                   (SELECT MIN(p.data_pasto) FROM clienti_prenotazioni p
                    WHERE p.cliente_id = c.id AND p.stato IN {STATI_OK}) as prima_visita
            FROM clienti c
            LEFT JOIN clienti_tag_assoc ta ON ta.cliente_id = c.id
            LEFT JOIN clienti_tag t ON t.id = ta.tag_id
            WHERE c.email IS NOT NULL AND c.email != ''
              AND c.newsletter = 1 AND c.attivo = 1
            GROUP BY c.id
        """).fetchall()

        # Calcola segmento per ogni cliente
        clients_data = []
        for r in rows:
            d = dict(r)
            n_pren = d.get("n_prenotazioni") or 0
            visite_anno = d.get("visite_anno") or 0
            ultima = d.get("ultima_visita")
            prima = d.get("prima_visita")

            if n_pren == 0:
                segmento = "mai_venuto"
            elif ultima and ultima < str(date.today() - timedelta(days=365)):
                segmento = "perso"
            elif prima and prima >= str(date.today() - timedelta(days=90)) and visite_anno <= 2:
                segmento = "nuovo"
            elif visite_anno >= 5:
                segmento = "abituale"
            elif visite_anno >= 1:
                segmento = "occasionale"
            else:
                segmento = "perso"

            tags_list = [t.strip() for t in (d.get("tags_str") or "").split(",") if t.strip()]

            clients_data.append({
                "email": d["email"],
                "nome": d["nome"],
                "cognome": d["cognome"],
                "telefono": d.get("telefono"),
                "data_nascita": d.get("data_nascita"),
                "citta": d.get("citta"),
                "rank": d.get("rank"),
                "segmento": segmento,
                "allergie": d.get("allergie"),
                "pref_cibo": d.get("pref_cibo"),
                "vip": d.get("vip"),
                "tags_list": tags_list,
            })

        if not clients_data:
            return JSONResponse({
                "status": "ok",
                "message": "Nessun cliente da sincronizzare (controlla che abbiano email + newsletter attiva)",
                "synced": 0, "errors": 0, "skipped": 0, "totale_candidati": 0,
            })

        result = sync_contacts(clients_data)
        result["totale_candidati"] = len(clients_data)
        result["status"] = "ok"
        return JSONResponse(result)

    except ValueError as ve:
        # Errore configurazione (API key mancante etc.)
        return JSONResponse({"status": "error", "error": str(ve)}, status_code=400)
    except Exception as e:
        logger.exception("Errore sync Mailchimp")
        raise HTTPException(500, str(e))
    finally:
        conn.close()
