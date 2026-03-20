"""
TRGB — Finanza Router
Gestione finanziaria completa: import Excel, doppia classificazione, riconciliazione.
"""

import io
import sqlite3
from datetime import datetime, date
from typing import Optional
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, Query
from pydantic import BaseModel

from app.services.auth_service import get_current_user

router = APIRouter(prefix="/finanza", tags=["finanza"])

DB_PATH = "app/data/foodcost.db"


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ═══════════════════════════════════════════════════════════════════
# IMPORT EXCEL
# ═══════════════════════════════════════════════════════════════════

MESI_MAP = {
    "GENNAIO": "gennaio", "FEBBRAIO": "febbraio", "MARZO": "marzo",
    "APRILE": "aprile", "MAGGIO": "maggio", "GIUGNO": "giugno",
    "LUGLIO": "luglio", "AGOSTO": "agosto", "SETTEMBRE": "settembre",
    "OTTOBRE": "ottobre", "NOVEMBRE": "novembre", "DICEMBRE": "dicembre",
}


def parse_date(val):
    """Converte valore Excel in stringa data YYYY-MM-DD."""
    if val is None:
        return None
    if isinstance(val, datetime):
        # Excel date seriale con base sbagliata (1900-01-03 = 3 gen "1900")
        # Se anno < 2000, probabilmente è un seriale sbagliato — skip
        if val.year < 2000:
            return None
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        if val.year < 2000:
            return None
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            d = datetime.strptime(s, fmt)
            if d.year < 2000:
                return None
            return d.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def safe_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def safe_int(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return None


@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
):
    """Importa il foglio DATI dal file Excel movimenti."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File deve essere .xlsx o .xls")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl non installato sul server")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

    if "DATI" not in wb.sheetnames:
        raise HTTPException(400, "Foglio 'DATI' non trovato nel file Excel")

    ws = wb["DATI"]
    conn = get_db()
    cur = conn.cursor()

    imported = 0
    skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        if len(row) < 15:
            skipped += 1
            continue

        # Colonne: P(0) N(1) S(2) Data(3) Desc(4) DescExt(5)
        #          Dare(6) Avere(7) Note(8) CatD(9) E/U(10)
        #          Anno(11) Mese(12) Cat1(13) Cat2(14)
        #          FinEU(15) FinAnno(16) FinMese(17) FinDesc(18) Cat1Fin(19) Cat2Fin(20)

        data_val = parse_date(row[3])
        descrizione = safe_str(row[4])

        if not data_val or not descrizione:
            skipped += 1
            continue

        stato = safe_str(row[0]).upper()
        if stato not in ("X", "C"):
            stato = ""

        desc_estesa = safe_str(row[5]) if len(row) > 5 else ""
        dare = safe_float(row[6]) if len(row) > 6 else 0
        avere = safe_float(row[7]) if len(row) > 7 else 0
        note = safe_str(row[8]) if len(row) > 8 else ""
        cat_debito = safe_str(row[9]) if len(row) > 9 else ""

        tipo_analitico = safe_str(row[10]) if len(row) > 10 else ""
        anno_analitico = safe_int(row[11]) if len(row) > 11 else None
        mese_analitico = safe_str(row[12]) if len(row) > 12 else ""
        cat1 = safe_str(row[13]) if len(row) > 13 else ""
        cat2 = safe_str(row[14]) if len(row) > 14 else ""

        tipo_finanziario = safe_str(row[15]) if len(row) > 15 else ""
        anno_finanziario = safe_int(row[16]) if len(row) > 16 else None
        mese_finanziario = safe_str(row[17]) if len(row) > 17 else ""
        desc_finanziaria = safe_str(row[18]) if len(row) > 18 else ""
        cat1_fin = safe_str(row[19]) if len(row) > 19 else ""
        cat2_fin = safe_str(row[20]) if len(row) > 20 else ""

        cur.execute("""
            INSERT INTO finanza_movimenti (
                data, descrizione, descrizione_estesa, dare, avere, note,
                stato, cat_debito,
                tipo_analitico, anno_analitico, mese_analitico, cat1, cat2,
                tipo_finanziario, anno_finanziario, mese_finanziario,
                descrizione_finanziaria, cat1_fin, cat2_fin
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data_val, descrizione, desc_estesa, dare, avere, note,
            stato, cat_debito,
            tipo_analitico, anno_analitico, mese_analitico, cat1, cat2,
            tipo_finanziario, anno_finanziario, mese_finanziario,
            desc_finanziaria, cat1_fin, cat2_fin,
        ))
        imported += 1

    # Auto-categorizzazione con regole esistenti
    auto_cat = 0
    try:
        regole = cur.execute("SELECT * FROM finanza_regole_cat").fetchall()
        for regola in regole:
            r = dict(regola)
            pattern = r["pattern"]
            if not pattern:
                continue
            sets = []
            params = []
            for col in ["cat1", "cat2", "tipo_analitico", "cat1_fin", "cat2_fin",
                         "tipo_finanziario", "descrizione_finanziaria", "cat_debito"]:
                val = r.get(col, "")
                if val:
                    sets.append(f"{col} = ?")
                    params.append(val)
            if not sets:
                continue
            sets.append("updated_at = CURRENT_TIMESTAMP")
            result = cur.execute(
                f"UPDATE finanza_movimenti SET {', '.join(sets)} WHERE UPPER(descrizione) LIKE ? AND (cat1 = '' OR cat1 IS NULL)",
                params + [f"%{pattern}%"],
            )
            matched = result.rowcount
            if matched > 0:
                cur.execute("UPDATE finanza_regole_cat SET num_match = num_match + ? WHERE id = ?", (matched, r["id"]))
                auto_cat += matched
    except Exception:
        pass  # Se la tabella regole non esiste ancora, ignora

    # Log
    cur.execute("""
        INSERT INTO finanza_import_log (filename, righe_importate, righe_scartate)
        VALUES (?, ?, ?)
    """, (file.filename, imported, skipped))

    conn.commit()
    conn.close()

    return {
        "ok": True,
        "righe_importate": imported,
        "righe_scartate": skipped,
        "auto_categorizzati": auto_cat,
        "filename": file.filename,
    }


# ═══════════════════════════════════════════════════════════════════
# LISTA MOVIMENTI con filtri
# ═══════════════════════════════════════════════════════════════════

@router.get("/movimenti")
def get_movimenti(
    data_da: Optional[str] = None,
    data_a: Optional[str] = None,
    stato: Optional[str] = None,
    tipo: Optional[str] = None,
    cat1: Optional[str] = None,
    cat2: Optional[str] = None,
    cat_debito: Optional[str] = None,
    search: Optional[str] = None,
    vista: str = "analitico",  # "analitico" | "finanziario"
    anno: Optional[int] = None,
    mese: Optional[str] = None,
    sort_by: str = "data",
    sort_dir: str = "desc",
    limit: int = Query(50, le=500),
    offset: int = 0,
    current_user=Depends(get_current_user),
):
    conn = get_db()
    wheres = []
    params = []

    if data_da:
        wheres.append("data >= ?")
        params.append(data_da)
    if data_a:
        wheres.append("data <= ?")
        params.append(data_a)
    if stato is not None and stato != "":
        if stato == "pending":
            wheres.append("(stato IS NULL OR stato = '')")
        else:
            wheres.append("stato = ?")
            params.append(stato.upper())
    if cat1:
        if vista == "finanziario":
            wheres.append("cat1_fin = ?")
        else:
            wheres.append("cat1 = ?")
        params.append(cat1)
    if cat2:
        if vista == "finanziario":
            wheres.append("cat2_fin = ?")
        else:
            wheres.append("cat2 = ?")
        params.append(cat2)
    if tipo:
        if vista == "finanziario":
            wheres.append("tipo_finanziario = ?")
        else:
            wheres.append("tipo_analitico = ?")
        params.append(tipo.upper())
    if cat_debito:
        wheres.append("cat_debito = ?")
        params.append(cat_debito)
    if anno:
        if vista == "finanziario":
            wheres.append("anno_finanziario = ?")
        else:
            wheres.append("anno_analitico = ?")
        params.append(anno)
    if mese:
        if vista == "finanziario":
            wheres.append("LOWER(mese_finanziario) = LOWER(?)")
        else:
            wheres.append("UPPER(mese_analitico) = UPPER(?)")
        params.append(mese)
    if search:
        wheres.append("(descrizione LIKE ? OR descrizione_estesa LIKE ? OR note LIKE ?)")
        s = f"%{search}%"
        params.extend([s, s, s])

    where_sql = (" WHERE " + " AND ".join(wheres)) if wheres else ""

    total = conn.execute(f"SELECT COUNT(*) FROM finanza_movimenti{where_sql}", params).fetchone()[0]

    # Ordinamento sicuro
    ALLOWED_SORT = {
        "data", "descrizione", "dare", "avere", "stato",
        "cat1", "cat2", "cat1_fin", "cat2_fin",
        "tipo_analitico", "tipo_finanziario",
        "descrizione_finanziaria", "cat_debito",
        "anno_analitico", "anno_finanziario",
    }
    col = sort_by if sort_by in ALLOWED_SORT else "data"
    direction = "ASC" if sort_dir.lower() == "asc" else "DESC"

    rows = conn.execute(
        f"SELECT * FROM finanza_movimenti{where_sql} ORDER BY {col} {direction}, id DESC LIMIT ? OFFSET ?",
        params + [limit, offset],
    ).fetchall()

    conn.close()
    return {
        "total": total,
        "movimenti": [dict(r) for r in rows],
    }


# ═══════════════════════════════════════════════════════════════════
# DASHBOARD — Vista Analitica e Finanziaria
# ═══════════════════════════════════════════════════════════════════

@router.get("/dashboard")
def get_dashboard(
    anno: int = Query(...),
    vista: str = "analitico",  # "analitico" | "finanziario"
    current_user=Depends(get_current_user),
):
    conn = get_db()

    if vista == "finanziario":
        anno_col = "anno_finanziario"
        tipo_col = "tipo_finanziario"
        mese_col = "mese_finanziario"
        cat1_col = "cat1_fin"
        cat2_col = "cat2_fin"
    else:
        anno_col = "anno_analitico"
        tipo_col = "tipo_analitico"
        mese_col = "mese_analitico"
        cat1_col = "cat1"
        cat2_col = "cat2"

    # Totali anno
    totali = conn.execute(f"""
        SELECT
            COALESCE(SUM(CASE WHEN dare < 0 THEN dare ELSE 0 END), 0) AS totale_uscite,
            COALESCE(SUM(CASE WHEN avere > 0 THEN avere ELSE 0 END), 0) +
            COALESCE(SUM(CASE WHEN dare > 0 THEN dare ELSE 0 END), 0) AS totale_entrate,
            COUNT(*) AS num_movimenti
        FROM finanza_movimenti
        WHERE {anno_col} = ?
    """, (anno,)).fetchone()

    # Per mese
    mesi_order = [
        "GENNAIO", "FEBBRAIO", "MARZO", "APRILE", "MAGGIO", "GIUGNO",
        "LUGLIO", "AGOSTO", "SETTEMBRE", "OTTOBRE", "NOVEMBRE", "DICEMBRE",
    ]
    mesi_order_lower = [m.lower() for m in mesi_order]

    mesi_rows = conn.execute(f"""
        SELECT
            {mese_col} AS mese,
            COALESCE(SUM(CASE WHEN dare < 0 THEN dare ELSE 0 END), 0) AS uscite,
            COALESCE(SUM(CASE WHEN avere > 0 THEN avere ELSE 0 END), 0) +
            COALESCE(SUM(CASE WHEN dare > 0 THEN dare ELSE 0 END), 0) AS entrate,
            COUNT(*) AS num
        FROM finanza_movimenti
        WHERE {anno_col} = ?
        GROUP BY {mese_col}
    """, (anno,)).fetchall()

    mesi_dict = {safe_str(r["mese"]).upper(): dict(r) for r in mesi_rows}
    mesi_data = []
    for m in mesi_order:
        row = mesi_dict.get(m, mesi_dict.get(m.lower(), None))
        # Try lowercase match too
        if not row:
            for k, v in mesi_dict.items():
                if k.lower() == m.lower() or (k and k.strip().upper() == m):
                    row = v
                    break
        if row:
            mesi_data.append({
                "mese": m[:3].capitalize(),
                "mese_full": m.capitalize(),
                "uscite": abs(row["uscite"]),
                "entrate": row["entrate"],
                "saldo": row["entrate"] + row["uscite"],
                "num": row["num"],
            })
        else:
            mesi_data.append({
                "mese": m[:3].capitalize(),
                "mese_full": m.capitalize(),
                "uscite": 0, "entrate": 0, "saldo": 0, "num": 0,
            })

    # Breakdown per Cat.1
    cat_rows = conn.execute(f"""
        SELECT
            {cat1_col} AS cat1,
            COALESCE(SUM(dare), 0) AS dare_tot,
            COALESCE(SUM(avere), 0) AS avere_tot,
            COUNT(*) AS num
        FROM finanza_movimenti
        WHERE {anno_col} = ? AND {cat1_col} != ''
        GROUP BY {cat1_col}
        ORDER BY dare_tot ASC
    """, (anno,)).fetchall()

    categorie = [dict(r) for r in cat_rows]

    # Riconciliazione stats
    riconc = conn.execute("""
        SELECT
            stato,
            COUNT(*) AS num,
            COALESCE(SUM(dare), 0) AS dare_tot,
            COALESCE(SUM(avere), 0) AS avere_tot
        FROM finanza_movimenti
        WHERE anno_analitico = ? OR anno_finanziario = ?
        GROUP BY stato
    """, (anno, anno)).fetchall()

    riconciliazione = {safe_str(r["stato"]) or "pending": {
        "num": r["num"], "dare": r["dare_tot"], "avere": r["avere_tot"]
    } for r in riconc}

    conn.close()

    return {
        "anno": anno,
        "vista": vista,
        "totali": {
            "totale_uscite": totali["totale_uscite"],
            "totale_entrate": totali["totale_entrate"],
            "saldo": totali["totale_entrate"] + totali["totale_uscite"],
            "num_movimenti": totali["num_movimenti"],
        },
        "mesi": mesi_data,
        "categorie": categorie,
        "riconciliazione": riconciliazione,
    }


# ═══════════════════════════════════════════════════════════════════
# VALORI UNICI (per dropdown)
# ═══════════════════════════════════════════════════════════════════

@router.get("/valori-unici")
def get_valori_unici(current_user=Depends(get_current_user)):
    conn = get_db()
    result = {}

    for col, key in [
        ("cat1", "cat1"), ("cat2", "cat2"),
        ("cat1_fin", "cat1_fin"), ("cat2_fin", "cat2_fin"),
        ("tipo_analitico", "tipi_analitico"), ("tipo_finanziario", "tipi_finanziario"),
        ("cat_debito", "cat_debiti"),
        ("descrizione_finanziaria", "desc_finanziarie"),
    ]:
        rows = conn.execute(
            f"SELECT DISTINCT {col} FROM finanza_movimenti WHERE {col} != '' ORDER BY {col}"
        ).fetchall()
        result[key] = [r[0] for r in rows]

    # Anni disponibili
    rows = conn.execute("""
        SELECT DISTINCT anno_analitico AS anno FROM finanza_movimenti WHERE anno_analitico IS NOT NULL
        UNION
        SELECT DISTINCT anno_finanziario AS anno FROM finanza_movimenti WHERE anno_finanziario IS NOT NULL
        ORDER BY anno
    """).fetchall()
    result["anni"] = [r[0] for r in rows]

    conn.close()
    return result


# ═══════════════════════════════════════════════════════════════════
# UPDATE MOVIMENTO (modifica classificazione)
# ═══════════════════════════════════════════════════════════════════

class UpdateMovimento(BaseModel):
    stato: Optional[str] = None
    cat_debito: Optional[str] = None
    tipo_analitico: Optional[str] = None
    anno_analitico: Optional[int] = None
    mese_analitico: Optional[str] = None
    cat1: Optional[str] = None
    cat2: Optional[str] = None
    tipo_finanziario: Optional[str] = None
    anno_finanziario: Optional[int] = None
    mese_finanziario: Optional[str] = None
    descrizione_finanziaria: Optional[str] = None
    cat1_fin: Optional[str] = None
    cat2_fin: Optional[str] = None
    note: Optional[str] = None


@router.patch("/movimenti/{mov_id}")
def update_movimento(
    mov_id: int,
    req: UpdateMovimento,
    current_user=Depends(get_current_user),
):
    conn = get_db()
    fields = []
    params = []
    for field_name, value in req.dict(exclude_none=True).items():
        fields.append(f"{field_name} = ?")
        params.append(value)

    if not fields:
        conn.close()
        raise HTTPException(400, "Nessun campo da aggiornare")

    fields.append("updated_at = CURRENT_TIMESTAMP")
    params.append(mov_id)

    cur = conn.cursor()
    cur.execute(
        f"UPDATE finanza_movimenti SET {', '.join(fields)} WHERE id = ?",
        params,
    )
    conn.commit()
    if cur.rowcount == 0:
        conn.close()
        raise HTTPException(404, "Movimento non trovato")
    conn.close()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════
# IMPORT LOG
# ═══════════════════════════════════════════════════════════════════

@router.get("/import-log")
def get_import_log(current_user=Depends(get_current_user)):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM finanza_import_log ORDER BY imported_at DESC LIMIT 20"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ═══════════════════════════════════════════════════════════════════
# STATISTICHE RAPIDE
# ═══════════════════════════════════════════════════════════════════

@router.get("/stats")
def get_stats(current_user=Depends(get_current_user)):
    conn = get_db()

    total = conn.execute("SELECT COUNT(*) FROM finanza_movimenti").fetchone()[0]
    riconciliati = conn.execute(
        "SELECT COUNT(*) FROM finanza_movimenti WHERE stato IN ('X', 'C')"
    ).fetchone()[0]
    da_riconciliare = conn.execute(
        "SELECT COUNT(*) FROM finanza_movimenti WHERE stato = '' OR stato IS NULL"
    ).fetchone()[0]

    anni = conn.execute("""
        SELECT DISTINCT anno_analitico FROM finanza_movimenti
        WHERE anno_analitico IS NOT NULL ORDER BY anno_analitico
    """).fetchall()

    conn.close()
    return {
        "totale_movimenti": total,
        "riconciliati": riconciliati,
        "da_riconciliare": da_riconciliare,
        "anni": [r[0] for r in anni],
    }


# ═══════════════════════════════════════════════════════════════════
# RESET (per reimportare da zero)
# ═══════════════════════════════════════════════════════════════════

@router.delete("/reset")
def reset_data(current_user=Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(403, "Solo admin")
    conn = get_db()
    conn.execute("DELETE FROM finanza_movimenti")
    conn.commit()
    conn.close()
    return {"ok": True, "message": "Tutti i movimenti finanza eliminati"}


# ═══════════════════════════════════════════════════════════════════
# REGOLE CATEGORIZZAZIONE
# ═══════════════════════════════════════════════════════════════════

class RegolaCreate(BaseModel):
    pattern: str
    fornitore_nome: str = ""
    fornitore_piva: str = ""
    cat1: str = ""
    cat2: str = ""
    tipo_analitico: str = ""
    cat1_fin: str = ""
    cat2_fin: str = ""
    tipo_finanziario: str = ""
    descrizione_finanziaria: str = ""
    cat_debito: str = ""


@router.get("/regole")
def get_regole(current_user=Depends(get_current_user)):
    """Lista tutte le regole di categorizzazione."""
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM finanza_regole_cat ORDER BY pattern"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.post("/regole")
def create_regola(req: RegolaCreate, current_user=Depends(get_current_user)):
    """Crea una nuova regola di categorizzazione."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO finanza_regole_cat (
            pattern, fornitore_nome, fornitore_piva,
            cat1, cat2, tipo_analitico,
            cat1_fin, cat2_fin, tipo_finanziario, descrizione_finanziaria,
            cat_debito
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        req.pattern.upper(), req.fornitore_nome, req.fornitore_piva,
        req.cat1, req.cat2, req.tipo_analitico,
        req.cat1_fin, req.cat2_fin, req.tipo_finanziario, req.descrizione_finanziaria,
        req.cat_debito,
    ))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return {"ok": True, "id": new_id}


@router.put("/regole/{regola_id}")
def update_regola(regola_id: int, req: RegolaCreate, current_user=Depends(get_current_user)):
    """Aggiorna una regola esistente."""
    conn = get_db()
    conn.execute("""
        UPDATE finanza_regole_cat SET
            pattern=?, fornitore_nome=?, fornitore_piva=?,
            cat1=?, cat2=?, tipo_analitico=?,
            cat1_fin=?, cat2_fin=?, tipo_finanziario=?, descrizione_finanziaria=?,
            cat_debito=?, updated_at=CURRENT_TIMESTAMP
        WHERE id=?
    """, (
        req.pattern.upper(), req.fornitore_nome, req.fornitore_piva,
        req.cat1, req.cat2, req.tipo_analitico,
        req.cat1_fin, req.cat2_fin, req.tipo_finanziario, req.descrizione_finanziaria,
        req.cat_debito, regola_id,
    ))
    conn.commit()
    conn.close()
    return {"ok": True}


@router.delete("/regole/{regola_id}")
def delete_regola(regola_id: int, current_user=Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(403, "Solo admin")
    conn = get_db()
    conn.execute("DELETE FROM finanza_regole_cat WHERE id=?", (regola_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@router.post("/auto-categorizza")
def auto_categorizza(current_user=Depends(get_current_user)):
    """Applica tutte le regole ai movimenti non ancora categorizzati."""
    if current_user["role"] != "admin":
        raise HTTPException(403, "Solo admin")

    conn = get_db()
    regole = conn.execute("SELECT * FROM finanza_regole_cat").fetchall()

    if not regole:
        conn.close()
        return {"ok": True, "aggiornati": 0, "message": "Nessuna regola definita"}

    total_updated = 0
    for regola in regole:
        r = dict(regola)
        pattern = r["pattern"]
        if not pattern:
            continue

        # Trova movimenti che matchano e non hanno ancora cat1 assegnata
        movimenti = conn.execute("""
            SELECT id FROM finanza_movimenti
            WHERE UPPER(descrizione) LIKE ?
            AND (cat1 = '' OR cat1 IS NULL)
        """, (f"%{pattern}%",)).fetchall()

        if not movimenti:
            continue

        # Costruisci SET dinamico solo per campi non vuoti
        sets = []
        params = []
        for col in ["cat1", "cat2", "tipo_analitico", "cat1_fin", "cat2_fin",
                     "tipo_finanziario", "descrizione_finanziaria", "cat_debito"]:
            val = r.get(col, "")
            if val:
                sets.append(f"{col} = ?")
                params.append(val)

        if not sets:
            continue

        sets.append("updated_at = CURRENT_TIMESTAMP")
        ids = [m["id"] for m in movimenti]
        placeholders = ",".join("?" * len(ids))

        conn.execute(
            f"UPDATE finanza_movimenti SET {', '.join(sets)} WHERE id IN ({placeholders})",
            params + ids,
        )

        # Aggiorna contatore match
        conn.execute(
            "UPDATE finanza_regole_cat SET num_match = num_match + ? WHERE id = ?",
            (len(ids), r["id"]),
        )
        total_updated += len(ids)

    conn.commit()
    conn.close()
    return {"ok": True, "aggiornati": total_updated}


@router.get("/fornitori-acquisti")
def get_fornitori_acquisti(current_user=Depends(get_current_user)):
    """Lista fornitori dal modulo acquisti (fe_fatture) per suggerimenti."""
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT DISTINCT fornitore_nome, fornitore_piva,
                   COUNT(*) AS num_fatture,
                   SUM(totale_fattura) AS totale
            FROM fe_fatture
            WHERE fornitore_nome != ''
            GROUP BY fornitore_piva
            ORDER BY fornitore_nome
        """).fetchall()
        conn.close()
        return [dict(r) for r in rows]
    except Exception:
        conn.close()
        return []


@router.get("/movimenti-non-categorizzati")
def get_non_categorizzati(
    limit: int = Query(20, le=100),
    current_user=Depends(get_current_user),
):
    """Lista descrizioni uniche senza categoria, per facilitare la creazione regole."""
    conn = get_db()
    rows = conn.execute("""
        SELECT descrizione, COUNT(*) AS num,
               COALESCE(SUM(dare), 0) AS tot_dare,
               COALESCE(SUM(avere), 0) AS tot_avere
        FROM finanza_movimenti
        WHERE (cat1 = '' OR cat1 IS NULL)
        GROUP BY UPPER(descrizione)
        ORDER BY num DESC
        LIMIT ?
    """, (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ═══════════════════════════════════════════════════════════════════
# ALBERO CATEGORIE FINANZA (cat1 → cat2, analitiche + finanziarie)
# ═══════════════════════════════════════════════════════════════════

class FinCatIn(BaseModel):
    nome: str
    vista: str  # 'A' o 'F'

class FinSubcatIn(BaseModel):
    nome: str

class FinCatRename(BaseModel):
    nome: str

class FinSubcatMove(BaseModel):
    new_cat_id: int


@router.get("/albero-categorie", summary="Albero categorie finanza (A=analitico, F=finanziario)")
def get_albero_categorie(current_user=Depends(get_current_user)):
    conn = get_db()
    cats = conn.execute("SELECT id, nome, vista, ordine FROM finanza_cat ORDER BY vista, ordine, nome").fetchall()
    result = {"A": [], "F": []}
    for c in cats:
        cd = dict(c)
        subs = conn.execute(
            "SELECT id, nome, ordine FROM finanza_subcat WHERE cat_id = ? ORDER BY ordine, nome",
            (c["id"],)
        ).fetchall()
        cd["sottocategorie"] = [dict(s) for s in subs]
        result[c["vista"]].append(cd)
    conn.close()
    return result


@router.post("/albero-categorie", summary="Crea cat1 finanza")
def create_fin_cat(body: FinCatIn, current_user=Depends(get_current_user)):
    if body.vista not in ("A", "F"):
        raise HTTPException(400, "vista deve essere 'A' o 'F'")
    conn = get_db()
    nome = body.nome.strip().upper()
    try:
        cur = conn.execute(
            "INSERT INTO finanza_cat (nome, vista, ordine) VALUES (?, ?, "
            "(SELECT COALESCE(MAX(ordine),0)+1 FROM finanza_cat WHERE vista = ?))",
            (nome, body.vista, body.vista)
        )
        conn.commit()
        cat_id = cur.lastrowid
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(400, detail=f"Categoria '{nome}' esiste già per questa vista")
    conn.close()
    return {"id": cat_id, "nome": nome, "vista": body.vista}


@router.post("/albero-categorie/{cat_id}/sotto", summary="Crea cat2 sotto una cat1")
def create_fin_subcat(cat_id: int, body: FinSubcatIn, current_user=Depends(get_current_user)):
    conn = get_db()
    parent = conn.execute("SELECT id, nome, vista FROM finanza_cat WHERE id = ?", (cat_id,)).fetchone()
    if not parent:
        conn.close()
        raise HTTPException(404, "Categoria padre non trovata")
    nome = body.nome.strip().upper()
    try:
        cur = conn.execute(
            "INSERT INTO finanza_subcat (cat_id, nome, ordine) VALUES (?, ?, "
            "(SELECT COALESCE(MAX(ordine),0)+1 FROM finanza_subcat WHERE cat_id = ?))",
            (cat_id, nome, cat_id)
        )
        conn.commit()
        sub_id = cur.lastrowid
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(400, detail=f"Sottocategoria '{nome}' esiste già in questa categoria")
    conn.close()
    return {"id": sub_id, "nome": nome, "cat_id": cat_id}


@router.put("/albero-categorie/{cat_id}", summary="Rinomina cat1 e propaga su movimenti/regole")
def rename_fin_cat(cat_id: int, body: FinCatRename, current_user=Depends(get_current_user)):
    conn = get_db()
    old = conn.execute("SELECT nome, vista FROM finanza_cat WHERE id = ?", (cat_id,)).fetchone()
    if not old:
        conn.close()
        raise HTTPException(404, "Categoria non trovata")
    new_nome = body.nome.strip().upper()
    old_nome = old["nome"]
    vista = old["vista"]

    conn.execute("UPDATE finanza_cat SET nome = ? WHERE id = ?", (new_nome, cat_id))

    # Propaga rinomina nei movimenti e regole
    if old_nome != new_nome:
        if vista == "A":
            conn.execute("UPDATE finanza_movimenti SET cat1 = ? WHERE cat1 = ?", (new_nome, old_nome))
            conn.execute("UPDATE finanza_regole_cat SET cat1 = ? WHERE cat1 = ?", (new_nome, old_nome))
        else:
            conn.execute("UPDATE finanza_movimenti SET cat1_fin = ? WHERE cat1_fin = ?", (new_nome, old_nome))
            conn.execute("UPDATE finanza_regole_cat SET cat1_fin = ? WHERE cat1_fin = ?", (new_nome, old_nome))
    conn.commit()
    conn.close()
    return {"ok": True, "nome": new_nome}


@router.put("/albero-categorie/sotto/{sub_id}", summary="Rinomina cat2 e propaga")
def rename_fin_subcat(sub_id: int, body: FinCatRename, current_user=Depends(get_current_user)):
    conn = get_db()
    old = conn.execute(
        "SELECT s.nome, s.cat_id, c.vista, c.nome AS cat_nome "
        "FROM finanza_subcat s JOIN finanza_cat c ON s.cat_id = c.id WHERE s.id = ?",
        (sub_id,)
    ).fetchone()
    if not old:
        conn.close()
        raise HTTPException(404, "Sottocategoria non trovata")
    new_nome = body.nome.strip().upper()
    old_nome = old["nome"]
    cat_nome = old["cat_nome"]
    vista = old["vista"]

    conn.execute("UPDATE finanza_subcat SET nome = ? WHERE id = ?", (new_nome, sub_id))

    if old_nome != new_nome:
        if vista == "A":
            conn.execute(
                "UPDATE finanza_movimenti SET cat2 = ? WHERE cat1 = ? AND cat2 = ?",
                (new_nome, cat_nome, old_nome)
            )
            conn.execute(
                "UPDATE finanza_regole_cat SET cat2 = ? WHERE cat1 = ? AND cat2 = ?",
                (new_nome, cat_nome, old_nome)
            )
        else:
            conn.execute(
                "UPDATE finanza_movimenti SET cat2_fin = ? WHERE cat1_fin = ? AND cat2_fin = ?",
                (new_nome, cat_nome, old_nome)
            )
            conn.execute(
                "UPDATE finanza_regole_cat SET cat2_fin = ? WHERE cat1_fin = ? AND cat2_fin = ?",
                (new_nome, cat_nome, old_nome)
            )
    conn.commit()
    conn.close()
    return {"ok": True, "nome": new_nome}


@router.delete("/albero-categorie/{cat_id}", summary="Elimina cat1 e tutte le sue cat2")
def delete_fin_cat(cat_id: int, current_user=Depends(get_current_user)):
    conn = get_db()
    conn.execute("DELETE FROM finanza_subcat WHERE cat_id = ?", (cat_id,))
    conn.execute("DELETE FROM finanza_cat WHERE id = ?", (cat_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@router.delete("/albero-categorie/sotto/{sub_id}", summary="Elimina una cat2")
def delete_fin_subcat(sub_id: int, current_user=Depends(get_current_user)):
    conn = get_db()
    conn.execute("DELETE FROM finanza_subcat WHERE id = ?", (sub_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@router.post("/albero-categorie/sotto/{sub_id}/sposta", summary="Sposta cat2 sotto un'altra cat1")
def move_fin_subcat(sub_id: int, body: FinSubcatMove, current_user=Depends(get_current_user)):
    conn = get_db()
    old = conn.execute(
        "SELECT s.nome, s.cat_id, c.vista, c.nome AS old_cat_nome "
        "FROM finanza_subcat s JOIN finanza_cat c ON s.cat_id = c.id WHERE s.id = ?",
        (sub_id,)
    ).fetchone()
    if not old:
        conn.close()
        raise HTTPException(404, "Sottocategoria non trovata")
    new_parent = conn.execute("SELECT id, nome, vista FROM finanza_cat WHERE id = ?", (body.new_cat_id,)).fetchone()
    if not new_parent:
        conn.close()
        raise HTTPException(404, "Nuova categoria padre non trovata")
    if new_parent["vista"] != old["vista"]:
        conn.close()
        raise HTTPException(400, "Non puoi spostare tra viste diverse (A/F)")

    sub_nome = old["nome"]
    old_cat = old["old_cat_nome"]
    new_cat = new_parent["nome"]
    vista = old["vista"]

    conn.execute("UPDATE finanza_subcat SET cat_id = ? WHERE id = ?", (body.new_cat_id, sub_id))

    # Propaga nei movimenti e regole
    if vista == "A":
        conn.execute(
            "UPDATE finanza_movimenti SET cat1 = ? WHERE cat1 = ? AND cat2 = ?",
            (new_cat, old_cat, sub_nome)
        )
        conn.execute(
            "UPDATE finanza_regole_cat SET cat1 = ? WHERE cat1 = ? AND cat2 = ?",
            (new_cat, old_cat, sub_nome)
        )
    else:
        conn.execute(
            "UPDATE finanza_movimenti SET cat1_fin = ? WHERE cat1_fin = ? AND cat2_fin = ?",
            (new_cat, old_cat, sub_nome)
        )
        conn.execute(
            "UPDATE finanza_regole_cat SET cat1_fin = ? WHERE cat1_fin = ? AND cat2_fin = ?",
            (new_cat, old_cat, sub_nome)
        )
    conn.commit()
    conn.close()
    return {"ok": True}
